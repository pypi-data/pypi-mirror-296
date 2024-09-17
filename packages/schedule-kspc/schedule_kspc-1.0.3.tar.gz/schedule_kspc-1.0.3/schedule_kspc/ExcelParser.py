import re
import openpyxl
from .Store import Lesson, Parser

# Re compile
SCHEDULE_RE_COMPILE = re.compile(pattern=r"\w+\s+.\..\.", flags=re.IGNORECASE | re.UNICODE)

# Offsets / padding
SCHEDULE_OFFSETS: dict = {
    "weekdays": 2,
    "max_row": 2,
    "groups": 2,
}


def _remove_whitespace(string: str) -> str:
    """
    This function returning str if 'string' type == str, else 'string' object
    :param string:
    :return str:
    """
    if isinstance(string, str):
        return " ".join(string.split())
    return string


def _xl_cell_is_empty(cell: object) -> bool:
    """
    This function returning True if cell has needed attributes, else False
    :param cell:
    :return bool:
    """
    if not (hasattr(cell, "row") and hasattr(cell, "column") and hasattr(cell, "value")):
        return True
    return False


def _xl_cell_is_empty_value(cell_value: object) -> bool:
    """
    :param cell_value:
    :return bool:
    """
    return not cell_value or cell_value is None or cell_value == ' '


class _XlParser:
    def __init__(self, workbook: openpyxl.Workbook.active):
        self._workbook: openpyxl.Workbook.active = workbook

    @property
    def workbook(self) -> openpyxl.Workbook:
        return self._workbook

    def parser_by_sections(self, **kw) -> dict:
        # Func var
        offset: int = 0
        sections: tuple = kw.get("sections")
        output: dict = dict({offset: list()})
        _iter: iter = self._xl_iterator(**kw)

        # Settings
        min_items: int = kw.get("min_items")
        values_only: bool = kw.get("values_only")

        for cells in _iter:
            # Get first cell in cells
            _first_cell: object = cells[0]

            # Validation first cell in cells
            if _xl_cell_is_empty(_first_cell):
                continue

            # Section counter
            if _first_cell.row in sections and _first_cell.row > sections[0]:
                offset += 1
                output[offset]: list = list()

            # Filtering cells / replace empty value to None / remove empty cell
            cells_filter: tuple = tuple(
                None if _xl_cell_is_empty_value(cell_value=cell.value) else _remove_whitespace(cell.value)
                if values_only else (cell.column, cell.row, _remove_whitespace(cell.value))
                for cell in cells if not _xl_cell_is_empty(cell=cell)
            )

            # Checking minimum elements in cells_filter
            if min_items and len(cells_filter) < min_items:
                continue

            # Adding into store
            output[offset].append(cells_filter)

        return output

    def parser(self, **kw) -> list:

        # Func var
        output: list = list()
        _iter: iter = self._xl_iterator(**kw)

        # Settings
        min_items: int = kw.get("min_items")
        values_only: bool = kw.get("values_only")

        for cells in _iter:
            # Checking minimum elements in cells
            if min_items and len(cells) < min_items: continue

            for cell in cells:
                # Filtering cells
                if _xl_cell_is_empty(cell) or _xl_cell_is_empty_value(cell_value=cell.value):
                    continue

                # Adding into store
                output.append(
                    _remove_whitespace(cell.value) if values_only
                    else (cell.column, cell.row, _remove_whitespace(cell.value))
                )

        return output

    def _xl_iterator(self, **kw) -> iter:
        try:
            return self._workbook.active.iter_rows(
                min_col=kw["col"][0],
                max_col=kw["col"][1],
                min_row=kw["row"][0],
                max_row=kw["row"][1],
                values_only=False
            )
        except (ValueError, TypeError, IndexError):
            return None


class ExcelParser(_XlParser):
    def __init__(self, filename: str):
        super().__init__(
            workbook=openpyxl.load_workbook(
                filename=filename,
                read_only=True,
                data_only=True
            )
        )

    def parse_groups_by_active_sheet(self, values_only: bool = False) -> list:
        # Get cell
        group_cell = self._workbook.active["D8"]

        # Valid empty cell
        if _xl_cell_is_empty(group_cell):
            return []

        # Parse groups
        groups: list = self.parser(
            col=(group_cell.column, self._workbook.active.max_column),
            row=(group_cell.row, group_cell.row),
            min_item=1, values_only=values_only
        )

        return [] if not len(groups) else groups

    def parse_days_by_active_sheet(self, values_only: bool = False) -> list:
        # Get cell
        days_cell = self._workbook.active["B10"]

        # Valid empty cell
        if _xl_cell_is_empty(days_cell):
            return []

        # Parse: days
        days: list = self.parser(
            col=(days_cell.column, days_cell.column),
            row=(days_cell.row, self._workbook.active.max_row),
            min_item=1, values_only=values_only
        )

        return [] if not len(days) else days

    def parse_schedule_by_active_sheet(self):
        output = dict()

        # Parse schedule by groups
        for group in self.parse_groups_by_active_sheet():
            output[group]: dict = self.parser_by_sections(
                col=(
                    group[Parser.COLUMN.value],
                    group[Parser.COLUMN.value] + SCHEDULE_OFFSETS["weekdays"]
                ),
                row=(
                    group[Parser.ROW.value] + SCHEDULE_OFFSETS["groups"],
                    self._workbook.active.max_row - SCHEDULE_OFFSETS["max_row"]
                ),
                sections=tuple(
                    items[Parser.ROW.value] for items in self.parse_days_by_active_sheet()
                ),
                min_items=3,
                values_only=True
            )

            """
                This code adds lesson indexes:
                    - Before add code: [('lesson', 'teacher', cabinet), ...]
                    - After add code: {0: ('lesson', 'teacher', cabinet), 1: (....)}
            """
            for day in output[group].keys():
                output[group][day] = {i: lesson for i, lesson in enumerate(output[group][day])}

        return output

    def parse_groups(self, values_only: bool = False) -> list:
        output: list = list()
        for i, sheet in enumerate(self._workbook.sheetnames):
            self._workbook.active = i
            groups: list = self.parse_groups_by_active_sheet(values_only=values_only)
            if groups:
                output.append(groups)
        return list(set(group for groups in output for group in groups))

    def parse_days(self, values_only: bool = False) -> list:
        output: list = list()
        for i, sheet in enumerate(self._workbook.sheetnames):
            self._workbook.active = i
            days: list = self.parse_groups_by_active_sheet(values_only=values_only)
            if days:
                output.append(days)
        return output

    def parse_sheetnames(self) -> list:
        output: list = list()
        for i, sheet in enumerate(self._workbook.sheetnames):
            self._workbook.active = i

            # Days & Groups
            days: list = self.parse_days_by_active_sheet()
            groups: list = self.parse_groups_by_active_sheet()

            if not days and not groups:
                continue

            output.append(sheet)
        return output

    def parse_sheetnames_days_and_groups(self, values_only: bool = False) -> dict:
        output: dict = dict()

        for i, sheet in enumerate(self._workbook.sheetnames):
            self._workbook.active = i

            # Days & Groups
            days: list = self.parse_days_by_active_sheet(values_only=values_only)
            groups: list = self.parse_groups_by_active_sheet(values_only=values_only)

            if not days and not groups:
                continue

            output[sheet] = {'days': days, 'groups': groups}
        return output

    def parse_teachers(self) -> list:
        output: set = set()

        for i, sheet in enumerate(self._workbook.sheetnames):
            self._workbook.active = i
            schedule: dict = self.parse_schedule_by_active_sheet()

            if not schedule:
                continue

            # Groups
            for j, group in enumerate(schedule):
                # Weekdays
                for weekday in range(len(schedule[group])):
                    # Lessons
                    for lesson in range(len(schedule[group][weekday])):
                        # Teachers
                        teacher: str | None = schedule[group][weekday][lesson][Lesson.TEACHER.value]

                        if teacher is None:
                            continue

                        for item in tuple(SCHEDULE_RE_COMPILE.findall(string=teacher)):
                            output.add(item)

        return sorted(list(output))

    def parse_schedule_sheetnames(self):
        # Create output data dict
        output: dict = dict()

        for i, sheet in enumerate(self._workbook.sheetnames):
            self._workbook.active = i
            schedule = self.parse_schedule_by_active_sheet()

            if not schedule:
                continue

            output[sheet]: dict = dict()

            # Parse schedule by groups
            for group in schedule.keys():
                # Parse schedule by weekday
                for weekday in range(len(schedule[group])):
                    # Parse schedule by lesson
                    for lesson in range(len(schedule[group][weekday])):
                        teacher: str | None = schedule[group][weekday][lesson][Lesson.TEACHER.value]

                        if teacher is None:
                            continue

                        schedule[group][weekday][lesson]: tuple = (
                            schedule[group][weekday][lesson][Lesson.LESSON_NAME.value],
                            tuple(SCHEDULE_RE_COMPILE.findall(string=teacher)),
                            schedule[group][weekday][lesson][Lesson.CABINET.value]
                        )

                output[sheet][group[Parser.VALUE.value]] = schedule[group]
        return output

    def parse_schedule_teachers(self) -> dict:
        output: dict = dict()
        schedule: dict = self.parse_schedule_sheetnames()

        if not schedule:
            return output

        # Sheet
        for i, sheet in enumerate(schedule.keys()):
            # Groups
            for j, group in enumerate(schedule[sheet]):
                # Weekdays
                for weekday in range(len(schedule[sheet][group])):
                    # Lessons
                    for lesson in range(len(schedule[sheet][group][weekday])):
                        # Teachers
                        teachers: tuple | None = schedule[sheet][group][weekday][lesson][Lesson.TEACHER.value]

                        if teachers is None:
                            continue

                        for teacher in teachers:
                            if teacher not in output:
                                # Create key and write output data
                                output[teacher] = dict({
                                        k: dict({
                                            l: (None, None, None, None)
                                            for l in range(len(schedule[sheet][group][weekday]))
                                        }) for k in range(len(schedule[sheet][group]))
                                    }
                                )

                            # Update output data
                            output[teacher][weekday][lesson]: tuple = tuple(
                                (
                                    schedule[sheet][group][weekday][lesson][Lesson.LESSON_NAME.value],
                                    schedule[sheet][group][weekday][lesson][Lesson.TEACHER.value],
                                    schedule[sheet][group][weekday][lesson][Lesson.CABINET.value],
                                    group
                                )
                            )
        return output
