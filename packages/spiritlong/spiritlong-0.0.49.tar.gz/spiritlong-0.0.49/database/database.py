## 
# Copyright (c) 2023 Chongqing Spiritlong Technology Co., Ltd.
# All rights reserved.
# 
# @author	arthuryang
# @brief	基于池的数据库驱动MySQL/postgreSQL
##

import	MySQLdb	# pip install mysqlclient
from	MySQLdb		import	OperationalError 
import	threading
import	logging
import	openpyxl
import	time
import	copy
from	dbutils.pooled_db	import PooledDB
import	redis

import	utility_tool
import	excel_tool
from excel_tool	import	cell_string, cell_fill

# 装饰器：线程同步
def synchronized(func):
	func.__lock__ = threading.Lock()

	def synced_func(*args, **kwargs):
		with func.__lock__:
			return func(*args, **kwargs)
        
	return synced_func

# 基于池的数据库类
class Database(object):
	# 确保此类只有一个对象
	instance = {}
	@synchronized
	def __new__(cls, *args, **kwargs):
		key 	= args[0]["host"] + args[0]["db"]
		if key not in cls.instance.keys():
			cls.instance[key] = super().__new__(cls)
		return cls.instance[key]

	## 类初始化
	#	_database_config	数据库配置参数，这是一个字典
	#	SQL_type		默认"MySQL"
	#	use_redis		None表示不使用redis，传入一个字典{host, port, index}，都可省略以使用默认值
	#	print_execute_time	是否打印执行时间
	def __init__(self, _database_config, SQL_type="MySQL", use_redis=None, print_execute_time=False):
		self.database_config			= _database_config
		self.database_config["blocking"]	= True	# 连接池中如果没有可用连接，阻塞等待

		# 默认字段
		self.default_field_names	= ['ID', 'CREATE_TIMESTAMP', 'UPDATE_TIMESTAMP', 'DELETE_TIMESTAMP']

		# 数值类型
		self.digital_field_type		= ["INT", "FLOAT", "DOUBLE", "DECIMAL"]

		# 日期类型
		self.timestamp_field_type	= ["TIMESTAMP", "TIMESTAMP ON UPDATE CURRENT_TIMESTAMP", "DATE", "DATETIME"]

		# 确保属性不会被重复初始化
		if not hasattr(self, 'connection_pool'):
			# 数据库连接池
			if SQL_type=="MySQL":
				# MySQL
				self.driver		= MySQLdb
			
			self.connection_pool	= PooledDB(self.driver, **self.database_config)
		
		# 日志初始化
		self.logger	= logging.getLogger(__name__)
		if not self.logger.handlers:
			# 避免重复添加handler
			self.logger.setLevel(logging.DEBUG)
			console_handler	= logging.StreamHandler()
			console_handler.setLevel(logging.DEBUG)
			console_handler.setFormatter(logging.Formatter('%(asctime)s %(message)s', datefmt='%Y-%m-%d %H:%M:%S'))
			self.logger.addHandler(console_handler)
		# 用于时间测试
		self.pref_counter	= time.perf_counter()

		# 是否测量执行时间
		self.print_execute_time	= print_execute_time

		# 更新redis配置。数据库配置中'redis'保存的是一个字典{host, port, index}，指出了使用的redis主机、端口和数据库（序号）
		if use_redis is not None:
			self.redis_host		= use_redis['host'] 	if 'host'	in use_redis else 'localhost'
			self.redis_port		= use_redis['port'] 	if 'port'	in use_redis else '6379'
			self.redis_index	= use_redis['index']	if 'index'	in use_redis else '0'
			self.redis		= redis.Redis(connection_pool=redis.ConnectionPool(
							host	= self.redis_host,
							port	= self.redis_port,
							db	= self.redis_index, 
							decode_responses=True))
		else:
			self.redis	= None

	## 时间测量
	def _time_used(self):
		t	= self.pref_counter
		self.pref_counter	= time.perf_counter()
		self.logger.debug(f"time_used: {self.pref_counter-t:.3}")

	## 过滤处理
	#	对要SQL语句中的部件进行过滤和安全处理
	#	除了特殊字符，还需要考虑
	def replace(self, string):
		if not string:
			return string
		string	= str(string)
		string	= string.replace("'", "''")	# 单引号换成两个单引号
		# string	= string.replace("[", "[[]")	# [ -> [[]
		# string	= string.replace("%", "[%]")	# % -> [%]
		# string	= string.replace("_", "[_]")	# _ -> [_]
		# string	= string.replace("^", "[^]")	# ^ -> [^]
		return string
	
	# ------------------------ 数据库基本操作 ------------------------
	# DDL 对数据库和表的操作	
	# 	CREATE DATABASE/TABLE
	# 	DROP DATABASE/TABLE
	# 	USE 
	# 	DESC
	#	ALTER TABLE ... ADD/DROP/CHANGE/MODIFY
	# DQL 查询
	# DML 插入/修改/删除，需要提交
	# DCL 授权
	# 操作接口：
	#	query_single	查询一条记录
	#	query		查询所有记录
	#	------------ 以下皆考虑redis缓存 ---------------
	#	query_redis	获取指定表的数据，可以指定需要的字段和ID
	#	insert		插入记录；若ID已经存在，则更新该记录
	#	update		更新记录，必须指定ID
	#	delete		删除记录（将DELET_TIMESTAMP设为当前时间，并非删除该记录）
	#	undelete	恢复删除的记录（将DELETE_TIMESTAMP设为NULL）

	## 执行SQL语句
	# 其他函数中应调用此函数来执行SQL语句
	# 私有方法，避免从外部直接调用
	#	SQL_string	要执行的SQL语句
	#	need_result	查询语句返回查询结果
	#	return_ID	当need_result==False时，返回插入语句的ID，多条语句将会返回第一个新纪录的ID；当need_result==True时，返回受影响的行数
	def _execute(self, SQL_string, need_result=False, return_ID=False):
		result	= None
		try:
			t1		= time.perf_counter()
			connection	= self.connection_pool.connection()
			cursor 		= connection.cursor()

			# 执行SQL语句
			cursor.execute(SQL_string)

			if need_result:
				columns = [desc[0] for desc in cursor.description]
				result = [dict(zip(columns, row)) for row in cursor.fetchall()]
			
			if (not need_result):
				if return_ID:
					# 获取第一个新插入记录的ID
					result	= cursor.lastrowid
				else:
					result	= cursor.rowcount
				
			cursor.close()
				
			if 'autocommit' not in self.database_config or not self.database_config['autocommit']:
				# 提交
				connection.commit()
			
			# 计时
			if self.print_execute_time:
				t2		= time.perf_counter()
				self.logger.info(f"SQL execute: {t2-t1:.4}s")

			return result
		except OperationalError as ex:
			# 操作失败
			self.logger.error(str(ex))
			cursor.close()
			if 'autocommit' not in self.database_config or not self.database_config['autocommit']:
				# 回滚
				connection.rollback()
			
			# 计时
			if self.print_execute_time:
				t2		= time.perf_counter()
				self.logger.info(f"SQL execute: {t2-t1:.4}s")
						
			return None
		except Exception as ex:
			
			cursor.close()
			if 'autocommit' not in self.database_config or not self.database_config['autocommit']:
				# 回滚
				connection.rollback()
			#self.logger.info(SQL_string)
			self.logger.error(str(ex))
			
			# 计时
			if self.print_execute_time:
				t2		= time.perf_counter()
				self.logger.info(f"SQL execute: {t2-t1:.4}s")
				
			return result
	
	## 查询一条记录
	def query_single(self, SQL_string, *args):
		result	= self._execute(SQL_string+" LIMIT 1", need_result=True)
		return result[0] if result else None
		
	## 查询所有记录
	def query(self, SQL_string, *args):
		return self._execute(SQL_string, need_result=True)

	# ------------------------- 缓存查询和更新 -------------------
	# 使用redis来缓存。名为"DB {table}"的键对应于一个SET，ID
	# 每个记录用"DB {table} {ID}"来记录，是一个HASH，field:value

	## 查询redis缓存
	#	table		表名
	#	ID_list		ID列表，[]表示全部ID，可以是一个ID
	#	field_list	字段列表，[]表示全部字段，可以是一个字段的字符串
	#	valid 		是否查询有效数据
	def query_redis(self, table, ID_list=None, field_list=None, valid=True):
		if ID_list is None:
			ID_list = []
		if field_list is None:
			field_list = []

		t1		= time.perf_counter()
		# table总是全大写
		table	= table.upper()
		# ID要是字符串
		if isinstance(ID_list, list):
			ID_list	= [str(ID) for ID in ID_list]
		else:
			# 兼容只有一个ID的情况
			ID_list	= [str(ID_list)]
		# 标记
		all_fields	= True if not field_list else False			
		if isinstance(field_list, str):
			# 兼容只有单个字段的情况
			field_list	= [field_list]
		# ID总是被查询
		if 'ID' not in field_list:
			# 确保查询ID
			field_list.append('ID')
			
		if 'DELETE_TIMESTAMP' not in field_list:
			# 确保查询DELETE_TIMESTAMP
			field_list.append('DELETE_TIMESTAMP')
		
		result	= []
		# "DB {table}" 是一个SET，记录了该表ID
		name_table	= f"DB {table}"
		if self.redis and self.redis.exists(name_table):
			# 支持redis缓存且该表已经缓存，直接查找该表的数据
			if not ID_list:
				names	= [f"DB {table} {ID}" for ID in self.redis.sscan_iter(name_table)]			
			else:
				names	= [f"DB {table} {ID}" for ID in ID_list]
			for name in names:
				# 逐个查找指定字段
				if all_fields:
					# 全部字段
					result.append(self.redis.hgetall(name))
				else:
					# 选出来的字段
					values	= self.redis.hmget(name, field_list)
					if valid and values[-1]!='':
						continue
					result.append({field_list[i]:values[i] for i in range(len(field_list))})
			if self.print_execute_time:
				t2		= time.perf_counter()
				self.logger.info(f"redis query: {t2-t1:.4}s")	
			field_list	= []
			return result
		
		# 查询数据库
		SQL_string	= f"SELECT {','.join(field_list) if not all_fields else '*'} FROM {table}"
		conditions	= []
		if ID_list:
			# 选择部分ID
			conditions.append(f"ID IN ({','.join(ID_list)})")
		if valid:
			# 检查DELETE_TIMESTAMP
			conditions.append(f"DELETE_TIMESTAMP IS NULL")
		if conditions:
			# SQL语句加上WHERE
			SQL_string	+= f" WHERE {' AND '.join(conditions)}"
		result	= self.query(SQL_string)

		# 如果有redis，则缓存该表
		if self.redis:
			# 查询全部记录以缓存
			SQL_string	= f"SELECT * FROM {table}"
			records		= self.query(SQL_string)
			for r in records:
				# 保存记录
				name	= f"DB {table} {r['ID']}"
				for f in r:
					self.redis.hset(name, f, str(r[f]) if r[f] else '')
				# 更新表
				self.redis.sadd(f"DB {table}", r['ID'])
		# 计时
		if self.print_execute_time:
			t2		= time.perf_counter()
			self.logger.info(f"redis update: {t2-t1:.4}s")
		field_list	= []
		return result
	
	## 多表联合查询
	#	table_list	表名列表
	#	field_list	字段列表
	#	where		查询条件字符串
	#	valid_main	是否查询主表有效数据
	#	valid_other	是否查询其他表有效数据
	def query_union(self, table_list, field_list, where=None, valid_main=False, valid_other=False):
		if not isinstance(table_list, list) or len(table_list)==1:
			return []
		
		if where == "" or where == None:
			where	= "1"
		
		# 查询ID
		table_main	= table_list[0]
		dict_field	= {table_main:[]}
		query_ID 	= f"`{table_main}`.ID"
		query_join 	= f""
		query_where 	= where + f" AND `{table_main}`.`DELETE_TIMESTAMP` IS NULL" if valid_main==False else ""
		
		for t in table_list[1:]:
			query_ID	= query_ID	+ f",`{table_main}`.`{t}_ID`"
			query_join 	= query_join 	+ f" LEFT JOIN `{t}` ON `{table_main}`.`{t}_ID`=`{t}`.`ID`"
			if valid_other==False:
				query_where	= query_where + f" AND `{t}`.`DELETE_TIMESTAMP` IS NULL "
			dict_field[t]	= []

		SQL_string	= f"SELECT {query_ID} FROM `{table_main}` {query_join} WHERE {query_where}"
		list_field	= self.query(SQL_string)
		
		# 查询值
		for f in field_list:
			array_f 	= f.split(".")
			if len(array_f)>1:
				if array_f[0] not in dict_field.keys() or array_f[1]=="ID":
					continue
				dict_field[array_f[0]].append(array_f[1])
			else:
				if array_f[0]=="ID":
					continue
				dict_field[table_main].append(array_f[0])
		
		# 查询值
		for t in table_list:
			if len(dict_field[t])==0:
				continue
			if t==table_main:
				values		= None
				ID_field	= "ID"
			else:
				values	= ",".join([f"{t}_{item}" for item in dict_field[t]])
				ID_field	= f"{t}_ID"

			self.replace_by_ID(list_field, t, ID_field=ID_field, fields=",".join(dict_field[t]), values=values, query_delete=valid_other)
		
		return list_field

	## 多表联合查询
	#	table_list	表名列表
	#	field_list	字段列表
	#	where		查询条件字符串
	# 	page_current	当前页数
	# 	page_size	每页显示条目个数
	#	valid_main	是否查询主表有效数据
	#	valid_other	是否查询其他表有效数据
	def query_page(self, table_list, field_list, where=None, page_current=1, page_size=20, valid_main=False, valid_other=False):
		page 	= {
			"count"	: 0,
			"data"	: []
		}
		if not isinstance(table_list, list):
			return page
		
		if where == "" or where == None:
			where	= "1"

		if page_current<1:
			page_current	= 1
		
		# 查询ID
		table_main	= table_list[0]
		dict_field	= {table_main:[]}
		query_ID 	= f"`{table_main}`.ID"
		query_join 	= f""
		query_where 	= where + f" AND `{table_main}`.`DELETE_TIMESTAMP` IS NULL" if valid_main==False else ""
		
		for t in table_list[1:]:
			query_ID	= query_ID	+ f",`{table_main}`.`{t}_ID`"
			query_join 	= query_join 	+ f" LEFT JOIN `{t}` ON `{table_main}`.`{t}_ID`=`{t}`.`ID`"
			if valid_other==False:
				query_where	= query_where + f" AND `{t}`.`DELETE_TIMESTAMP` IS NULL "
			dict_field[t]	= []

		SQL_string	= f"SELECT {query_ID} FROM `{table_main}` {query_join} WHERE {query_where} LIMIT {(page_current-1)*page_size},{page_size}"
		list_field	= self.query(SQL_string)
		
		SQL_string	= f"SELECT COUNT(`{table_main}`.`ID`) AS COUNT FROM `{table_main}` {query_join} WHERE {query_where}"
		page["count"]	= self.query_single(SQL_string)["COUNT"]
		
		# 查询值
		for f in field_list:
			array_f 	= f.split(".")
			if len(array_f)>1:
				if array_f[0] not in dict_field.keys() or array_f[1]=="ID":
					continue
				dict_field[array_f[0]].append(array_f[1])
			else:
				if array_f[0]=="ID":
					continue
				dict_field[table_main].append(array_f[0])
		
		# 查询值
		for t in table_list:
			if len(dict_field[t])==0:
				continue
			if t==table_main:
				values		= None
				ID_field	= "ID"
			else:
				values	= ",".join([f"{t}_{item}" for item in dict_field[t]])
				ID_field	= f"{t}_ID"

			self.replace_by_ID(list_field, t, ID_field=ID_field, fields=",".join(dict_field[t]), values=values, query_delete=valid_other)
		page["data"]	= list_field
		
		return page
			
	## 插入记录，若ID已经存在，则更新该记录。本实现中，insert包含了update
	#	table	要插入的表
	#	records	记录的列表
	# 如果只有一条记录，则返回其ID；否则返回None
	def insert(self, table, records):
		if not records:
			self.logger.info("新增/更新0条记录")
			return None
		
		if isinstance(records, dict):
			records		= [records]

		if not isinstance(records, list):
			self.logger.info("记录要用列表")
			return None
		
		# 已有ID的列表
		if self.driver==MySQLdb:
			SQL_string	= f"SELECT ID FROM `{table}`"

		result		= self.query(SQL_string)
		ID_all		= [int(i['ID']) for i in result]

		# 把没有ID的选出来
		records_insert	= []
		records_update	= []
		for r in records:
			if 'ID' not in r:
				# 无ID
				records_insert.append(r)
			elif r['ID']=='NULL' or not r['ID']:
				# ID为空
				r.pop('ID')
				records_insert.append(r)
			else:
				if ID_all and int(r['ID']) in ID_all:
					# ID已经存在
					records_update.append(r)
				else:
					# ID不存在
					records_insert.append(r)
		result	= None

		# UPDATE		
		if records_update:
			self.update(table, records_update)		

		# INSERT
		if records_insert: 	
			# 字典列表合并到一个字典
			field_names	= {}
			for r in records_insert:
				field_names	|= r

			# 构建SQL语句：若该字段若没有，则代之以NULL
			field_names	= list(field_names)
			if self.driver==MySQLdb:
				SQL_string	= f"INSERT INTO `{table}` (`{'`,`'.join(field_names)}`) VALUES"
			value_strings	= []

			for r in records_insert:
				values	= []
				for f in field_names:
					if f in r:
						if r[f]==None or r[f]=='NULL':
							values.append('NULL')
						else:
							values.append(f"'{self.replace(r[f])}'")
					else:
						values.append('NULL')
				value_strings.append(f"( {','.join(values)} )")
			SQL_string	+= f" {','.join(value_strings)}"

			if self.redis and self.redis.exists(table):
				# 如果要缓存的话，在执行INSERT之前先检查最后一个ID
				result	= self.query(f"SELECT MAX(ID) FROM {table}")
				max_ID	= int(result[0]['MAX(ID)'])

			result	= self._execute(SQL_string, return_ID=True)
			if result:
				self.logger.info(f"表{table}新增{len(records_insert)}条记录，last_ID={result}")
			else:
				self.logger.info(f"表{table}新增{len(records_insert)}条记录失败")
			if self.redis and self.redis.exists(table):
				# 仅当已经缓存该表的时候才更新
				SQL_string	= f"SELECT * FROM {table} WHERE ID>{max_ID}"
				records	= self.query(SQL_string)
				for r in records:
					# 保存记录
					self.redis.hmset(f"DB {table} {r['ID']}", r)
					# 更新表
					self.redis.sadd(f"DB {table}", r['ID'])
		return result	

	## 更新记录
	# 	table	记录所在的表
	# 	records	一条或多条记录的列表，每个记录必须包含ID，更新字段可以不一致
	def update(self, table, records):
		if not records:
			self.logger.info("无更新数据")
			return False
		
		if isinstance(records, dict):
			records	= [records]
			
		# 对每个字段，生成需要更新的ID和值
		values	= {}
		for r in records:
			if 'ID' in r:
				# 必须要包含ID字段
				for f in r.keys():
					if  f not in values:
						values[f]	= [(r['ID'], r[f])]
					else:
						# 使用(ID, value)来存储
						values[f].append((r['ID'], r[f]))

		# 生成SQL语句，要注意长度不要超过SQL服务器的要求
		values_string	= []
		IDs_string	= []
		for f in values.keys():
			if self.driver==MySQLdb:
				s	= f"`{f}` = CASE `ID` "

			for p in values[f]:
				value	= f"'{self.replace(p[1])}'" if p[1]!=None and p[1]!='NULL' else 'NULL'
				s	+= f" WHEN '{p[0]}' THEN {value} "
				if str(p[0]) not in IDs_string:
					IDs_string.append(str(p[0]))
			s	+= 'END'
			values_string.append(s)
		if self.driver==MySQLdb:
			SQL_string	=  f"UPDATE `{table}` SET {','.join(values_string)} WHERE `ID` IN ({','.join(IDs_string)})"
				
		# 检查长度限制 TODO

		# 执行
		result	= self._execute(SQL_string)

		self.logger.info(f"表{table}更新{result}条记录")

		if self.redis and self.redis.exists(table):
			# 仅当已经缓存该表的时候才更新，直接按ID更新全部内容
			SQL_string	= f"SELECT * FROM {table} WHERE ID IN ({','.join(IDs_string)})"
			records	= self.query(SQL_string)
			for r in records:
				# 保存记录
				self.redis.hmset(f"DB {table} {r['ID']}", r)
				# 更新表
				self.redis.sadd(f"DB {table}", r['ID'])

		return result
		
	## 删除记录（将DELET_TIMESTAMP设为当前时间，并非删除该记录）
	def delete(self, table, ID):
		if self.driver==MySQLdb:
			SQL_string	= f"UPDATE `{table}` SET DELETE_TIMESTAMP=CURRENT_TIMESTAMP()  WHERE `ID`={ID}"
		self._execute(SQL_string)

		if self.redis and self.redis.exists(table):
			# 仅当已经缓存该表的时候才更新
			SQL_string	= f"SELECT DELETE_TIMESTAMP FROM {table} WHERE ID={ID}"
			result	= self.query_single(SQL_string)
			if result:
				self.redis.hset(f"DB {table} {ID}", 'DELETE_TIMESTAMP', result['DELETE_TIMESTAMP'])

	## 恢复删除的记录（将DELETE_TIMESTAMP设为NULL）
	def undelete(self, table, ID):
		if self.driver==MySQLdb:
			SQL_string	= f"UPDATE `{table}` SET DELETE_TIMESTAMP=NULL  WHERE `ID`={ID}"
		self._execute(SQL_string)
		
		if self.redis and self.redis.exists(table):
			# 仅当已经缓存该表的时候才更新
			self.redis.hset(f"DB {table} {ID}", 'DELETE_TIMESTAMP', 'NULL')

	# ************************************************************************
	# 				MySQL
	# ************************************************************************
	# ------------------------ 工具函数 ------------------------
	## 将列表或逗号分隔的字符串转换为SQL语句中的字段列表
	def fields_string(self, table_name, fields):
		if isinstance(fields, str):
			# 字符串转列表
			fields	= fields.split(',')

		if isinstance(fields, list):
			# 列表
			try:
				fields_string	= ",".join([f"`{table_name}`.`{f.strip()}`" for f in fields])
			except Exception:
				# 可能会失败，万一不是字符串列表呢
				return None
		else:
			# 本来就不是字符串或字段列表
			return None
		
		if '*' in fields_string:
			# 不支持通配符*
			return None

		return fields_string

	## 查询满足条件的所有记录，不支持JOIN
	#	table_name	表名
	#	where		WHERE语句
	#	fields		要查询的字段，可以是一个逗号分隔的字符串，或字段列表，无效或通配符则返回ID
	#	order		默认ID降序
	#	groups		分组列名，可以是一个逗号分隔的字符串，或字段名列表
	def all_where(self, table_name, where, fields=None, orders="", groups="", print_SQL_string=False, limit=0):
		fields	= self.fields_string(fields)
		fields	= f"`{table_name}`.`ID`" if not fields else fields
		orders	= orders if orders else f"`{table_name}`.`ID` DESC"
		groups	= f"GROUP BY {self.fields_string(groups)}" if groups else ""
		limit	= f"LIMIT {limit}" if limit>0 else ""

		SQL_string	= f'''
			SELECT {fields}
			FROM `{table_name}` 
			WHERE `{table_name}`.`DELETE_TIMESTAMP` IS NULL AND {where}
			{groups}
			ORDER BY {orders}
			{limit}
		'''
		
		if print_SQL_string:
			print(SQL_string)
		return self.query(SQL_string)
	
	## 查询满足条件的单个记录
	def one_where(self, table_name, where, fields=None, orders="", groups="", print_SQL_string=False):
		result	= self.all_where(table_name, where, fields=None, orders="", groups="", print_SQL_string=False, limit=1)
		return result[0] if result else None
	

	## 从数据库按ID查询
	#	table_name	表名
	#	ID		可以是单个ID，ID列表，字符串ID
	#	fields		要查询的字段，可以是一个逗号分隔的字符串，或字段列表，无效或通配符则返回ID
	def query_by_ID(self, table_name, ID, fields):
		fields	= self.fields_string(table_name,fields)
		fields	= f"`{table_name}`.`ID`" if not fields else fields
		if isinstance(ID,int):
			ID_string	= str(ID)
		if isinstance(ID, str):
			# 字符串转列表
			ID	= ID.split(',')
		if isinstance(ID, list):
			# 列表
			ID_string	= ",".join([f"{str(f).strip()}" for f in ID])
		
		where	= f"" 

		SQL_string	= f'''
			SELECT {fields}
			FROM `{table_name}` 
			WHERE `{table_name}`.`DELETE_TIMESTAMP` IS NULL AND
				`{table_name}`.`ID` IN ({ID_string})
		'''
		return self.query(SQL_string)

	## 根据查询结果中其他表的ID，添加为其他表的指定字段
	#	data		查询结果，可以是单个字典，或字典列表
	#	ID_field	查询结果中的ID字段，它们都是同一个表
	#	table_name	ID字段对应的表
	#	fields		需要添加的字段，可以是单个字段名或逗号分隔的字段名
	# 	values		fields对应的值的别名,可以是单个字段名或逗号分隔的字段名,为空则同fields
	# 	query_delete	是否查询已经删除的记录
	# data
	def replace_by_ID(self, data, table_name, ID_field="ID", fields="NAME", values=None, query_delete=False):
		if utility_tool.is_empty(data):
			return data

		if values is None:
			values	= fields
		fields	= fields.split(',')
		values	= values.split(',')
		
		# 统一格式
		if isinstance(data, list):
			list_data 	= data
		else:
			list_data 	= [data]
		
		# 数据查询
		fields_string	= ",".join([f"`{table_name}`.`{f.strip()}`" for f in fields])
		ID_string	= utility_tool.list_to_tuple_string(utility_tool.list_dict_to_list(list_data, ID_field))
		where_string	= f"`{table_name}`.`ID` IN {ID_string}"
		if not query_delete:
			where_string	= where_string +f''' AND `{table_name}`.`DELETE_TIMESTAMP` IS NULL'''
			
		SQL_string	= f'''
				SELECT {fields_string},`{table_name}`.`ID`
				FROM `{table_name}` 
				WHERE {where_string}
			'''
		
		# 数据转换
		dict_record	= {}
		if len(values)>len(fields):
			values	= values[:len(fields)]	
		for item in self.query(SQL_string):
			dict_record[item["ID"]]	= {}
			for index,key in enumerate(fields):
				dict_record[item["ID"]][values[index]]	= item[key]
		
		for item in list_data:
			for key in values:
				if item[ID_field] not in dict_record.keys() or utility_tool.format_int(item[ID_field])==0:
					item[key]	= ""
				else:
					item[key]	= dict_record[item[ID_field]][key]
		
		# 返回
		if isinstance(data, list):
			return list_data
			
		return list_data[0]

	# ------------------------ 输入输出和数据库管理操作 ------------------------
	## 字段COLUMN信息转换为我们自定义的字段字典
	def convert_COLUMN_field_MySQL(self, COLUMN_field):
		# 设置字段信息
		field		= self.blank_field_string_MySQL()
		field['name']	= COLUMN_field['COLUMN_NAME']

		# 保存原始的类型字符串
		field['MySQL_Type']	= COLUMN_field['COLUMN_TYPE']

		type_string	= COLUMN_field['COLUMN_TYPE'].split('(')
		field['type']	= type_string[0].upper()
		
		if field['type']=="TIMESTAMP" and "on update current_timestamp()" in COLUMN_field['EXTRA']:
			field['type']	+= " ON UPDATE CURRENT_TIMESTAMP"
			
		if len(type_string)>1:
			field['length']	= type_string[1][:-1]
		
		field['null']	= True if COLUMN_field['IS_NULLABLE']=="YES" else False
		
		if COLUMN_field['COLUMN_DEFAULT']=="current_timestamp()":
			field['default']	= "CURRENT_TIMESTAMP"
		elif COLUMN_field['COLUMN_DEFAULT']=="current_timestamp()":
			field['default']	= COLUMN_field['COLUMN_DEFAULT']
			
		if COLUMN_field['COLUMN_COMMENT']!='None':
			field['comment']	= COLUMN_field['COLUMN_COMMENT']

		if COLUMN_field['COLUMN_KEY']=='UNI':
			field['unique']	= True
		
		return field

	## 获取字段信息
	def get_field_MySQL(self, table, field_name):
		# 获取字段信息
		SQL_string	= f'''
			SELECT * 
			FROM information_schema.`COLUMNS` 
			WHERE `TABLE_SCHEMA`='{self.database_config['db']}' AND 
				`TABLE_NAME`='{table}' AND 
				`COLUMN_NAME`='{field_name}'
			'''
		r	= self.query_single(SQL_string)
		if r:
			return self.convert_COLUMN_field_MySQL(r)
		else:
			return None
		

	## 获取表的结构
	def get_table_MySQL(self, table):
		t	= self.query(f'''
			SELECT table_name, table_comment 
			FROM information_schema.TABLES 
			WHERE table_schema = '{self.database_config['db']}' AND 
				table_name='{table}'
			''')
		if len(t)<1:
			# 没找到
			return None
		
		t	= t[0]
		
		fields = self.query(f'''
			SELECT * 
			FROM information_schema.`COLUMNS` 
			WHERE `TABLE_SCHEMA`='{self.database_config['db']}' AND 
				`TABLE_NAME`='{table}'
			''')
		
		# 排序后依次加入
		t['fields']	= []
		for f in sorted(fields, key=lambda item: item['ORDINAL_POSITION']):
			t['fields'].append(self.convert_COLUMN_field_MySQL(f))
		
		return t

	## 获取数据库的结构
	def get_schema_MySQL(self):
		# 读取表
		tables	= self.query(f'''
			SELECT table_name, table_comment 
			FROM information_schema.TABLES 
			WHERE table_schema = '{self.database_config['db']}'
		''')
		
		if tables==None:
			# 操作失败
			return None

		
		fields = self.query(f'''
			SELECT * 
			FROM information_schema.`COLUMNS` 
			WHERE `TABLE_SCHEMA`='{self.database_config['db']}'
		''')
		
		for t in tables:
			# 选择该表字段
			fs	= [ f for f in fields if f['TABLE_NAME']==t['table_name'] ]
			# 排序后依次加入
			t['fields']	= []
			for f in sorted(fs, key=lambda item: item['ORDINAL_POSITION']):
				t['fields'].append(self.convert_COLUMN_field_MySQL(f))

			# 唯一约束
			if self.driver==MySQLdb:
				# 该表的键（除ID主键外）
				keys	= self.query(f"SELECT * FROM information_schema.`KEY_COLUMN_USAGE` WHERE `TABLE_SCHEMA`='{self.database_config['db']}' and `TABLE_NAME`='{t['table_name']}'")
				if len(keys)>1:
					# 注意ID总是主键，排除这种情况
					constraints	= {}
					for k in keys:
						if k['CONSTRAINT_NAME'] in constraints:
							constraints[k['CONSTRAINT_NAME']].append(k['COLUMN_NAME'])
						else:
							constraints[k['CONSTRAINT_NAME']]	= [k['COLUMN_NAME']]
					t['constraints']	= constraints
		# tables按表名排序
		tables.sort(key=lambda t:t['table_name'])
		return tables

	## 空的字段SQL字符串(MySQL)
	def blank_field_string_MySQL(self):
		return {
			'name'			: None,		# 字段名
			'type'			: None,		# 字段类型INT/DOUBLE/DECIMAL/CHAR/TEXT/DATETIME
			'length'		: None,		# 字段类型长度DECIMAL(10,2)/CHAR(255)
			'default'		: "NULL",	# 默认值
			'null'			: False,	# 是否可以为空，默认不能必填
			'unique'		: False,	# 是否唯一键
			'comment'		: None,		# 注释
		}

	## 从字段信息字典生成字段SQL字符串
	#	field	字段信息字典（包含name,type,length,default,null,comment）
	# 返回None表示构造失败
	def generate_field_string_MySQL(self, field):
		# 检查键是否存在
		if not all(k in field for k in self.blank_field_string_MySQL()):
			return None

		if not field['name']:
			return None
			
		MySQL_Type	= field['type']
		if field['length']:
			# 类型可能附带长度，要加小括号附在类型后
			MySQL_Type	+= f"({field['length']})"
		elif field['type'] in ("CHAR", "VARCHAR"):
			# CHAR/VARCHAR只使用255长度
			MySQL_Type	+= "(255)"
		s	= f"`{field['name']}` {MySQL_Type}"
		
		# 排序规则
		if field['type'] in ["CHAR", "VARCHAR", "TEXT"]:
			s	+= " CHARACTER SET utf8mb4 COLLATE utf8mb4_unicode_ci "
		
		# default中是初值
		if field['default'] and field['default']!='NULL':
			if field['type'] in self.digital_field_type:
				s	+= f" DEFAULT {field['default']} "
			elif field['type']=='TIMESTAMP':
				s	+= f" DEFAULT current_timestamp() " if field['default']=='CURRENT_TIMESTAMP' else ""
			else:
				s	+= f" DEFAULT '{self.replace(field['default'])}' "
		
		# 是否可以为空
		if field['null']:
			s	+= " NULL "
		else:
			s	+= " NOT NULL "

		# 注释
		if field['comment']:
			s	+= f" COMMENT '{self.replace(field['comment'])}'"

		return s

	## 增加表
	#	name	表名
	#	fields	字段信息，是一个list，每个成员是一个记录该字段信息的字典（包含name,type,length,default,on_update,null,primary,auto_increment,comment）
	#	comment	表注释
	def new_table_MySQL(self, name, fields, comment=None):
		fields_strings	= []
		# 增加默认的ID字段
		fields_strings.append("`ID` INT NOT NULL AUTO_INCREMENT COMMENT 'ID'")

		# 增加各字段
		for i in range(len(fields)):
			s	= self.generate_field_string_MySQL(fields[i])
			if s:
				fields_strings.append(s)

		# 增加默认的时间戳字段
		fields_strings.append("`CREATE_TIMESTAMP` TIMESTAMP NOT NULL DEFAULT CURRENT_TIMESTAMP COMMENT '创建时间戳'")
		fields_strings.append("`UPDATE_TIMESTAMP` TIMESTAMP ON UPDATE CURRENT_TIMESTAMP NOT NULL DEFAULT CURRENT_TIMESTAMP COMMENT '更新时间戳'")
		fields_strings.append("`DELETE_TIMESTAMP` TIMESTAMP NULL DEFAULT NULL COMMENT '删除时间戳，NULL表示未删除'")
	
		fields_strings.append('PRIMARY KEY(`ID`)')
		SQL_string	= f"CREATE TABLE `{self.database_config['db']}`.`{name}` (	\
				{','.join(fields_strings)}					\
				)  CHARACTER SET utf8mb4 COLLATE utf8mb4_unicode_ci ENGINE=InnoDB" 
		
		# 表注释
		if comment:
			SQL_string	+=  f" COMMENT='{self.replace(comment)}'"
	
		self._execute(SQL_string)
	
	## 修改/增加字段
	#	field	要修改/增加的字段属性，只包含要修改的属性，name必须要有
	def update_field_MySQL(self, table, field, first=False):
		# field中必须要有name
		if 'name' not in field:
			return None
		
		# 读取原有字段信息
		field_old	= self.get_field_MySQL(table, field['name'])
		
		if field_old:
			# 该字段已经存在	
			SQL_string	= f"ALTER TABLE `{table}` CHANGE `{field['name']}` "
		else:	
			# 需要添加字段
			SQL_string	= f"ALTER TABLE `{table}` ADD "
			field_old	= self.blank_field_string_MySQL()
			field_old['name']	= field['name']
			field_old['MySQL_Type']	= field['MySQL_Type']

		if 'type' in field:
			field_old['type']	= field['type']
			if 'length' in field:
				field_old['length']	= field['length']
			else:
				field_old['length']	= None
			# 实际上是使用这个来决定数据库字段类型的
			field_old['MySQL_Type']		= field['MySQL_Type']
		
		if 'default' in field:
			field_old['default']	= field['default']
		
		if 'null' in field:
			field_old['null']	= field['null']

		if 'comment' in field:
			field_old['comment']	= field['comment']
		
		SQL_string	+= self.generate_field_string_MySQL(field_old)

		if first:
			# 排到首位去
			SQL_string	+= " FIRST "

		self._execute(SQL_string)

		# 唯一性要单独处理
		if 'unique' in field:
			if field_old['unique']!= field['unique']:
				if field['unique']:	
					# 增加唯一性
					self._execute(f"ALTER TABLE `{table}` ADD UNIQUE (`{field['name']}`)")
				else:	
					# 去掉唯一性
					self._execute(f"ALTER TABLE `{table}` DROP INDEX `{field['name']}`")

		# json类型修改排序规则
		if field['type'] in ["LONGTEXT", "JSON"]:
			self._execute(f"ALTER TABLE `{table}` CHANGE `{field['name']}` `{field['name']}` LONGTEXT CHARACTER SET utf8mb4 COLLATE utf8mb4_unicode_ci NULL DEFAULT NULL COMMENT '{field['comment']}'")

	## 列出/删除此数据库中所有的位于DELETE_TIMESTAMP之后的字段
	# 在导入excel文件后，将会把excel内没有的字段移动到DELETE_TIMESTAMP之后，调用此函数检查无误后，再调用此函数删除	
	def obsolete_fields_MySQL(self, delete=False):
		tables	= self.get_schema_MySQL()
		
		result	= []
		for t in tables:
			field_names	= []
			for f in t['fields']:
				field_names.append(f['name'])
			i	= field_names.index('DELETE_TIMESTAMP')
			for k in range(i+1, len(field_names)):
				result.append(f"{t['table_name']:20}{t['fields'][k]['name']:20}{t['fields'][k]['comment']}")
		return result
	
	# ************************************************************************
	# 				EXCEL 输入输出
	# ************************************************************************
	## 从excel导入结构
	def import_excel_schema(self, filename):
		# 读取文件
		try:
			book	= openpyxl.load_workbook(filename)
			sheet	= book['数据库表结构']			
		except Exception as ex:
			self.logger.error(str(ex))
			return None

		tables	= []
		for row in range(2, sheet.max_row+1):
			if not cell_string(sheet, row, 1):
				# 第一列为空：表名和注释
				tables.append({
					'table_name'	: cell_string(sheet, row, 2),
					'table_comment'	: cell_string(sheet, row, 3),
					'constraints'	: cell_string(sheet, row, 4),
					'fields'	: [],
				})
			else:
				type, length	= self.excel_type_string[cell_string(sheet, row, 4)]
				default	= cell_string(sheet, row, 5) 
				if not default:
					default	= "NULL"
				
				if self.driver==MySQLdb:
					MySQL_Type	= type
					if length:
						# 类型可能附带长度，要加小括号附在类型后
						MySQL_Type	+= f"({length})"
					elif type in ("CHAR", "VARCHAR"):
						# CHAR/VARCHAR只使用255长度
						MySQL_Type	+= "(255)"
				else:
					MySQL_Type	= ""

				tables[-1]['fields'].append({
					'name'		: cell_string(sheet, row, 2),		# 字段名
					'type'		: type,					# 字段类型INT/DOUBLE/DECIMAL/CHAR/TEXT/DATETIME
					'length'	: length,				# 字段类型长度DECIMAL(10,2)/CHAR(255)
					'default'	: default,				# 默认值
					'null'		: cell_string(sheet, row, 6)=='' ,	# 是否可以为空（只要有任何值都表示不能为空）
					'unique'	: False,				# 是否唯一键
					'comment'	: cell_string(sheet, row, 3),		# 注释
					'MySQL_Type'	: MySQL_Type,				# MySQL字段类型字符串
				})
		
		if self.driver==MySQLdb:
			current_tables	= self.get_schema_MySQL()
		
		if current_tables==None:
			# 读取失败
			self.logger.error("读取表结构失败")
			return None
		
		# 当前数据库表名列表
		current_table_names	= []
		for t in current_tables:
			current_table_names.append(t['table_name'])
		
		# Excel中的表名列表
		table_names		= []
		for t in tables:
			table_names.append(t['table_name'])

		# 删除excel中不存在的表
		for t in current_table_names:
			if t not in table_names:
				self._execute(f"DROP TABLE {t}")
		
		# 按excel逐个表处理
		for t in tables:
			try:
				i	= current_table_names.index(t['table_name'])
			except Exception:
				i	= -1

			if i==-1:
				# 表还不存在，创建表
				if self.driver==MySQLdb:
					self.new_table_MySQL(t['table_name'], t['fields'], t['table_comment'])
				
				self.logger.info(f"创建了表{t['table_name']}({t['table_comment']})")
			else:
				# 表已经存在，需要更新字段：只增加不删除
				# Excel中没有字段将在现有字段的备注上标注待删除并移动到末尾
				# 	注意current_tables和current_table_names是顺序一致的
				self.logger.info(f"更新表{t['table_name']}")
				
				# 表中必须有四个默认字段：ID, CREATE_TIMESTAMP, UPDATE_TIMESTAMP, DELETE_TIMESTAMP
				for field_name in ['ID', 'CREATE_TIMESTAMP', 'UPDATE_TIMESTAMP', 'DELETE_TIMESTAMP']:
					if self.driver==MySQLdb:
						result	= self.get_field_MySQL(t['table_name'], field_name)
						if not result:
							# 该字段不存在，需要创建
							if field_name=='ID':
								# ID必须存在
								pass
							elif field_name=='CREATE_TIMESTAMP':
								# 增加默认的时间戳字段CREATE_TIMESTAMP
								self._execute(f"ALTER TABLE `{t['table_name']}` ADD`CREATE_TIMESTAMP` TIMESTAMP NOT NULL DEFAULT CURRENT_TIMESTAMP COMMENT '创建时间戳'")
							elif field_name=='UPDATE_TIMESTAMP':
								# 增加默认的时间戳字段UPDATE_TIMESTAMP
								self._execute(f"ALTER TABLE `{t['table_name']}` ADD`UPDATE_TIMESTAMP` TIMESTAMP ON UPDATE CURRENT_TIMESTAMP NOT NULL DEFAULT CURRENT_TIMESTAMP COMMENT '更新时间戳'")
							elif field_name=='DELETE_TIMESTAMP':
								# 增加默认的时间戳字段DELETE_TIMESTAMP
								self._execute(f"ALTER TABLE `{t['table_name']}` ADD`DELETE_TIMESTAMP` TIMESTAMP NULL DEFAULT NULL COMMENT '删除时间戳，NULL表示未删除'")							
									
				# 重新排序：把ID+t['fields']+CREATE_TIMESTAMP+UPDATE_TIMESTAMP+DELETE_TIMESTAMP倒序放到开头
				# 排序过程中会使用修改的字段定义，增加excel中新的字段，在excel中没有的字段将会被移动到末尾
				# 如果表中没有这默认三字段，将会导致数据库操作返回错误
				action	= 'MODIFY' 
				self._execute(f"ALTER TABLE `{t['table_name']}` {action} `DELETE_TIMESTAMP` TIMESTAMP NULL DEFAULT NULL COMMENT '删除时间戳，NULL表示未删除' FIRST")
				self._execute(f"ALTER TABLE `{t['table_name']}` {action} `UPDATE_TIMESTAMP` TIMESTAMP ON UPDATE CURRENT_TIMESTAMP NOT NULL DEFAULT CURRENT_TIMESTAMP COMMENT '更新时间戳' FIRST")
				self._execute(f"ALTER TABLE `{t['table_name']}` {action} `CREATE_TIMESTAMP` TIMESTAMP NOT NULL DEFAULT CURRENT_TIMESTAMP COMMENT '创建时间戳' FIRST")
				for i in range(len(t['fields'])-1, -1, -1):
					f	= t['fields'][i]
					if self.driver==MySQLdb:
						self.update_field_MySQL(t['table_name'], f, first=True)
				self._execute(f"ALTER TABLE `{t['table_name']}` MODIFY `ID` INT(11) AUTO_INCREMENT FIRST")

				# 查询约束
				constraints	= self.query(f"SELECT * FROM information_schema.`TABLE_CONSTRAINTS` \
					WHERE `TABLE_SCHEMA`='{self.database_config['db']}' \
						AND `TABLE_NAME`='{t['table_name']}'\
						AND `CONSTRAINT_TYPE`='UNIQUE'	")
				# 去掉约束
				for c in constraints:
					self._execute(f"ALTER TABLE `{t['table_name']}` DROP INDEX `{c['CONSTRAINT_NAME']}`")
				# 更新表的注释
				self._execute(f"ALTER TABLE `{t['table_name']}`  COMMENT '{t['table_comment']}'")
			# 更新约束
			if t['constraints']!='':
				# 分号划分
				constraints	= t['constraints'].split(';')
				for c in constraints:
					name, columns	= c.split(':')
					columns	= columns.split(',')
					self._execute(f"ALTER TABLE `{t['table_name']}` ADD UNIQUE `{name}`(`{'`,`'.join(columns)}`)")
				
	## 导出结构到excel
	def export_excel_schema(self, filename=None):
		if not filename:
			# 使用默认文件名
			filename	= f"数据库{self.database_config['db']}设计.xlsx"

		book	= openpyxl.Workbook()
		
		# SHEET 数据库表
		sheet		= book.active
		sheet.title	= '数据库表结构'
			
		# 标题
		row	= 1
		cell_fill(sheet, row, 1, '序号',	self.cell_style_title)	
		cell_fill(sheet, row, 2, '字段名',	self.cell_style_title)	
		cell_fill(sheet, row, 3, '说明',	self.cell_style_title)	
		cell_fill(sheet, row, 4, '类型',	self.cell_style_title)	
		cell_fill(sheet, row, 5, '默认值',	self.cell_style_title)	
		cell_fill(sheet, row, 6, '必填',	self.cell_style_title)	
		cell_fill(sheet, row, 7, '备注',	self.cell_style_title)	

		# 获取表结构
		if self.driver==MySQLdb:
			tables	= self.get_schema_MySQL()
		row	= 2
		# 设置格式验证，产生下拉菜单
		validation	= openpyxl.worksheet.datavalidation.DataValidation(
			type		= 'list',
			formula1	='"'+','.join(self.excel_type_string.keys())+'"',
			allow_blank	= False,
		)
				
		if tables:
			# 逐个表生成
			for t in tables:
				# 表名和表注释
				cell_fill(sheet, row, 2, t['table_name'], self.cell_style_table_name)
				cell_fill(sheet, row, 3, t['table_comment'], self.cell_style_table_comment)
				# 约束
				if 'constraints' in t:
					constraints	= []
					for c in t['constraints']:
						if c != 'PRIMARY':
							constraints.append(f"{c}:{','.join(t['constraints'][c])}")

					cell_fill(sheet, row, 4, ';'.join(constraints), self.cell_style_table_comment)
				row	+= 1

				# 各字段
				for i in range(len(t['fields'])):
					f	= t['fields'][i]
					if f['name'] not in self.default_field_names:
						cell_fill(sheet, row, 1, i, 		self.cell_style_ID)
						cell_fill(sheet, row, 2, f['name'],	self.cell_style_default)
						cell_fill(sheet, row, 3, f['comment'],	self.cell_style_default)
						# 需要将MySQL中的字段名转换为Excel中的统一名称
						if self.driver==MySQLdb:
							cell_fill(sheet, row, 4, self.MySQL_type_string[f['type']],	self.cell_style_default)
						
						# 此单元格需要验证
						validation.add(f'D{row}')

						if f['default']!="NULL":
							cell_fill(sheet, row, 5, f['default'],	self.cell_style_default)
						if f['null']==False:
							cell_fill(sheet, row, 6, '●', self.cell_style_ID)
						row	+= 1
			# 添加验证
			sheet.add_data_validation(validation)
		else:
			self.logger.info("没有数据库表")
			return

		# 自动调整列宽
		excel_tool.adjust_all_column_width(sheet)
		
		# # SHEET 数据库表
		sheet	= book.create_sheet('数据库表', 1)
		cell_fill(sheet, 1, 1, '表名', self.cell_style_title)
		cell_fill(sheet, 1, 2, '说明', self.cell_style_title)
		for i in range(len(tables)):
			# 填写excel自动计算，但这会让速度变慢
			#self.excel_cell_fill(sheet, i+2, 1, f'=INDEX(数据库表结构!$B:$B,SMALL(IF(数据库表结构!$A:$A="",ROW(数据库表结构!$A:$A),65536),ROW(A{i+2})))')
			#self.excel_cell_fill(sheet, i+2, 2, f'=INDEX(数据库表结构!$C:$C,SMALL(IF(数据库表结构!$A:$A="",ROW(数据库表结构!$A:$A),65536),ROW(A{i+2})))')
			# 直接填写表名
			cell_fill(sheet, i+2, 1, tables[i]['table_name'], 	self.cell_style_table_comment)
			cell_fill(sheet, i+2, 2, tables[i]['table_comment'],	self.cell_style_default)
		# # 自动调整列宽
		excel_tool.adjust_all_column_width(sheet)

		# SHEET 字段类型
		sheet	= book.create_sheet('字段类型', 1)
		cell_fill(sheet, 1, 1, '字段类型', 	self.cell_style_title)
		cell_fill(sheet, 1, 2, 'MySQL类型', 	self.cell_style_title)
		cell_fill(sheet, 1, 3, '备注', 		self.cell_style_title)
		row	= 2
		for k in self.excel_type_string:
			cell_fill(sheet, row, 1, k,					self.cell_style_default)
			cell_fill(sheet, row, 2, self.excel_type_string[k][0],		self.cell_style_default)
			cell_fill(sheet, row, 3, str(self.excel_type_string[k][1]),	self.cell_style_default)
			row	+= 1
	
		# # 自动调整列宽
		excel_tool.adjust_all_column_width(sheet)

		# 保存
		book.save(filename)

	## 导出数据到excel
	# 每个SHEET是一个表，SHEET名称就是表名称，第一行总是字段名称
	#	table_names	表名列表
	def export_excel_data(self, table_names=None, filename=None):
		if not filename:
			# 使用默认文件名
			filename	= f"数据库{self.database_config['db']}数据 {time.strftime('%Y-%m-%d %H_%M_%S')}.xlsx"

		book	= openpyxl.Workbook()
		# 新建有一个sheet，删除之
		book.remove(book.active)

		if table_names:
			tables	= []
			for t in table_names:
				if self.driver==MySQLdb:
					tt	= self.get_table_MySQL(t)
				if tt:
					tables.append(tt)
		else:
			# 没有指定表名就导出全部表
			if self.driver==MySQLdb:
				tables	= self.get_schema_MySQL()
		for t in tables:
			sheet		= book.create_sheet(t['table_name'])
			field_names	= []
			i		= 1
			for f in t['fields']:
				# 标题行：字段名
				cell_fill(sheet, 1, i, f['name'], self.cell_style_title)
				comment				= openpyxl.comments.Comment(f['comment'], author="EXON")
				comment.width			= 200
				comment.hight			= 25				
				sheet.cell(1, i).comment	= comment
				i				+= 1
				field_names.append(f['name'])
			# 查询数据
			row	= 2

			if self.driver==MySQLdb:
				result	= self.query(f"SELECT * FROM `{t['table_name']}`")

			for r in result:
				for i in range(len(field_names)):
					value		= r[field_names[i]]
					type		= t['fields'][i]['type']
					number_format	= None
					if value!=None and type in self.digital_field_type[1:]:
						# 数值类型
						if type=="INT":
							value	= int(r[field_names[i]])
						else:
							value	= float(r[field_names[i]])
					elif type in self.timestamp_field_type:
						# 时间类
						number_format	= "yyyy-mm-dd hh:mm:ss"

					if value != None and value != 'NULL':
						cell_fill(sheet, row, i+1, value, { 'number_format' : number_format} )
					
				row	+= 1
			# 自动调整列宽
			excel_tool.adjust_all_column_width(sheet)
		# 保存
		book.save(filename)	

	## 从excel导入数据
	# 	filename	excel文件名
	#	overwrite	是否覆盖已有数据，默认False，将excel中的数据添加到末尾，若ID已经存在则更新数据；True则先清除该表
	def import_excel_data(self, filename, overwrite=False):
		self._time_used()
		# 读取文件
		try:
			book	= openpyxl.load_workbook(filename)
		except Exception as ex:
			self.logger.error(str(ex))
			return None
		
		self._time_used()
		for sheet_name in book.sheetnames:
			sheet	= book[sheet_name]
			# SHEET名称必须是表名称
			table	= sheet.title

			if overwrite:
				# 清除表中数据
				self._execute(f"TRUNCATE {table}")

			# 首行必须是字段名称
			field_names	= []
			for i in range(sheet.max_column):
				field_names.append(str(sheet.cell(1, i+1).value))
			# 插入，若ID存在则将更新该记录
			records	= []
			for r in range(2, sheet.max_row+1):
				record	= {}
				for k in range(len(field_names)):
					value	= sheet.cell(r, k+1).value
					record[field_names[k]]	= value if value!=None else "NULL"
				records.append(record)
			self.logger.info(f"导入数据{len(records)}条到表{table}")
			self._time_used()
			self.insert(table, records)
			self._time_used()

	## 配置和定义
	
	# excel中的字段名称，值是MySQL中对应的类型和长度
	excel_type_string	= {
		"STRING"	: ["CHAR", 	"255"],
		"INT"		: ["INT",	"11"],
		"JSON"		: ["JSON", 	None],		# 在 MariaDB 的实现中， JSON 类型为 longtext 类型的别名
		"DATETIME"	: ["DATETIME", 	None],
		"FLOAT"		: ["DECIMAL", 	"10,5"],	# FLOAT实际上是10.5的DECIMAL，而不是浮点数
		"DECIMAL"	: ["DECIMAL", 	"10,2"],
		"DOUBLE"	: ["DOUBLE", 	None],
		"BLOB"		: ["LONGBLOB", 	None],		# BLOB都是LONGBLOB
		"TEXT"		: ["TEXT", 	None],		# 文本只有小于255和TEXT两种
		"TIMESTAMP"	: ["TIMESTAMP", None],
	}

	# MySQL中读取的字段名称，值是excel中的对应名称
	MySQL_type_string	= {
		"CHAR"					: "STRING",
		"DATE"					: "DATETIME",
		"DATETIME"				: "DATETIME",
		"DOUBLE"				: "DOUBLE",
		"FLOAT"					: "DECIMAL",	# FLOAT实际上是10.5的DECIMAL，而不是浮点数
		"DECIMAL"				: "DECIMAL",
		"INT"					: "INT",
		"BLOB"					: "BLOB",	# BLOB都是LONGBLOB
		"LONGBLOB"				: "BLOB",	# BLOB都是LONGBLOB
		"LONGTEXT"				: "JSON",	# 在 MariaDB 的实现中， JSON 类型为 longtext 类型的别名
		"MEDIUMBLOB"				: "BLOB",
		"TEXT"					: "TEXT",
		"TIMESTAMP"				: "TIMESTAMP",
		"TIMESTAMP ON UPDATE CURRENT_TIMESTAMP"	: "TIMESTAMP",
		"VARCHAR"				: "STRING",
		"JSON"					: "JSON",
	}

	# 默认单元格样式
	cell_style_default	= {
		'font'		: excel_tool.font('Calibri'),
		'fill'		: None,
		'border'	: None, 
		'alignment'	: excel_tool.alignment_left, 
		'number_format'	: None,
	}

	# 标题的单元格样式
	cell_style_title	= {
		'font'		: excel_tool.font(name='Calibri', color='FFFFFFFF', bold=True),
		'fill'		: excel_tool.fill_with_color('FF4169E1'),
		'border'	: None, 
		'alignment'	: excel_tool.alignment_center, 
		'number_format'	: None,
	}

	# 表名的单元格样式
	cell_style_table_name	= {
		'font'		: excel_tool.font(name='Calibri', color='FFFFFFFF', bold=True),
		'fill'		: excel_tool.fill_with_color('FFFC760F'),
		'border'	: None, 
		'alignment'	: excel_tool.alignment_left, 
		'number_format'	: None,
	}

	# 表注释单元格样式
	cell_style_table_comment	= {
		'font'		: excel_tool.font(name='Calibri', color='FFFC760F'),
		'fill'		: None,
		'border'	: None, 
		'alignment'	: excel_tool.alignment_left, 
		'number_format'	: None,
	}

	# ID单元格样式
	cell_style_ID		= {
		'font'		: excel_tool.font(name='Calibri', color='FF4169E1'),
		'fill'		: None,
		'border'	: None, 
		'alignment'	: excel_tool.alignment_center, 
		'number_format'	: None,
	}
	

# 调试/测试代码
if __name__ == '__main__':
	pass
