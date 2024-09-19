import requests
import math
import os
from datetime import datetime, timedelta
import threading
from typing import Optional
from getpass import getpass
import time
import json
from datetime import datetime
import os
from pathlib import Path
import pandas as pd
from openpyxl import load_workbook

def get_desktop_path():
    """
    Retrieve the path to the user's desktop directory.

    This function uses the `Path.home()` method to get the user's home directory
    and appends 'Desktop' to this path to construct the full path to the desktop.

    Returns:
        Path: A `Path` object representing the path to the user's desktop.
    """
    # Get the user's home directory
    home = Path.home()
    # Return the desktop path
    return home / 'Desktop'

def append_json_with_timestamp(new_data, file_name='data_log.json'):
    """
    Append new data with a timestamp to a JSON file on the desktop.

    This function adds a timestamp to each entry in `new_data` and appends it
    to a JSON file named `file_name` located on the user's desktop. If the file
    or the desktop directory does not exist, they are created.

    Args:
        new_data (dict or list of dicts): The new data to append. Must be either
            a dictionary or a list of dictionaries.
        file_name (str, optional): The name of the JSON file on the desktop.
            Defaults to 'data_log.json'.

    Raises:
        ValueError: If `new_data` is neither a dictionary nor a list of dictionaries.

    Example:
        append_json_with_timestamp({'event': 'start'}, 'log.json')
        append_json_with_timestamp([{'event': 'start'}, {'event': 'stop'}], 'log.json')
    """
    # Get desktop path and define full file path
    desktop_path = get_desktop_path()
    file_path = desktop_path / file_name

    # Create the desktop directory if it doesn't exist
    if not desktop_path.exists():
        desktop_path.mkdir(parents=True)

    # Read existing data from the JSON file
    try:
        with open(file_path, 'r') as file:
            existing_data = json.load(file)
    except FileNotFoundError:
        existing_data = []

    # Add timestamp to dictionary or list elements
    if isinstance(new_data, dict):
        # Append a timestamp to the dictionary and add it to existing data
        new_data['timestamp'] = datetime.now().isoformat()
        existing_data.append(new_data)
    elif isinstance(new_data, list):
        # Append a timestamp to each dictionary in the list and extend existing data
        for item in new_data:
            if isinstance(item, dict):
                item['timestamp'] = datetime.now().isoformat()
        existing_data.extend(new_data)
    else:
        # Print an error if new_data is neither a dictionary nor a list of dictionaries
        print("Error: new_data must be a dictionary or list of dictionaries")
        return

    # Write the updated data back to the JSON file
    try:
        with open(file_path, 'w') as file:
            json.dump(existing_data, file, indent=4)
    except Exception as e:
        print(f"Error writing to file: {e}")
        return

    print(f"Data with timestamp has been appended to {file_path}")
class InvalidCredentialsError(Exception):
    """Exception raised for invalid credentials."""
    def __init__(self, message="Invalid credentials provided"):
        self.message = message
        super().__init__(self.message)
class PermissionDeniedError(Exception):
    """Exception raised for permission denied errors."""
    def __init__(self, message="Permission denied"):
        self.message = message
        super().__init__(self.message)
class InvalidInputError(Exception):
    pass
class StoreTurnOffRestrictedError(Exception):
    """Exception raised for Store Unable to be updated."""
    def __init__(self, message="Unable to update store"):
        self.message = message
        super().__init__(self.message)
from typing import List, Optional
from typing import Optional, Dict
class Billing_Address:
    pass
class Items:
    def __init__(
        self,
        sku: str,  # Required field
        additional_data: Optional[str] = None,
        amount_refunded: Optional[float] = None,
        applied_rule_ids: Optional[str] = None,
        base_amount_refunded: Optional[float] = None,
        base_cost: Optional[float] = None,
        base_discount_amount: Optional[float] = None,
        base_discount_invoiced: Optional[float] = None,
        base_discount_refunded: Optional[float] = None,
        base_discount_tax_compensation_amount: Optional[float] = None,
        base_discount_tax_compensation_invoiced: Optional[float] = None,
        base_discount_tax_compensation_refunded: Optional[float] = None,
        base_original_price: Optional[float] = None,
        base_price: Optional[float] = None,
        base_price_incl_tax: Optional[float] = None,
        base_row_invoiced: Optional[float] = None,
        base_row_total: Optional[float] = None,
        base_row_total_incl_tax: Optional[float] = None,
        base_tax_amount: Optional[float] = None,
        base_tax_before_discount: Optional[float] = None,
        base_tax_invoiced: Optional[float] = None,
        base_tax_refunded: Optional[float] = None,
        base_weee_tax_applied_amount: Optional[float] = None,
        base_weee_tax_applied_row_amnt: Optional[float] = None,
        base_weee_tax_disposition: Optional[float] = None,
        base_weee_tax_row_disposition: Optional[float] = None,
        created_at: Optional[str] = None,
        description: Optional[str] = None,
        discount_amount: Optional[float] = None,
        discount_invoiced: Optional[float] = None,
        discount_percent: Optional[float] = None,
        discount_refunded: Optional[float] = None,
        discount_tax_compensation_amount: Optional[float] = None,
        discount_tax_compensation_canceled: Optional[float] = None,
        discount_tax_compensation_invoiced: Optional[float] = None,
        discount_tax_compensation_refunded: Optional[float] = None,
        event_id: Optional[int] = None,
        ext_order_item_id: Optional[str] = None,
        extension_attributes: Optional[Dict] = None,
        free_shipping: Optional[int] = None,
        gw_base_price: Optional[float] = None,
        gw_base_price_invoiced: Optional[float] = None,
        gw_base_price_refunded: Optional[float] = None,
        gw_base_tax_amount: Optional[float] = None,
        gw_base_tax_amount_invoiced: Optional[float] = None,
        gw_base_tax_amount_refunded: Optional[float] = None,
        gw_id: Optional[int] = None,
        gw_price: Optional[float] = None,
        gw_price_invoiced: Optional[float] = None,
        gw_price_refunded: Optional[float] = None,
        gw_tax_amount: Optional[float] = None,
        gw_tax_amount_invoiced: Optional[float] = None,
        gw_tax_amount_refunded: Optional[float] = None,
        is_qty_decimal: Optional[int] = None,
        is_virtual: Optional[int] = None,
        item_id: Optional[int] = None,
        locked_do_invoice: Optional[int] = None,
        locked_do_ship: Optional[int] = None,
        name: Optional[str] = None,
        no_discount: Optional[int] = None,
        order_id: Optional[int] = None,
        original_price: Optional[float] = None,
        parent_item: Optional[Dict] = None,  # Recursive type
        parent_item_id: Optional[int] = None,
        price: Optional[float] = None,
        price_incl_tax: Optional[float] = None,
        product_id: Optional[int] = None,
        product_option: Optional[Dict] = None,  # Assuming catalog-data-product-option-interface is a dict here
        product_type: Optional[str] = None,
        qty_backordered: Optional[float] = None,
        qty_canceled: Optional[float] = None,
        qty_invoiced: Optional[float] = None,
        qty_ordered: Optional[float] = None,
        qty_refunded: Optional[float] = None,
        qty_returned: Optional[float] = None,
        qty_shipped: Optional[float] = None,
        quote_item_id: Optional[int] = None,
        row_invoiced: Optional[float] = None,
        row_total: Optional[float] = None,
        row_total_incl_tax: Optional[float] = None,
        row_weight: Optional[float] = None,
        store_id: Optional[int] = None,
        tax_amount: Optional[float] = None,
        tax_before_discount: Optional[float] = None,
        tax_canceled: Optional[float] = None,
        tax_invoiced: Optional[float] = None,
        tax_percent: Optional[float] = None,
        tax_refunded: Optional[float] = None,
        updated_at: Optional[str] = None,
        weee_tax_applied: Optional[str] = None,
        weee_tax_applied_amount: Optional[float] = None,
        weee_tax_applied_row_amount: Optional[float] = None,
        weee_tax_disposition: Optional[float] = None,
        weee_tax_row_disposition: Optional[float] = None,
        weight: Optional[float] = None
    ):
        self.sku = sku
        self.additional_data = additional_data
        self.amount_refunded = amount_refunded
        self.applied_rule_ids = applied_rule_ids
        self.base_amount_refunded = base_amount_refunded
        self.base_cost = base_cost
        # ... continue with other fields

    def to_json(self):
        return [{key: value for key, value in self.__dict__.items() if value is not None}]

class Entity:
    """
    Represents an order interface in a sales data system, capturing comprehensive details about an order,
    including customer information, order items, pricing, and shipping details.

    Attributes:
        adjustment_negative (float): Negative adjustment amount applied to the order.
        adjustment_positive (float): Positive adjustment amount applied to the order.
        applied_rule_ids (str): Comma-separated list of applied rule IDs.
        base_adjustment_negative (float): Negative adjustment amount in the base currency.
        base_adjustment_positive (float): Positive adjustment amount in the base currency.
        base_currency_code (str): Currency code of the base currency.
        base_discount_amount (float): Discount amount in the base currency.
        base_discount_canceled (float): Discount amount that was canceled in the base currency.
        base_discount_invoiced (float): Discount amount that was invoiced in the base currency.
        base_discount_refunded (float): Discount amount that was refunded in the base currency.
        base_discount_tax_compensation_amount (float): Tax compensation amount for the discount in the base currency.
        base_discount_tax_compensation_invoiced (float): Tax compensation amount for the discount that was invoiced in the base currency.
        base_discount_tax_compensation_refunded (float): Tax compensation amount for the discount that was refunded in the base currency.
        base_grand_total (float): Grand total of the order in the base currency. (Required)
        base_shipping_amount (float): Shipping amount in the base currency.
        base_shipping_canceled (float): Shipping amount that was canceled in the base currency.
        base_shipping_discount_amount (float): Discount on the shipping amount in the base currency.
        base_shipping_discount_tax_compensation_amnt (float): Tax compensation amount for the shipping discount in the base currency.
        base_shipping_incl_tax (float): Shipping amount including tax in the base currency.
        base_shipping_invoiced (float): Shipping amount that was invoiced in the base currency.
        base_shipping_refunded (float): Shipping amount that was refunded in the base currency.
        base_shipping_tax_amount (float): Tax amount on shipping in the base currency.
        base_shipping_tax_refunded (float): Tax amount on shipping that was refunded in the base currency.
        base_subtotal (float): Subtotal of the order in the base currency.
        base_subtotal_canceled (float): Subtotal that was canceled in the base currency.
        base_subtotal_incl_tax (float): Subtotal including tax in the base currency.
        base_subtotal_invoiced (float): Subtotal that was invoiced in the base currency.
        base_subtotal_refunded (float): Subtotal that was refunded in the base currency.
        base_tax_amount (float): Tax amount in the base currency.
        base_tax_canceled (float): Tax amount that was canceled in the base currency.
        base_tax_invoiced (float): Tax amount that was invoiced in the base currency.
        base_tax_refunded (float): Tax amount that was refunded in the base currency.
        base_to_global_rate (float): Exchange rate from base to global currency.
        base_to_order_rate (float): Exchange rate from base to order currency.
        base_total_canceled (float): Total amount that was canceled in the base currency.
        base_total_due (float): Total amount due in the base currency.
        base_total_invoiced (float): Total amount invoiced in the base currency.
        base_total_invoiced_cost (float): Total cost that was invoiced in the base currency.
        base_total_offline_refunded (float): Total amount refunded offline in the base currency.
        base_total_online_refunded (float): Total amount refunded online in the base currency.
        base_total_paid (float): Total amount paid in the base currency.
        base_total_qty_ordered (float): Total quantity of items ordered in the base currency.
        base_total_refunded (float): Total amount refunded in the base currency.
        billing_address (dict): Billing address information. (Assumed to be a dictionary representing the sales-data-order-address-interface)
        billing_address_id (int): ID of the billing address.
        can_ship_partially (int): Flag indicating whether the order can be partially shipped.
        can_ship_partially_item (int): Flag indicating whether individual items can be partially shipped.
        coupon_code (str): Coupon code applied to the order.
        created_at (str): Timestamp when the order was created.
        customer_dob (str): Customer's date of birth.
        customer_email (str): Customer's email address. (Required)
        customer_firstname (str): Customer's first name.
        customer_gender (int): Customer's gender.
        customer_group_id (int): ID of the customer group.
        customer_id (int): ID of the customer.
        customer_is_guest (int): Flag indicating whether the customer is a guest.
        customer_lastname (str): Customer's last name.
        customer_middlename (str): Customer's middle name.
        customer_note (str): Note added by the customer.
        customer_note_notify (int): Flag indicating whether to notify the customer about their note.
        customer_prefix (str): Customer's prefix (e.g., Mr., Ms.).
        customer_suffix (str): Customer's suffix (e.g., Jr., Sr.).
        customer_taxvat (str): Customer's tax/VAT number.
        discount_amount (float): Discount amount applied to the order.
        discount_canceled (float): Discount amount that was canceled.
        discount_description (str): Description of the discount applied.
        discount_invoiced (float): Discount amount that was invoiced.
        discount_refunded (float): Discount amount that was refunded.
        discount_tax_compensation_amount (float): Tax compensation amount for the discount.
        discount_tax_compensation_invoiced (float): Tax compensation amount for the discount that was invoiced.
        discount_tax_compensation_refunded (float): Tax compensation amount for the discount that was refunded.
        edit_increment (int): Edit increment value.
        email_sent (int): Flag indicating whether the order email was sent.
        entity_id (int): ID of the order entity.
        ext_customer_id (str): External customer ID.
        ext_order_id (str): External order ID.
        extension_attributes (dict): Additional extension attributes. (Assumed to be a dictionary representing the sales-data-order-extension-interface)
        forced_shipment_with_invoice (int): Flag indicating whether the shipment is forced with an invoice.
        global_currency_code (str): Currency code for the global currency.
        grand_total (float): Grand total of the order. (Required)
        hold_before_state (str): State of the order before it was put on hold.
        hold_before_status (str): Status of the order before it was put on hold.
        increment_id (str): Increment ID of the order.
        is_virtual (int): Flag indicating whether the order is virtual.
        items (List[dict]): List of items in the order. (Assumed to be a list of dictionaries representing sales-data-order-item-interface) (Required)
        order_currency_code (str): Currency code for the order.
        original_increment_id (str): Original increment ID of the order.
        payment (dict): Payment information. (Assumed to be a dictionary representing sales-data-order-payment-interface)
        payment_auth_expiration (int): Payment authorization expiration.
        payment_authorization_amount (float): Payment authorization amount.
        protect_code (str): Protect code for the order.
        quote_address_id (int): Quote address ID.
        quote_id (int): Quote ID.
        relation_child_id (str): Relation child ID.
        relation_child_real_id (str): Real ID of the relation child.
        relation_parent_id (str): Relation parent ID.
        relation_parent_real_id (str): Real ID of the relation parent.
        remote_ip (str): Remote IP address of the customer.
        shipping_amount (float): Shipping amount for the order.
        shipping_canceled (float): Shipping amount that was canceled.
        shipping_description (str): Description of the shipping method.
        shipping_discount_amount (float): Discount amount on the shipping.
        shipping_discount_tax_compensation_amount (float): Tax compensation amount for the shipping discount.
        shipping_incl_tax (float): Shipping amount including tax.
        shipping_invoiced (float): Shipping amount that was invoiced.
        shipping_refunded (float): Shipping amount that was refunded.
        shipping_tax_amount (float): Tax amount on the shipping.
        shipping_tax_refunded (float): Tax amount on the shipping that was refunded.
        state (str): State of the order.
        status (str): Status of the order.
        status_histories (List[dict]): List of status histories for the order. (Assumed to be a list of dictionaries representing sales-data-order-status-history-interface)
        store_currency_code (str): Currency code for the store.
        store_id (int): ID of the store.
        store_name (str): Name of the store.
        store_to_base_rate (float): Exchange rate from store to base currency.
        store_to_order_rate (float): Exchange rate from store to order currency.
        subtotal (float): Subtotal of the order.
        subtotal_canceled (float): Subtotal amount that was canceled.
        subtotal_incl_tax (float): Subtotal including tax.
        subtotal_invoiced (float): Subtotal amount that was invoiced.
        subtotal_refunded (float): Subtotal amount that was refunded.
        tax_amount (float): Tax amount for the order.
        tax_canceled (float): Tax amount that was canceled.
        tax_invoiced (float): Tax amount that was invoiced.
        tax_refunded (float): Tax amount that was refunded.
        total_canceled (float): Total amount that was canceled.
        total_due (float): Total amount due for the order.
        total_invoiced (float): Total amount that was invoiced.
        total_item_count (int): Total number of items in the order.
        total_offline_refunded (float): Total amount refunded offline.
        total_online_refunded (float): Total amount refunded online.
        total_paid (float): Total amount paid for the order.
        total_qty_ordered (float): Total quantity of items ordered.
        total_refunded (float): Total amount refunded.
        updated_at (str): Timestamp when the order was last updated.
        weight (float): Weight of the order.
        x_forwarded_for (str): X-Forwarded-For header value.

    Methods:
        __init__: Initializes a new instance of the SalesDataOrderInterface class.
    """
    def __init__(
        self,
        base_grand_total: float,
        customer_email: str,
        grand_total: float,
        items: Items,  # Assuming sales-data-order-item-interface is a dict here
        adjustment_negative: Optional[float] = None,
        adjustment_positive: Optional[float] = None,
        applied_rule_ids: Optional[str] = None,
        base_adjustment_negative: Optional[float] = None,
        base_adjustment_positive: Optional[float] = None,
        base_currency_code: Optional[str] = None,
        base_discount_amount: Optional[float] = None,
        base_discount_canceled: Optional[float] = None,
        base_discount_invoiced: Optional[float] = None,
        base_discount_refunded: Optional[float] = None,
        base_discount_tax_compensation_amount: Optional[float] = None,
        base_discount_tax_compensation_invoiced: Optional[float] = None,
        base_discount_tax_compensation_refunded: Optional[float] = None,
        base_shipping_amount: Optional[float] = None,
        base_shipping_canceled: Optional[float] = None,
        base_shipping_discount_amount: Optional[float] = None,
        base_shipping_discount_tax_compensation_amnt: Optional[float] = None,
        base_shipping_incl_tax: Optional[float] = None,
        base_shipping_invoiced: Optional[float] = None,
        base_shipping_refunded: Optional[float] = None,
        base_shipping_tax_amount: Optional[float] = None,
        base_shipping_tax_refunded: Optional[float] = None,
        base_subtotal: Optional[float] = None,
        base_subtotal_canceled: Optional[float] = None,
        base_subtotal_incl_tax: Optional[float] = None,
        base_subtotal_invoiced: Optional[float] = None,
        base_subtotal_refunded: Optional[float] = None,
        base_tax_amount: Optional[float] = None,
        base_tax_canceled: Optional[float] = None,
        base_tax_invoiced: Optional[float] = None,
        base_tax_refunded: Optional[float] = None,
        base_to_global_rate: Optional[float] = None,
        base_to_order_rate: Optional[float] = None,
        base_total_canceled: Optional[float] = None,
        base_total_due: Optional[float] = None,
        base_total_invoiced: Optional[float] = None,
        base_total_invoiced_cost: Optional[float] = None,
        base_total_offline_refunded: Optional[float] = None,
        base_total_online_refunded: Optional[float] = None,
        base_total_paid: Optional[float] = None,
        base_total_qty_ordered: Optional[float] = None,
        base_total_refunded: Optional[float] = None,
        billing_address: Optional[dict] = None,  # Assuming sales-data-order-address-interface is a dict here
        billing_address_id: Optional[int] = None,
        can_ship_partially: Optional[int] = None,
        can_ship_partially_item: Optional[int] = None,
        coupon_code: Optional[str] = None,
        created_at: Optional[str] = None,
        customer_dob: Optional[str] = None,
        customer_firstname: Optional[str] = None,
        customer_gender: Optional[int] = None,
        customer_group_id: Optional[int] = None,
        customer_id: Optional[int] = None,
        customer_is_guest: Optional[int] = None,
        customer_lastname: Optional[str] = None,
        customer_middlename: Optional[str] = None,
        customer_note: Optional[str] = None,
        customer_note_notify: Optional[int] = None,
        customer_prefix: Optional[str] = None,
        customer_suffix: Optional[str] = None,
        customer_taxvat: Optional[str] = None,
        discount_amount: Optional[float] = None,
        discount_canceled: Optional[float] = None,
        discount_description: Optional[str] = None,
        discount_invoiced: Optional[float] = None,
        discount_refunded: Optional[float] = None,
        discount_tax_compensation_amount: Optional[float] = None,
        discount_tax_compensation_invoiced: Optional[float] = None,
        discount_tax_compensation_refunded: Optional[float] = None,
        edit_increment: Optional[int] = None,
        email_sent: Optional[int] = None,
        entity_id: Optional[int] = None,
        ext_customer_id: Optional[str] = None,
        ext_order_id: Optional[str] = None,
        extension_attributes: Optional[dict] = None,  # Assuming sales-data-order-extension-interface is a dict here
        forced_shipment_with_invoice: Optional[int] = None,
        global_currency_code: Optional[str] = None,
        hold_before_state: Optional[str] = None,
        hold_before_status: Optional[str] = None,
        increment_id: Optional[str] = None,
        is_virtual: Optional[int] = None,
        order_currency_code: Optional[str] = None,
        original_increment_id: Optional[str] = None,
        payment: Optional[dict] = None,  # Assuming sales-data-order-payment-interface is a dict here
        payment_auth_expiration: Optional[int] = None,
        payment_authorization_amount: Optional[float] = None,
        protect_code: Optional[str] = None,
        quote_address_id: Optional[int] = None,
        quote_id: Optional[int] = None,
        relation_child_id: Optional[str] = None,
        relation_child_real_id: Optional[str] = None,
        relation_parent_id: Optional[str] = None,
        relation_parent_real_id: Optional[str] = None,
        remote_ip: Optional[str] = None,
        shipping_amount: Optional[float] = None,
        shipping_canceled: Optional[float] = None,
        shipping_description: Optional[str] = None,
        shipping_discount_amount: Optional[float] = None,
        shipping_discount_tax_compensation_amount: Optional[float] = None,
        shipping_incl_tax: Optional[float] = None,
        shipping_invoiced: Optional[float] = None,
        shipping_refunded: Optional[float] = None,
        shipping_tax_amount: Optional[float] = None,
        shipping_tax_refunded: Optional[float] = None,
        state: Optional[str] = None,
        status: Optional[str] = None,
        status_histories: Optional[List[dict]] = None,  # Assuming sales-data-order-status-history-interface is a dict here
        store_currency_code: Optional[str] = None,
        store_id: Optional[int] = None,
        store_name: Optional[str] = None,
        store_to_base_rate: Optional[float] = None,
        store_to_order_rate: Optional[float] = None,
        subtotal: Optional[float] = None,
        subtotal_canceled: Optional[float] = None,
        subtotal_incl_tax: Optional[float] = None,
        subtotal_invoiced: Optional[float] = None,
        subtotal_refunded: Optional[float] = None,
        tax_amount: Optional[float] = None,
        tax_canceled: Optional[float] = None,
        tax_invoiced: Optional[float] = None,
        tax_refunded: Optional[float] = None,
        total_canceled: Optional[float] = None,
        total_due: Optional[float] = None,
        total_invoiced: Optional[float] = None,
        total_item_count: Optional[int] = None,
        total_offline_refunded: Optional[float] = None,
        total_online_refunded: Optional[float] = None,
        total_paid: Optional[float] = None,
        total_qty_ordered: Optional[float] = None,
        total_refunded: Optional[float] = None,
        updated_at: Optional[str] = None,
        weight: Optional[float] = None,
        x_forwarded_for: Optional[str] = None,
    ):
        self.adjustment_negative = adjustment_negative
        self.adjustment_positive = adjustment_positive
        self.applied_rule_ids = applied_rule_ids
        self.base_adjustment_negative = base_adjustment_negative
        self.base_adjustment_positive = base_adjustment_positive
        self.base_currency_code = base_currency_code
        self.base_discount_amount = base_discount_amount
        self.base_discount_canceled = base_discount_canceled
        self.base_discount_invoiced = base_discount_invoiced
        self.base_discount_refunded = base_discount_refunded
        self.base_discount_tax_compensation_amount = base_discount_tax_compensation_amount
        self.base_discount_tax_compensation_invoiced = base_discount_tax_compensation_invoiced
        self.base_discount_tax_compensation_refunded = base_discount_tax_compensation_refunded
        self.base_grand_total = base_grand_total
        self.base_shipping_amount = base_shipping_amount
        self.base_shipping_canceled = base_shipping_canceled
        self.base_shipping_discount_amount = base_shipping_discount_amount
        self.base_shipping_discount_tax_compensation_amnt = base_shipping_discount_tax_compensation_amnt
        self.base_shipping_incl_tax = base_shipping_incl_tax
        self.base_shipping_invoiced = base_shipping_invoiced
        self.base_shipping_refunded = base_shipping_refunded
        self.base_shipping_tax_amount = base_shipping_tax_amount
        self.base_shipping_tax_refunded = base_shipping_tax_refunded
        self.base_subtotal = base_subtotal
        self.base_subtotal_canceled = base_subtotal_canceled
        self.base_subtotal_incl_tax = base_subtotal_incl_tax
        self.base_subtotal_invoiced = base_subtotal_invoiced
        self.base_subtotal_refunded = base_subtotal_refunded
        self.base_tax_amount = base_tax_amount
        self.base_tax_canceled = base_tax_canceled
        self.base_tax_invoiced = base_tax_invoiced
        self.base_tax_refunded = base_tax_refunded
        self.base_to_global_rate = base_to_global_rate
        self.base_to_order_rate = base_to_order_rate
        self.base_total_canceled = base_total_canceled
        self.base_total_due = base_total_due
        self.base_total_invoiced = base_total_invoiced
        self.base_total_invoiced_cost = base_total_invoiced_cost
        self.base_total_offline_refunded = base_total_offline_refunded
        self.base_total_online_refunded = base_total_online_refunded
        self.base_total_paid = base_total_paid
        self.base_total_qty_ordered = base_total_qty_ordered
        self.base_total_refunded = base_total_refunded
        self.billing_address = billing_address
        self.billing_address_id = billing_address_id
        self.can_ship_partially = can_ship_partially
        self.can_ship_partially_item = can_ship_partially_item
        self.coupon_code = coupon_code
        self.created_at = created_at
        self.customer_dob = customer_dob
        self.customer_email = customer_email
        self.customer_firstname = customer_firstname
        self.customer_gender = customer_gender
        self.customer_group_id = customer_group_id
        self.customer_id = customer_id
        self.customer_is_guest = customer_is_guest
        self.customer_lastname = customer_lastname
        self.customer_middlename = customer_middlename
        self.customer_note = customer_note
        self.customer_note_notify = customer_note_notify
        self.customer_prefix = customer_prefix
        self.customer_suffix = customer_suffix
        self.customer_taxvat = customer_taxvat
        self.discount_amount = discount_amount
        self.discount_canceled = discount_canceled
        self.discount_description = discount_description
        self.discount_invoiced = discount_invoiced
        self.discount_refunded = discount_refunded
        self.discount_tax_compensation_amount = discount_tax_compensation_amount
        self.discount_tax_compensation_invoiced = discount_tax_compensation_invoiced
        self.discount_tax_compensation_refunded = discount_tax_compensation_refunded
        self.edit_increment = edit_increment
        self.email_sent = email_sent
        self.entity_id = entity_id
        self.ext_customer_id = ext_customer_id
        self.ext_order_id = ext_order_id
        self.extension_attributes = extension_attributes
        self.forced_shipment_with_invoice = forced_shipment_with_invoice
        self.global_currency_code = global_currency_code
        self.grand_total = grand_total
        self.hold_before_state = hold_before_state
        self.hold_before_status = hold_before_status
        self.increment_id = increment_id
        self.is_virtual = is_virtual
        self.items = items.to_json()
        self.order_currency_code = order_currency_code
        self.original_increment_id = original_increment_id
        self.payment = payment
        self.payment_auth_expiration = payment_auth_expiration
        self.payment_authorization_amount = payment_authorization_amount
        self.protect_code = protect_code
        self.quote_address_id = quote_address_id
        self.quote_id = quote_id
        self.relation_child_id = relation_child_id
        self.relation_child_real_id = relation_child_real_id
        self.relation_parent_id = relation_parent_id
        self.relation_parent_real_id = relation_parent_real_id
        self.remote_ip = remote_ip
        self.shipping_amount = shipping_amount
        self.shipping_canceled = shipping_canceled
        self.shipping_description = shipping_description
        self.shipping_discount_amount = shipping_discount_amount
        self.shipping_discount_tax_compensation_amount = shipping_discount_tax_compensation_amount
        self.shipping_incl_tax = shipping_incl_tax
        self.shipping_invoiced = shipping_invoiced
        self.shipping_refunded = shipping_refunded
        self.shipping_tax_amount = shipping_tax_amount
        self.shipping_tax_refunded = shipping_tax_refunded
        self.state = state
        self.status = status
        self.status_histories = status_histories
        self.store_currency_code = store_currency_code
        self.store_id = store_id
        self.store_name = store_name
        self.store_to_base_rate = store_to_base_rate
        self.store_to_order_rate = store_to_order_rate
        self.subtotal = subtotal
        self.subtotal_canceled = subtotal_canceled
        self.subtotal_incl_tax = subtotal_incl_tax
        self.subtotal_invoiced = subtotal_invoiced
        self.subtotal_refunded = subtotal_refunded
        self.tax_amount = tax_amount
        self.tax_canceled = tax_canceled
        self.tax_invoiced = tax_invoiced
        self.tax_refunded = tax_refunded
        self.total_canceled = total_canceled
        self.total_due = total_due
        self.total_invoiced = total_invoiced
        self.total_item_count = total_item_count
        self.total_offline_refunded = total_offline_refunded
        self.total_online_refunded = total_online_refunded
        self.total_paid = total_paid
        self.total_qty_ordered = total_qty_ordered
        self.total_refunded = total_refunded
        self.updated_at = updated_at
        self.weight = weight
        self.x_forwarded_for = x_forwarded_for



    def to_json(self):
        entity = {}

        for key, value in self.__dict__.items():
            if value is not None:
                entity[key] = value

        return json.dumps({"entity": entity}, indent=4)
    
class AttributeSet:
    def __init__(self, attribute_set_name:str, sort_order:int, skeleton_id:int, attribute_set_id: int = None, entity_type_id: int = None, extension_attributes=None):
        """
        Initialize an AttributeSet object.

        Args:
            attribute_set_name (str): The name of the attribute set.
            sort_order (int): The sort order of the attribute set.
            skeleton_id (int): The skeleton ID for the attribute set.
            attribute_set_id (int, optional): The ID of the attribute set. Defaults to None.
            entity_type_id (int, optional): The entity type ID associated with the attribute set. Defaults to None.
            extension_attributes (dict, optional): Additional extension attributes. Defaults to None.
        """
        self.attribute_set_name = attribute_set_name
        self.sort_order = sort_order
        self.skeleton_id = skeleton_id
        self.attribute_set_id = attribute_set_id
        self.entity_type_id = entity_type_id
        self.extension_attributes = extension_attributes
    def to_dict(self) -> dict:
        """
        Convert the AttributeSet instance to a dictionary with the desired format.
        
        Returns:
            dict: A dictionary representation of the AttributeSet instance.
        """
        attribute_set_dict = {
            "attribute_set_name": self.attribute_set_name,
            "sort_order": self.sort_order,
            "extension_attributes": self.extension_attributes
        }

        if self.attribute_set_id is not None:
            attribute_set_dict["attribute_set_id"] = self.attribute_set_id

        if self.entity_type_id is not None:
            attribute_set_dict["entity_type_id"] = self.entity_type_id

        return {
            "attributeSet": attribute_set_dict,
            "skeletonId": self.skeleton_id
        }
class Attribute:
    def __init__(self, attribute_code: str, attribute_group_id: int, attribute_set_id: int, sort_order: int):
        """
        Initialize an Attribute object.

        Args:
            attribute_code (str): The code of the attribute.
            attribute_group_id (int): The ID of the attribute group to which the attribute belongs.
            attribute_set_id (int): The ID of the attribute set to which the attribute belongs.
            sort_order (int): The sort order of the attribute within the attribute set.
        """
        self.attribute_code = attribute_code
        self.attribute_group_id = attribute_group_id
        self.attribute_set_id = attribute_set_id
        self.sort_order = sort_order

    def to_dict(self) -> dict:
        """
        Convert the Attribute instance to a dictionary with the desired format.
        
        Returns:
            dict: A dictionary representation of the Attribute instance.
        """
        return {
            "attributeCode": self.attribute_code,
            "attributeGroupId": self.attribute_group_id,
            "attributeSetId": self.attribute_set_id,
            "sortOrder": self.sort_order
        }
class Filter:
    def __init__(self, conditionType, field, value):
        """
        Initialize a Filter object.

        Args:
            conditionType (str): The condition type for the filter.
            field (str): The field to filter by.
            value (str): The value to filter by.
        """
        self.conditionType = conditionType
        self.field = field
        self.value = value
    
    def to_dict(self):
        """
        Convert the Filter instance to a dictionary.

        Returns:
            dict: A dictionary representation of the Filter instance.
        """
        return {
            'conditionType': self.conditionType,
            'field': self.field,
            'value': self.value
        }
class Product:
    def __init__(self, id):
        pass

class FilterGroup:
    def __init__(self, filters=None):
        """
        Initialize a FilterGroup object.

        Args:
            filters (list of Filter, optional): A list of Filter objects representing the filters to be applied. Defaults to None.
        """
        self.filters = filters if filters else []

    def add_filter(self, conditionType, field, value):
        """
        Add a filter to the FilterGroup.

        Args:
            conditionType (str): The condition type for the filter.
            field (str): The field to filter by.
            value (str): The value to filter by.
        """
        self.filters.append(Filter(conditionType, field, value))
    def add_date_filter(self, date_from, date_to):
        """
        Add a date filter to the filter group.

        Args:
            date (str): The date value for filtering.
            conditionType (str): The condition type for the date filter.
        """

        self.filters.append(Filter('from', 'created_at', date_from))
        self.filters.append(Filter("to", 'created_at', date_to))

    def to_dict(self):
        """
        Convert the FilterGroup instance to a dictionary.

        Returns:
            dict: A dictionary representation of the FilterGroup instance.
        """
        return {
            'filtersGroups': [f.to_dict() for f in self.filters]
        }
    
class Group:
    def __init__(self, attribute_group_id: int, attribute_group_name: str, attribute_set_id: int, extension_attributes=None):
        """
        Initialize a Group object.

        Args:
            attribute_group_id (int): The ID of the attribute group.
            attribute_group_name (str): The name of the attribute group.
            attribute_set_id (int): The ID of the attribute set to which the group belongs.
            extension_attributes (dict, optional): Additional extension attributes. Defaults to None.
        """
        
        self.attribute_group_id = attribute_group_id
        self.attribute_group_name = attribute_group_name
        self.attribute_set_id = attribute_set_id
        self.extension_attributes = extension_attributes
    def to_dict(self):
        """
        Convert the Group instance to a dictionary with the desired format.
        
        Returns:
            dict: A dictionary representation of the Group instance.
        """
        group_dict = {
            "group": {
                "attribute_group_id": self.attribute_group_id,
                "attribute_group_name": self.attribute_group_name,
                "attribute_set_id": self.attribute_set_id,
                "extension_attributes": self.extension_attributes
            }
        }
        return group_dict


class APIRequestError(Exception):
    """Exception raised for errors during API requests.

    Attributes:
        message -- explanation of the error
    """

    def __init__(self, message):
        self.message = message
        super().__init__(self.message)
        
class Params:
    def __init__(self, currentPage=None, filterGroups:FilterGroup=None, pageSize=None, sortOrders:dict=None):
        """
        Initialize a Params object.

        Args:
            currentPage (int, optional): The current page number. Defaults to None.
            filterGroups (FilterGroup, optional): An instance of FilterGroup representing the filter groups. Defaults to None.
            pageSize (int, optional): The number of items per page. Defaults to None.
            sortOrders (list of dict, optional): A list of dictionaries containing sorting orders. Each dictionary should contain 'direction' and 'field' keys. Defaults to None.
        """
        self.currentPage = currentPage
        self.filterGroups = filterGroups
        self.pageSize = pageSize
        self.sortOrders = sortOrders
    
    def to_dict(self):
        """
        Convert the Params instance to a dictionary with the desired format.

        Returns:
            dict: A dictionary representation of the Params instance.
        """
        params_dict = {
            'searchCriteria': {}
        }

        if self.currentPage is not None:
            params_dict['searchCriteria']['currentPage'] = self.currentPage
        if self.filterGroups is not None:
            filter_groups_dict = self.filterGroups.to_dict()
            params_dict['searchCriteria']['filterGroups'] = []
            for idx, filter_data in enumerate(filter_groups_dict['filtersGroups']):
                filter_group = {
                    'filters': [{
                        'conditionType': filter_data['conditionType'],
                        'field': filter_data['field'],
                        'value': filter_data['value']
                    }]
                }
                params_dict['searchCriteria']['filterGroups'].append(filter_group)
        if self.pageSize is not None:
            params_dict['searchCriteria']['pageSize'] = self.pageSize
        if self.sortOrders is not None:
            params_dict['searchCriteria']['sortOrders'] = []
            for idx, sort_order in enumerate(self.sortOrders):
                params_dict['searchCriteria']['sortOrders'].append({
                    'direction': sort_order['direction'],
                    'field': sort_order['field']
                })

        return params_dict
    


class LoginController:
    """
    A controller class for managing the login process to the API. 

    This class handles user authentication, stores the session token, 
    and checks the validity of the login session based on the token expiration time.
    """

    def __init__(self, endpoint, username=None, password=None):
        """Initialize the LoginController with a username and password. 
        If not provided, prompts the user for input."""
        self.token = None
        self.store = "default"
        self.logged_in = False
        self.api_endpoint = f"{endpoint}{self.store}/V1"
        self.token_expiration = timedelta(hours=4)  # Set the token expiration time (4 hours in this example)
        self.username = username
        self.password = password
        if not self.username or not self.password:
            self.username = input("Username: ")
            self.password = getpass()

    def login(self):
        """Logs in the user by sending a POST request to the API's token endpoint.
        If successful, stores the token and sets the logged-in status and login time."""
        endpoint = self.api_endpoint + "/integration/admin/token"
        payload = {"username": self.username, "password": self.password}
        headers = {"Content-Type": "application/json"}

        response = requests.post(endpoint, headers=headers, params=payload)
        if response.status_code == requests.codes.ok:
            # The request was successful
            self.token = response.json()
            self.login_time = datetime.now()
            self.logged_in = True
            print("Login successful")
            return True
        else:
            # The request failed, print error message or handle accordingly
            print("Login failed")
            raise InvalidCredentialsError("Please log in to access product details.")

    def is_logged_in(self):
        """Checks if the user is currently logged in by verifying the token's validity.
        If the token has expired, returns False and updates the logged-in status."""
        if not self.logged_in:
            return False

        current_time = datetime.now()
        elapsed_time = current_time - self.login_time

        # Check if the token has expired
        if elapsed_time > self.token_expiration:
            print("Token has expired.")
            self.logged_in = False
            print("Attempting to re-login")
            return self.login()
        else:
            return True


class Magento:
    def make_api_request(self, endpoint: str, params: Params = None, request_type: str = "get", data: dict = None, json:dict = None) -> dict:
        if not self.login_controller.is_logged_in():
            raise InvalidCredentialsError("Please log in to access product details.")
        """
        Makes an API request of the specified type (GET, POST, DELETE, PUT) to the given endpoint.

        Args:
            endpoint (str): The API endpoint to send the request to.
            params (Optional[Params]): An instance of the Params class containing query parameters.
            request_type (str): The type of HTTP request to make (e.g., "get", "post", "delete"). Default is "get".
            data (Optional[Dict]): The data to send with the request (for POST requests).

        Raises:
            InvalidCredentialsError: If the user is not logged in.
            PermissionDeniedError: If the user does not have permission to access the data.
            APIRequestError: If the request fails.

        Returns:
            dict: The response data from the API request if successful.
        """
        headers = {
        "Authorization": f"Bearer {self.login_controller.token}",
        "Accept": "application/json",
        "Content-Type" : "application/json"
        }
        if request_type.lower() == "get":
            response = requests.get(endpoint, headers=headers, params=params.to_dict() if params is not None else None)
            append_json_with_timestamp(response.json()) if response.json() else None
        elif request_type.lower() == "post":
            response = requests.post(endpoint, headers=headers, params=params.to_dict() if params is not None else None, data=data if data is not None else None, json=json if json is not None else None)
            print(response.text)
            append_json_with_timestamp(response.json()) if response.json() else None
        elif request_type.lower() == "delete":
            response = requests.delete(endpoint, headers=headers, params=params.to_dict() if params is not None else None)
            append_json_with_timestamp(response.json()) if response.json() else None

        elif request_type.lower() == "put":
            response = requests.put(endpoint, headers=headers, params=params.to_dict() if params is not None else None,data=data if data is not None else None, json=json if json is not None else None)
            print(response.text)
            append_json_with_timestamp(response.json()) if response.json() else None

        else:
            raise ValueError(f"Unsupported request type: {request_type}")

        if response.status_code == requests.codes.ok:
            return response.json()
        elif response.status_code == 401:
            raise PermissionDeniedError()
        else:
            raise APIRequestError(f"Failed to fetch data from {endpoint}" + response.text)
    
    # def save_details(self, data, type):
    #     self.type = type
    #     if not self.login_controller.is_logged_in():
    #         raise InvalidCredentialsError("Please log in to access product details.")
    #     try:
    #         if self.type == 1:
    #             print("Product")
    #             df = json_normalize(data)
    #             # Specify the columns containing nested dictionaries that you want to normalize further
    #             columns_to_normalize = ['media_gallery_entries', 'custom_attributes','extension_attributes.category_links','extension_attributes.configurable_product_options']
    #             for column in columns_to_normalize:
    #                 if column in df.columns:
    #                     normalized_data = json_normalize(df.pop(column))
    #                     df = df.join(normalized_data.add_prefix(f'{column}.'))

    #             df.to_excel('./Output.xlsx', index=False)

    #         elif self.type == 2:
    #             print("Order")
    #             df = json_normalize(data)
    #             # Specify the columns containing nested dictionaries that you want to normalize further
    #             df.to_excel('./Output.xlsx', index=False)
    #     except PermissionError as e:
    #         print("Sorry You don't have access to write to that file please close it.")
    def get_all_stores(self) -> dict:
        """
        Retrieves a mapping of all store IDs to their corresponding store codes.

        This method sends a request to the store views API endpoint and constructs
        a dictionary that maps store IDs (as strings) to their respective store codes.

        Returns:
            dict: A dictionary where the keys are store IDs (as strings) and the values
                are store codes.

        Example:
            # Get all store mappings
            store_mapping = get_all_stores()
            print(store_mapping)
            # Output: {'1': 'default', '2': 'store_code', ...}

        Notes:
            - This method assumes that the API response contains a list of store data, where each
            store data item includes 'website_id' and 'code' keys.
            - If the structure of the API response is different, adjust the code accordingly.
            - If any store data item does not have the expected structure, a warning message
            will be printed to the console.

        """
        endpoint = f"{self.login_controller.api_endpoint}/store/storeViews"
        stores_data = self.make_api_request(endpoint)
        # Adjust this part based on the actual structure of the API response
        store_mapping = {}
        for store in stores_data:
            if 'website_id' in store and 'code' in store:
                store_mapping[str(store['website_id'])] = store['code']
            else:
                print(f"Warning: Unexpected structure in store data: {store}")

        return store_mapping
    # def export_details(self,type):
    #             # Read the Excel file into a Pandas DataFrame
    #     df = read_excel('./Output.xlsx')

    #     # Reverse the normalization process to reconstruct the nested structure
    #     columns_to_denormalize = ['media_gallery_entries', 'custom_attributes','extension_attributes.category_links','extension_attributes.configurable_product_options']
    #     denormalized_data = df.copy()

    #     for column in columns_to_denormalize:
    #         if f"{column}.0" in denormalized_data.columns:
    #             # Collect columns related to the normalized data
    #             related_columns = [col for col in denormalized_data.columns if col.startswith(f"{column}.")]
    #             # Reconstruct the nested structure
    #             nested_data = denormalized_data[related_columns].to_dict(orient='records')
    #             # Replace the normalized columns with the reconstructed nested structure
    #             denormalized_data[column] = nested_data
    #             # Drop the intermediate columns
    #             denormalized_data.drop(columns=related_columns, inplace=True)

    #     # Convert the DataFrame to a list of dictionaries
    #     denormalized_list = denormalized_data.to_dict(orient='records')
    #     print(denormalized_list)
    
class Customer(Magento):
    def __init__(self, login_controller):
        super().__init__()
        self.product_data = None
        self.login_controller = login_controller
        if not self.login_controller.is_logged_in():
            self.login_controller.login()
    def search(self, firstname=None,lastname=None, email=None, phone=None) -> dict:
        filtergroup = FilterGroup()
        if firstname is not None:
            filtergroup.add_filter('eq','device_order_id',firstname)
        if lastname is not None:
            filtergroup.add_filter('eq','firstname',lastname)
        if email is not None: 
            filtergroup.add_filter('eq','email',email)
        if phone is not None:
            filtergroup.add_filter('eq','sms_mobile_phone_number', phone)
        if filtergroup is not None:
            params = Params(1,filtergroup,pageSize=2000)
        else:
            params = Params(1,pageSize=500)
        endpoint = f"{self.login_controller.api_endpoint}/customers/search"
        # endpoint = "https://www.reedssports.com/rest/V8/pos/orders"
        # Start the timer
        start_time = time.time()
        print(params.to_dict())
        # Make the API request
        response = self.make_api_request(endpoint, params=params)
        
        # Stop the timer
        end_time = time.time()
        
        # Calculate the elapsed time
        elapsed_time = end_time - start_time
        
        # Print the elapsed time
        print(f"Time taken for the request: {elapsed_time:.2f} seconds")
        
        return response
    def test(self):
        endpoint = "https://www.reedssports.com/rest/V8/pos/orders?searchCriteria%5Bfilter_groups%5D%5B0%5D%5Bfilters%5D%5B0%5D%5Bfield%5D=created_at&searchCriteria%5Bfilter_groups%5D%5B0%5D%5Bfilters%5D%5B0%5D%5Bvalue%5D=0001-01-01&searchCriteria%5Bfilter_groups%5D%5B0%5D%5Bfilters%5D%5B0%5D%5Bcondition_type%5D=gteq&searchCriteria%5Bfilter_groups%5D%5B1%5D%5Bfilters%5D%5B0%5D%5Bfield%5D=created_at&searchCriteria%5Bfilter_groups%5D%5B1%5D%5Bfilters%5D%5B0%5D%5Bvalue%5D=2024-06-06&searchCriteria%5Bfilter_groups%5D%5B1%5D%5Bfilters%5D%5B0%5D%5Bcondition_type%5D=lteq&searchCriteria%5Bfilter_groups%5D%5B2%5D%5Bfilters%5D%5B0%5D%5Bfield%5D=entity_id&searchCriteria%5Bfilter_groups%5D%5B2%5D%5Bfilters%5D%5B0%5D%5Bvalue%5D=25000o2yeF3300003236&searchCriteria%5Bfilter_groups%5D%5B2%5D%5Bfilters%5D%5B0%5D%5Bcondition_type%5D=like&searchCriteria%5Bfilter_groups%5D%5B2%5D%5Bfilters%5D%5B1%5D%5Bfield%5D=increment_id&searchCriteria%5Bfilter_groups%5D%5B2%5D%5Bfilters%5D%5B1%5D%5Bvalue%5D=%2525000o2yeF3300003236%2525&searchCriteria%5Bfilter_groups%5D%5B2%5D%5Bfilters%5D%5B1%5D%5Bcondition_type%5D=like&searchCriteria%5Bfilter_groups%5D%5B2%5D%5Bfilters%5D%5B2%5D%5Bfield%5D=customer_email&searchCriteria%5Bfilter_groups%5D%5B2%5D%5Bfilters%5D%5B2%5D%5Bvalue%5D=%2525000o2yeF3300003236%2525&searchCriteria%5Bfilter_groups%5D%5B2%5D%5Bfilters%5D%5B2%5D%5Bcondition_type%5D=like&searchCriteria%5Bfilter_groups%5D%5B2%5D%5Bfilters%5D%5B3%5D%5Bfield%5D=customer_lastname&searchCriteria%5Bfilter_groups%5D%5B2%5D%5Bfilters%5D%5B3%5D%5Bvalue%5D=%2525000o2yeF3300003236%2525&searchCriteria%5Bfilter_groups%5D%5B2%5D%5Bfilters%5D%5B3%5D%5Bcondition_type%5D=like&searchCriteria%5Bfilter_groups%5D%5B2%5D%5Bfilters%5D%5B4%5D%5Bfield%5D=customer_firstname&searchCriteria%5Bfilter_groups%5D%5B2%5D%5Bfilters%5D%5B4%5D%5Bvalue%5D=%2525000o2yeF3300003236%2525&searchCriteria%5Bfilter_groups%5D%5B2%5D%5Bfilters%5D%5B4%5D%5Bcondition_type%5D=like&searchCriteria%5Bfilter_groups%5D%5B2%5D%5Bfilters%5D%5B5%5D%5Bfield%5D=device_order_id&searchCriteria%5Bfilter_groups%5D%5B2%5D%5Bfilters%5D%5B5%5D%5Bvalue%5D=000o2yeF3300003236&searchCriteria%5Bfilter_groups%5D%5B2%5D%5Bfilters%5D%5B5%5D%5Bcondition_type%5D=like&searchCriteria%5Bfilter_groups%5D%5B2%5D%5Bfilters%5D%5B6%5D%5Bfield%5D=shipping_method&searchCriteria%5Bfilter_groups%5D%5B2%5D%5Bfilters%5D%5B6%5D%5Bvalue%5D=%2525000o2yeF3300003236%2525&searchCriteria%5Bfilter_groups%5D%5B2%5D%5Bfilters%5D%5B6%5D%5Bcondition_type%5D=like&searchCriteria%5BcurrentPage%5D=1&searchCriteria%5BpageSize%5D=10&searchCriteria%5BsortOrders%5D%5B0%5D%5Bfield%5D=created_at&searchCriteria%5BsortOrders%5D%5B0%5D%5Bdirection%5D=desc"

        # Start the timer
        start_time = time.time()
        response = self.make_api_request(endpoint)
        # Stop the timer
        end_time = time.time()

        elapsed_time = end_time - start_time
        print(response)

        print(f"Time taken for the request: {elapsed_time:.2f} seconds")


class Bulk(Magento):
    def __init__(self, login_controller):
        super().__init__()
        self.product_data = None
        self.login_controller = login_controller
        if not self.login_controller.is_logged_in():
            self.login_controller.login()
    def bulk(self,params: Params) -> dict:
        """
        Fetches bulk data from the API endpoint associated with the login controller.

        This method retrieves a list of items from a store API endpoint using the bulk API functionality.
        It first checks if the user is logged in. If the user is not logged in, it raises an `InvalidCredentialsError`.
        If the user is logged in, it sends a GET request to the specified API endpoint with the appropriate 
        authorization header and query parameters. The method processes the response and returns the product data 
        if the request is successful.

        Args:
            params (Params): An instance of the Params class containing query parameters.

        Raises:
            InvalidCredentialsError: If the user is not logged in.
            PermissionDeniedError: If the user does not have permission to access the data.

        Returns:
            dict: A dictionary containing the product data if the request is successful.
            None: If the request fails or the user is not authorized.
        """
        endpoint = f"{self.login_controller.api_endpoint}/bulk"
        return self.make_api_request(endpoint, params=params)
    def detailed_status(self,bulkUuid: str) -> dict:
        """
        Retrieves detailed status information for a bulk operation identified by its UUID.

        This method fetches detailed status information for a specific bulk operation from the store API endpoint.
        It first checks if the user is logged in. If the user is not logged in, it raises an `InvalidCredentialsError`.
        If the user is logged in, it sends a GET request to the specified API endpoint with the appropriate 
        authorization header and the UUID of the bulk operation. The method processes the response and returns the 
        detailed status data if the request is successful.

        Args:
            bulkUuid (str): The UUID (Universally Unique Identifier) of the bulk operation for which detailed status
                            information is requested.

        Raises:
            InvalidCredentialsError: If the user is not logged in.
            PermissionDeniedError: If the user does not have permission to access the data.

        Returns:
            dict: A dictionary containing the detailed status information of the bulk operation if the request 
                is successful.
            None: If the request fails or the user is not authorized.
        """
        
        endpoint = f"{self.login_controller.api_endpoint}/bulk/{bulkUuid}/detailed-status"
        return self.make_api_request(endpoint)
    def operation_status(self,bulkUuid: str,status = int) -> dict:
        """
        Retrieves operation status information for a specific status of a bulk operation identified by its UUID.

        This method fetches operation status information for a specific status of a bulk operation from the store API endpoint.
        It first checks if the user is logged in. If the user is not logged in, it raises an `InvalidCredentialsError`.
        If the user is logged in, it sends a GET request to the specified API endpoint with the appropriate 
        authorization header, the UUID of the bulk operation, and the status of the operation.
        The method processes the response and returns the operation status data if the request is successful.

        Args:
            bulkUuid (str): The UUID (Universally Unique Identifier) of the bulk operation for which operation status
                            information is requested.
            status (int): The status of the operation for which status information is requested.

        Raises:
            InvalidCredentialsError: If the user is not logged in.
            PermissionDeniedError: If the user does not have permission to access the data.

        Returns:
            dict: A dictionary containing the operation status information of the bulk operation for the specified status
                if the request is successful.
            None: If the request fails or the user is not authorized.
        """
        endpoint = f"{self.login_controller.api_endpoint}/bulk/{bulkUuid}/operation-status/{status}"
        return self.make_api_request(endpoint)
    def status(self,bulkUuid: str) -> dict:
        """
        Fetches the status of a bulk operation identified by its UUID.

        This method retrieves the status information of a specific bulk operation from the store API endpoint.
        It first checks if the user is logged in. If the user is not logged in, it raises an `InvalidCredentialsError`.
        If the user is logged in, it sends a GET request to the specified API endpoint with the appropriate 
        authorization header and the UUID of the bulk operation. The method processes the response and returns the 
        status information if the request is successful.

        Args:
            bulkUuid (str): The UUID (Universally Unique Identifier) of the bulk operation for which status information
                            is requested.

        Raises:
            InvalidCredentialsError: If the user is not logged in.
            PermissionDeniedError: If the user does not have permission to access the data.

        Returns:
            dict: A dictionary containing the status information of the bulk operation if the request is successful.
            None: If the request fails or the user is not authorized.
        """
        endpoint = f"{self.login_controller.api_endpoint}/bulk/{bulkUuid}/status"
        return self.make_api_request(endpoint)
        
class Store(Magento):
    def __init__(self, login_controller):
        super().__init__()
        self.product_data = None
        self.login_controller = login_controller
        if not self.login_controller.is_logged_in():
            self.login_controller.login()
    def websites(self) -> dict:
        """
        Fetches website data from the API endpoint associated with the login controller.

        This method retrieves a list of websites from a store API endpoint. It first checks if the user is logged in.
        If the user is not logged in, it raises an `InvalidCredentialsError`. If the user is logged in, it sends
        a GET request to the specified API endpoint with the appropriate authorization header. The method processes 
        the response and returns the product data if the request is successful.

        Raises:
            InvalidCredentialsError: If the user is not logged in.

        Returns:
            dict: A dictionary containing the product data if the request is successful.
            None: If the request fails or the user is not authorized.
        """
        endpoint = f"{self.login_controller.api_endpoint}/store/websites"  # Example endpoint
        return self.make_api_request(endpoint)
    def storeViews(self) -> dict:
        """
        Fetches store views from the API endpoint associated with the login controller.

        This method retrieves the store views from the store API endpoint.
        It first checks if the user is logged in. If the user is not logged in, it raises an `InvalidCredentialsError`.
        If the user is logged in, it sends a GET request to the specified API endpoint with the appropriate 
        authorization header. The method processes the response and returns the store views 
        if the request is successful.

        Raises:
            InvalidCredentialsError: If the user is not logged in.
            PermissionDeniedError: If the user does not have permission to access the data.

        Returns:
            dict: A dictionary containing the store views if the request is successful.
            None: If the request fails or the user is not authorized.
        """
        endpoint = f"{self.login_controller.api_endpoint}/store/storeViews"  # Example endpoint
        return self.make_api_request(endpoint)

    def storeConfigs(self) -> dict:
        """
        Fetches store configurations from the API endpoint associated with the login controller.

        This method retrieves the configurations of the store from the store API endpoint.
        It first checks if the user is logged in. If the user is not logged in, it raises an `InvalidCredentialsError`.
        If the user is logged in, it sends a GET request to the specified API endpoint with the appropriate 
        authorization header. The method processes the response and returns the store configurations 
        if the request is successful.

        Raises:
            InvalidCredentialsError: If the user is not logged in.
            PermissionDeniedError: If the user does not have permission to access the data.

        Returns:
            dict: A dictionary containing the store configurations if the request is successful.
            None: If the request fails or the user is not authorized.
        """
        endpoint = f"{self.login_controller.api_endpoint}/store/storeConfigs"  # Example endpoint
        return self.make_api_request(endpoint)

class ProductHandler(Magento):
    def __init__(self, login_controller):
        super().__init__()
        self.product_data = None
        self.login_controller = login_controller
        if not self.login_controller.is_logged_in():
            self.login_controller.login()

    def attribute_sets(self, attributeset: AttributeSet) -> dict:
        """
        Creates a new attribute set for products in the Magento 2 system.

        This method sends a POST request to the Magento 2 API to create a new attribute set using the details provided in the `attributeset` parameter.

        Parameters:
        attributeset (AttributeSet): An instance of the AttributeSet class containing the details of the attribute set to be created.

        Returns:
        dict: The response from the Magento 2 API, typically containing details of the created attribute set or an error message if the request fails.

        Example:
            attribute_set = AttributeSet(attribute_set_name="Clothing", sort_order=1, skeleton_id=4)
            response = ProductHandler.attribute_sets(attribute_set)
            print(response)

        Raises:
        ValueError: If the response from the API indicates an error.
        """
        endpoint = f"{self.login_controller.api_endpoint}/products/attribute-sets"
        return self.make_api_request(endpoint, type='post',data=attributeset.to_dict())
    def attribute_sets_attributes(self, attribute: Attribute ) -> dict:
        """
        Create or update attributes assigned to an attribute set.

        Args:
            attribute (Attribute): An instance of the Attribute class representing the attribute to be assigned to the attribute set.

        Returns:
            dict: A dictionary containing the response from the API request.
        """
        endpoint = f"{self.login_controller.api_endpoint}/products/attribute-sets/attributes"
        return self.make_api_request(endpoint, type='post',data=attribute.to_dict())
    def group_attributes(self, group: Group) -> dict:
        """
        Create or update attributes assigned to an attribute set group.

        Args:
            group (Group): An instance of the Group class representing the attribute group to be assigned to the attribute set.

        Returns:
            dict: A dictionary containing the response from the API request.
        """
        endpoint = f"{self.login_controller.api_endpoint}/products/attribute-sets/groups"
        return self.make_api_request(endpoint, type='post', data=group.to_dict())
    def get_attribute_set_groups_list(self, params:Params) -> dict:
        """
        Retrieve a list of attribute groups for attribute sets.

        Args:
            params (Params): An instance of Params containing the parameters for the request, including:
                - currentPage (int, optional): The current page number.
                - pageSize (int, optional): The number of items per page.
                - sortOrders (list of dict, optional): A list of dictionaries containing sorting orders, with each dictionary containing 'direction' and 'field' keys.
                - filterGroups (FilterGroup, optional): An instance of FilterGroup representing the filter groups.

        Returns:
            dict: A dictionary containing the response from the API request.
        """
        endpoint = f"{self.login_controller.api_endpoint}/products/attribute-sets/groups/list"
        return self.make_api_request(endpoint, params=params, type='get')


    def get_product_details(self, sku: str) -> dict:
        """
        Retrieves details of a product from the Magento 2 system using the provided SKU.

        This method sends a GET request to the Magento 2 API to retrieve details of the product identified by the given SKU.

        Parameters:
        sku (str): The Stock Keeping Unit (SKU) of the product to retrieve details for.

        Returns:
        dict: A dictionary containing details of the product, including attributes such as name, price, status, visibility, etc.

        Example:
            product_details = ProductHandler.get_product_details("SKU123")
            print(product_details)

        Raises:
        ValueError: If the response from the API indicates an error or if the provided SKU is invalid.
        """
        endpoint = f"{self.login_controller.api_endpoint}/products/{sku}"
        return self.make_api_request(endpoint)

    def save_details(self, data):
        return super().save_details(data,1)
    def get_media(self,sku:str) -> dict:
        """
        Fetches media information for a product identified by its SKU from the API endpoint associated with the login controller.

        This method retrieves the media information for a specific product from the products API endpoint.
        It first checks if the user is logged in. If the user is not logged in, it raises an `InvalidCredentialsError`.
        If the user is logged in, it sends a GET request to the specified API endpoint with the appropriate 
        authorization header and the SKU of the product. The method processes the response and returns the media information 
        if the request is successful.

        Args:
            sku (str): The SKU (Stock Keeping Unit) of the product for which media information is requested.

        Raises:
            InvalidCredentialsError: If the user is not logged in.
            PermissionDeniedError: If the user does not have permission to access the data.

        Returns:
            dict: A dictionary containing the media information of the product if the request is successful.
            None: If the request fails or the user is not authorized.
        """
        endpoint = f"{self.login_controller.api_endpoint}/products/{sku}/media"  # Example endpoint
        return self.make_api_request(endpoint)
    
    def delete_media(self,sku,entryID:int) -> dict:
        """
        Deletes media information for a product identified by its SKU and media entry ID from the API endpoint associated with the login controller.

        This method deletes the media information for a specific product from the store API endpoint.
        It first checks if the user is logged in. If the user is not logged in, it raises an `InvalidCredentialsError`.
        If the user is logged in, it sends a DELETE request to the specified API endpoint with the appropriate 
        authorization header, the SKU of the product, and the media entry ID. The method processes the response and returns 
        the result of the delete operation if the request is successful.

        Args:
            sku (str): The SKU (Stock Keeping Unit) of the product for which media information is to be deleted.
            entryID (int): The ID of the media entry to be deleted.

        Raises:
            InvalidCredentialsError: If the user is not logged in.
            PermissionDeniedError: If the user does not have permission to delete the data.
            APIRequestError: If the request fails.

        Returns:
            dict: A dictionary containing the result of the delete operation if the request is successful.
            None: If the request fails or the user is not authorized.
        """
        endpoint = f"{self.login_controller.api_endpoint}/products/{sku}/media/{entryID}"  # Example endpoint
        return self.make_api_request(endpoint=endpoint, request_type='delete')
    def turn_offs(self, sku:str, store_codes:list, archive: bool = False) -> dict:
        """
        This method is used to turn off specific stores for a given product SKU.

        Args:
            sku (str): The SKU of the product for which stores need to be turned off.
            store_codes (list of str): A list of store codes to be turned off for the product.
            archive (bool, optional): Whether to archive the product or not. Defaults to False.

        Returns:
            dict: A dictionary containing the updated product data after turning off the specified stores.

        Raises:
            StoreTurnOffRestrictedError: If attempting to turn off a store that is restricted or if exception source is not blank.

        Notes:
            - This method retrieves product details for the specified SKU and checks if the provided store codes can be turned off.
            - If the `archive` parameter is set to True, the product will be archived instead of turning off stores.
            - The method checks if specific store names are already applied to the product and adjusts the store codes accordingly.
            - It validates if the provided store codes are associated with the product and then updates the store associations accordingly.
            - If any errors occur during the process, appropriate messages are printed, and the function returns None.

        Example:
            # Turn off specific stores for a product
            product_data = turn_offs('884110754721', ['amazon_default', 'reedsdu_default'])
            
        """
        # Getting the product details for the SKU.
        self.product_data = self.get_product_details(sku)

        product_data = self.product_data
        website_ids = product_data.get('extension_attributes', {}).get('website_ids', [])
        exception_source = None
        for attr in product_data.get('custom_attributes', []):
            if attr['attribute_code'] == 'exception_source':
                exception_source = attr['value']
        # Fetch all stores to map IDs to store codes
        all_stores = self.get_all_stores()

        applied_store_codes = [all_stores.get(str(website_id), f"Unknown Store: {website_id}") for website_id in website_ids]
        print("All store codes applied to the current product:", ', '.join(applied_store_codes))
        if not archive:
                # Check if the specific store name is already applied
            store_name_to_check = "amazon_default"
            if store_name_to_check in applied_store_codes:
                print(f"The Amazon Store is applied to the product")
                store_codes.append('amazon_default')
            nonprofit_names_to_check = ['reedsdu_default','deltawaterfowldirect_default','rmefevents_default','muledeerevents_default','nwtfdistribution_default','reedsrgs_default','reedsnda_default','claytargetteamsupply_default','reedscca_default','dutrap_default','ducal_default','nwtfcal_default','npcal_default']

            for store_name in nonprofit_names_to_check:
                if store_name in applied_store_codes or exception_source is not None:
                    print(f"NP Stores Turned on: {store_name}")
                    raise StoreTurnOffRestrictedError(f"Unable to turn off stores because of {store_name}. The Exception Source is {exception_source}")
        # Check if store_codes are provided and are associated with the product
        # Get the store IDs corresponding to the provided store_codes
        store_ids = [int(key) for key, value in all_stores.items() if value in store_codes]
        if store_ids:
            endpoint = f"{self.login_controller.api_endpoint}/products/{sku}"
            product_stores = {
                "product": {
                    "sku": sku,
                    "extension_attributes": {
                        "website_ids": store_ids,
                        
                    },
                    "custom_attributes" : [
                        {
                        "attribute_code":"amazon_can_list",
                        "value":0
                        }
                    ]
                }
            }
            product_data = self.make_api_request(endpoint, request_type='put',json=product_stores)
            # Print both provided store_codes and actual store codes
            print(f"Provided store_codes: {', '.join(store_codes)}")
            # Process and return product data or perform other operations
            print("Store Codes Removed")
            return product_data

        else:
            print(f"Provided store_codes {', '.join(store_codes)} not found in actual store codes:", ', '.join(all_stores.values()))
            print("Those Stores are not assigned to that Product.")

    
    def update_product_stock_status(self, sku: str, is_in_stock:bool= True) -> dict:
        """
        Updates the stock status of a product.

        This method updates the stock status of a specified product SKU to either in stock or out of stock.

        Args:
            sku (str): The SKU of the product to update.
            is_in_stock (bool, optional): The desired stock status of the product. Defaults to True.

        Returns:
            dict: The API response containing the updated product data.

        Example:
            # Update the stock status of a product to out of stock
            updated_product = update_product_stock_status('884110754721', is_in_stock=False)
            print(updated_product)

        Notes:
            - This method constructs the necessary data payload to update the stock status and sends a PUT request to the product API endpoint.
            - The `is_in_stock` attribute is updated in the `stock_item` section of the `extension_attributes`.

        """

        product_data = {
            "product": {
                "sku": sku,
                "extension_attributes": {
                    "stock_item": {
                            "is_in_stock": is_in_stock,
                            "low_stock_date": None,
                            "stock_status_changed_auto": 1
                    }
                }
            }
        }

        endpoint = f"https://admin.reedssports.com/rest//V1/products/{sku}"
        return self.make_api_request(endpoint,request_type='put',data=product_data)
    def update_product_price(self, sku: str, new_price: float) -> dict:
        """
        Updates the price of a product in Magento 2.

        Args:
            sku (str): The SKU of the product to update.
            new_price (float): The new price to set for the product.

        Returns:
            dict: The API response containing the updated product data.

        Example:
            # Update the price of a product
            updated_product = update_product_price('884110754721', 99.99)
            print(updated_product)
        """
        product_data = {
            "product": {
                "sku": sku,
                "price": new_price
            }
        }

        endpoint = f"{self.login_controller.api_endpoint}/products/{sku}"
        return self.make_api_request(endpoint, request_type='put', data=product_data)
    def product_get_attributes(self, attribute_code:list) -> dict:
        attribute_dict = {}
        endpoint = f"{self.login_controller.api_endpoint}/products/attributes"
        filtergroup = FilterGroup()
        for attribute in attribute_code:
            filtergroup.add_filter('in', 'attribute_code', attribute)
        params = Params(filterGroups=filtergroup)
        attribute_codes_data = self.make_api_request(endpoint, request_type='get',params=params)
        for attribute_data in attribute_codes_data.get('items'):
            if attribute_data.get('attribute_code') in attribute_code:
                attribute_code = attribute_data.get('attribute_code')
                attribute_label = attribute_data.get('default_frontend_label')
                options = attribute_data.get('options')
                custom_attributes = attribute_data.get('custom_attributes')
                frontend_labels = attribute_data.get('frontend_labels')
                attribute_dict.update({
                    attribute_code: {
                    'attribute_label':attribute_label,
                    'options': options,
                    'custom_attributes': custom_attributes,
                    'frontend_labels': frontend_labels
                    }
                })

        return attribute_dict
    def product_search(self, skus: list) -> dict:
        """
        Search for product by sku.

        Args:
            skus (list): List of sku numbers of products
        """
        endpoint = f"{self.login_controller.api_endpoint}/products"
        product_data = []
        all_responses = []
        chunk_size = 500

        for i in range(0, len(skus), chunk_size):
            sku_chunk = skus[i:i + chunk_size]
            filtergroup = FilterGroup()
            filtergroup.add_filter('in', 'sku', ','.join(sku_chunk))
            params = Params(filterGroups=filtergroup)
            try:
                response = self.make_api_request(endpoint, params=params)
                response.raise_for_status()  # Raise HTTPError for bad responses (4xx and 5xx)

                product_chunk_data = response.json()
                all_responses.extend(product_chunk_data.get('items', []))
            except requests.exceptions.RequestException as e:
                print("An error occurred while searching for the product: %s", str(e))
                return None

        product_data = all_responses
        return product_data
    def save_to_excel(df, file_path):
        file_exists = os.path.exists(file_path)

        if file_exists:
            while True:
                try:
                    wb = load_workbook(file_path)
                    break  # If loading succeeds, break out of the loop
                except PermissionError as e:
                    print(f"Permission denied. Retrying in 5 seconds...")
                    time.sleep(5)
                except Exception as e:
                    print(f"Error loading workbook: {e}")
                    return

            with pd.ExcelWriter(file_path, engine='openpyxl', mode='a') as writer:
                try:
                    idx = writer.book.sheetnames.index('All_Orders')
                    writer.book.remove(writer.book.worksheets[idx])
                except ValueError:
                    pass  # 'All_Orders' sheet doesn't exist, no need to remove

                df.to_excel(writer, index=False, sheet_name='All_Orders', header=True)
        else:
            with pd.ExcelWriter(file_path, engine='openpyxl') as writer:
                df.to_excel(writer, index=False, sheet_name='All_Orders')
        
class OrderHandler(Magento):
    """
    A class to handle order-related operations using the Magento API.
    
    This class inherits from the Magento class and manages tasks like fetching orders,
    canceling orders, adding totals to order data, and saving the data to an Excel file.
    """

    def __init__(self, login_controller):
        """Initializes the OrderHandler with a login controller.
        
        Sets up threading events and locks, checks login status, and logs in if necessary.
        """
        self.error_event = threading.Event()  # Event to signal an error
        self.orders_lock = threading.Lock()   # Lock for synchronizing access to orders list
        super().__init__()
        self.type = 2
        self.login_controller = login_controller
        if not self.login_controller.is_logged_in():
            self.login_controller.login()

    def fetch_orders_page(self, api_endpoint, headers, params):
        """Fetches a single page of orders from the API.
        
        Sends a GET request to the orders endpoint and returns the response.
        """
        orders_endpoint = f"{api_endpoint}/orders"
        response = requests.get(orders_endpoint, headers=headers, params=params)
        return response

    def cancel_order(self, order_number: str) -> dict:
        """Attempts to cancel an order given its order number.
        
        Fetches order details, determines the order ID, and sends a POST request 
        to cancel the order. Raises an error if the order cannot be canceled.
        """
        order_data = self.get_order_details(order_number)
        order_id = order_data.get('items', [])[0].get("entity_id")
        orders_endpoint = f"{self.login_controller.api_endpoint}/orders/{order_id}/cancel"
        
        if self.make_api_request(orders_endpoint, request_type="post") == False:
            raise InvalidInputError(f"Order {order_number.rstrip()} Cannot Be Cancelled")
        else:
            return self.make_api_request(orders_endpoint, request_type="post")

    def add_totals(self, input_df):
        """Calculates the totals for specific numeric columns in the DataFrame.
        
        Sums up the numeric columns and appends a totals row to the DataFrame.
        """
        numeric_columns = ['SubTotal', 'Shipping and Handling', 'Total Due', 'Total Paid', 'Grand Total', 'Total Refunded']
        totals = input_df[numeric_columns].sum()

        # Create a dictionary for the totals row
        totals_row = {'Order#': '', 'Order Date': '', 'Customer Name': '',
                      'Carrier Type': '', 'Ship-to Name': '', 'Shipping Address': 'Total',
                      'SubTotal': int(totals['SubTotal']),
                      'Shipping and Handling': int(totals['Shipping and Handling']),
                      'Total Due': int(totals['Total Due']),
                      'Total Paid': int(totals['Total Paid']),
                      'Grand Total': int(totals['Grand Total']),
                      'Total Refunded': int(totals['Total Refunded']),
                      'PO Number': ''}

        # Append the totals row to the DataFrame
        df_with_totals = pd.concat([input_df, pd.DataFrame([totals_row])], ignore_index=True)
        return df_with_totals

    def save_to_excel(self, df, file_path):
        """Saves the DataFrame to an Excel file.
        
        If the file already exists, it attempts to open and modify the existing file.
        If not, it creates a new file. The data is written to a sheet named 'All_Orders'.
        """
        file_exists = os.path.exists(file_path)

        if file_exists:
            try:
                wb = load_workbook(file_path)
            except Exception as e:
                print(f"Error loading workbook: {e}")
                return

            with pd.ExcelWriter(file_path, engine='openpyxl', mode='a') as writer:
                try:
                    # Remove the 'All_Orders' sheet if it already exists
                    idx = writer.book.sheetnames.index('All_Orders')
                    writer.book.remove(writer.book.worksheets[idx])
                except ValueError:
                    pass  # 'All_Orders' sheet doesn't exist, no need to remove

                # Write the DataFrame to the 'All_Orders' sheet
                df.to_excel(writer, index=False, sheet_name='All_Orders', header=True)
        else:
            # Create a new Excel file and write the DataFrame to the 'All_Orders' sheet
            with pd.ExcelWriter(file_path, engine='openpyxl') as writer:
                df.to_excel(writer, index=False, sheet_name='All_Orders')
    def fetch_order_data(self, params: Params):
        endpoint = f"{self.login_controller.api_endpoint}/orders"
        return self.make_api_request(endpoint,request_type='get', params=params)

    def fetch_orders(self, headers, params: Params, orders):
        """
        Fetches orders from the API based on the provided headers and parameters.
        
        Args:
            headers (dict): The headers to include in the API request, typically containing authorization info.
            params (Params): The query parameters to include in the API request.
            orders (list): A list to store the fetched orders.
        
        This method sends an API request to fetch orders based on the given parameters.
        If the request is successful, the orders are processed and added to the shared orders list.
        If an error occurs, it logs the error and sets the error event.
        """
        # Fetch a page of orders from the API
        response = self.fetch_orders_page(self.login_controller.api_endpoint, headers, params)
        
        if response.status_code == 200:
            # If the request is successful, process and add orders to the list
            fetched_orders = self.process_orders_data(response.json())
            
            # Ensure thread-safe access to the orders list using a lock
            with self.orders_lock:
                orders.extend(fetched_orders)
        else:
            # Log the error if the request fails
            print(f"Error occurred while retrieving order data: {response.text}")
            
            # Set the error event to indicate an issue
            self.error_event.set()

    def Reeds_API_orders_Search(self, store, file_save_location, file_name, date=None, shipment_date=None, customer_email=None, customer_last=None, total_paid=None, total_due=None, status=None, page=1, page_size=500, excel_file_path=None, skus=None):
        """
        Searches for orders using the Reeds API and exports them to an Excel file.
        
        Args:
            store (str): The store identifier.
            access_token (str): The access token for API authentication.
            file_save_location (str): The directory where the Excel file will be saved.
            file_name (str): The name of the Excel file.
            date (tuple, optional): A date range for filtering orders.
            shipment_date (tuple, optional): A shipment date range for filtering orders.
            customer_email (str, optional): The customer's email address.
            customer_last (str, optional): The customer's last name.
            total_paid (float, optional): The total amount paid for filtering orders.
            total_due (float, optional): The total amount due for filtering orders.
            status (str, optional): The status of the orders for filtering.
            page (int, optional): The current page of results to retrieve. Defaults to 1.
            page_size (int, optional): The number of orders per page. Defaults to 500.
            excel_file_path (str, optional): The path to save the Excel file.
            skus (list, optional): A list of SKUs to filter orders by.
        
        This method retrieves orders from the Reeds API based on various filters,
        processes the results using multiple threads, and exports the orders to an Excel file.
        """
        headers = {"Authorization": f"Bearer {self.login_controller.token}"}
        orders = []

        # Construct initial query parameters based on provided filters
        params = self.construct_query_params(store, date, shipment_date, customer_email, customer_last, total_paid, total_due, status, page, 1, skus)
        
        # Fetch the first page of orders to determine the total count
        response = self.fetch_orders_page(self.login_controller.api_endpoint, headers, params)
        
        if response.status_code == 200:
            total_count = response.json().get('total_count', 0)
            if total_count > 0:
                print("Total Count:", total_count)
                threads = []
                
                while True:
                    # Construct parameters for the current page
                    params = self.construct_query_params(store, date, shipment_date, customer_email, customer_last, total_paid, total_due, status, page, page_size, skus)
                    
                    # Start a new thread to fetch orders for the current page
                    thread = threading.Thread(target=self.fetch_orders, args=(headers, params, orders))
                    threads.append(thread)
                    thread.start()

                    # Limit the number of active threads to avoid overloading the system
                    if len(threads) >= 5:
                        for thread in threads:
                            thread.join()  # Wait for threads to finish
                            if self.error_event.is_set():
                                print("An error occurred in one of the threads. Aborting.")
                                return
                        threads = []  # Reset the threads list after joining

                    # Check if there are more pages to fetch or if an error occurred
                    if not self.has_more_pages(page, total_count, page_size) or self.error_event.is_set():
                        break

                    # Move to the next page
                    page += 1

                # Join any remaining threads
                for thread in threads:
                    thread.join()

                # Export the collected orders to an Excel file after all threads finish
                self.export_orders_to_excel(orders, file_save_location, file_name, date)
            else:
                print("No items found matching the search criteria.")
        else:
            # Log the error if the initial request fails
            print(f"Error occurred while retrieving order data: {response.text}")

    def construct_query_params(self, store, date, shipment_date, customer_email, customer_last, total_paid, total_due, status, page, page_size, skus):
        """
        Constructs the query parameters for the API request based on the provided filters.
        
        Args:
            store (str): The store identifier.
            date (tuple, optional): A date range for filtering orders.
            shipment_date (tuple, optional): A shipment date range for filtering orders.
            customer_email (str, optional): The customer's email address.
            customer_last (str, optional): The customer's last name.
            total_paid (float, optional): The total amount paid for filtering orders.
            total_due (float, optional): The total amount due for filtering orders.
            status (str, optional): The status of the orders for filtering.
            page (int): The current page of results to retrieve.
            page_size (int): The number of orders per page.
            skus (list, optional): A list of SKUs to filter orders by.
        
        Returns:
            dict: A dictionary of query parameters to be used in the API request.
        
        This method builds and returns a dictionary of parameters that will be used 
        to filter and paginate the orders retrieved from the API.
        """
        # Convert the provided date range into datetime objects for filtering
        date1 = datetime.strptime(date[0], "%Y-%m-%d")
        date2 = datetime.strptime(date[1], "%Y-%m-%d")
        
        # Adjust the date range by one day before and after the given dates
        date1 = date1 + timedelta(days=-1)
        date2 = date2 + timedelta(days=1)

        # Initialize the filter group for the query
        filtergroup = FilterGroup()
        filtergroup.add_filter("eq", customer_email, 'customer_email')
        filtergroup.add_filter("",)  # This seems to be a placeholder; may need further refinement

        # Set up the initial parameters including pagination settings
        params = {
            'searchCriteria[pageSize]': page_size,
            'searchCriteria[currentPage]': int(page)  # Set the current page number
        }
        
        # Add filters to the parameters based on the provided values
        if customer_email:
            params['searchCriteria[filterGroups][1][filters][0][field]'] = 'customer_email'
            params['searchCriteria[filterGroups][1][filters][0][value]'] = customer_email
            params['searchCriteria[filterGroups][1][filters][0][conditionType]'] = 'eq'

        if date:
            params['searchCriteria[filterGroups][2][filters][0][field]'] = 'created_at'
            params['searchCriteria[filterGroups][2][filters][0][value]'] = date1
            params['searchCriteria[filterGroups][2][filters][0][conditionType]'] = 'from'
            params['searchCriteria[filterGroups][3][filters][0][field]'] = 'created_at'
            params['searchCriteria[filterGroups][3][filters][0][value]'] = date2
            params['searchCriteria[filterGroups][3][filters][0][conditionType]'] = 'to'

        count = 4  # Used to keep track of the filter group index

        if customer_last:
            params[f'searchCriteria[filterGroups][{count}][filters][0][field]'] = 'customer_lastname'
            params[f'searchCriteria[filterGroups][{count}][filters][0][value]'] = customer_last
            params[f'searchCriteria[filterGroups][{count}][filters][0][conditionType]'] = 'eq'
            count += 1

        if shipment_date is not None:
            params[f'searchCriteria[filterGroups][{count}][filters][0][field]'] = 'shipment_date'
            params[f'searchCriteria[filterGroups][{count}][filters][0][value]'] = shipment_date[0]
            params[f'searchCriteria[filterGroups][{count}][filters][0][conditionType]'] = 'from'
            count += 1
            params[f'searchCriteria[filterGroups][{count}][filters][0][field]'] = 'dispatch_date'
            params[f'searchCriteria[filterGroups][{count}][filters][0][value]'] = shipment_date[1]
            params[f'searchCriteria[filterGroups][{count}][filters][0][conditionType]'] = 'to'
            count += 1

        return params





    def process_orders_data(self, order_data):
        """
        Processes order data to extract and format relevant order information.

        Args:
            order_data (dict): A dictionary containing orders and their details.

        Returns:
            list: A list of dictionaries, each representing a processed order.
        """
        order_list = []  # Initialize a list to store individual orders (unused in current code)
        orders = []  # Initialize a list to store the processed order dictionaries
        item_list = []  # Initialize a list to store items (unused in current code)

        # Iterate through each order in the order_data
        for order in order_data.get('items', []):
            # Extract order_id from each item in the order
            for item in order.get('items', []):
                order_id = item.get('order_id')  # Extract order_id from items

            # Extract the shipping address from order extension attributes
            shipping_address = order.get('extension_attributes', {}).get('shipping_assignments', [{}])[0].get('shipping', {}).get('address', {})

            # Extract item descriptions and quantities, filtering by product type
            item_descriptions_bundle = [item.get('name', '') for item in order.get('items', []) if item.get('product_type') == 'bundle']
            item_descriptions_simple = [item.get('name', '') for item in order.get('items', []) if item.get('product_type') == 'simple']
            item_qty = [item.get('qty_ordered', '0') for item in order.get('items', []) if item.get('product_type') == 'simple']
            print(item_qty)  # Print quantities for debugging

            # Extract item barcodes and shipped quantities, filtering by product type
            item_barcodes_bundle = [item.get('sku', '') for item in order.get('items', []) if item.get('product_type') == 'bundle']
            item_barcodes_simple = [item.get('sku', '') for item in order.get('items', []) if item.get('product_type') == 'simple']
            item_shipped_simple_qty = [item.get('qty_shipped', '') for item in order.get('items', []) if item.get('product_type') == 'simple']
            item_shipped_bundle_qty = [item.get('qty_shipped', '') for item in order.get('items', []) if item.get('product_type') == 'bundle']

            # Filter out blank or specific barcodes (e.g., '400140098286')
            non_blank_barcodes_descriptions_bundle = [desc for desc in item_barcodes_bundle if desc != '400140098286']
            non_blank_barcodes_descriptions_simple = [desc for desc in item_barcodes_simple if desc != '400140098286']

            # Create a dictionary for each order with relevant details
            order_entry = {
                'Order#': order.get('increment_id'),  # Order increment ID
                'Order Date': order.get('created_at'),  # Date the order was created
                'Customer Name': f"{order.get('customer_firstname')} {order.get('customer_lastname')}",  # Customer full name
                'Store Name': order.get('store_name', ''),  # Name of the store
                'Customer Email': shipping_address.get('email', ''),  # Customer's email address
                'Carrier Type': order.get('shipping_description'),  # Description of the shipping carrier
                'Ship-to Name': f"{shipping_address.get('firstname', '')} {shipping_address.get('lastname', '')}",  # Full name of the recipient
                'Shipping Address': f"{shipping_address.get('street', [''])[0]}, {shipping_address.get('city', '')}, {shipping_address.get('region', '')}, {shipping_address.get('postcode', '')}",  # Complete shipping address
                'SubTotal': int(order.get('subtotal', 0)),  # Order subtotal
                'Shipping and Handling': int(order.get('shipping_amount', 0)),  # Shipping and handling fees
                'Total Due': order.get('total_due', 0),  # Total amount due
                'Total Paid': order.get('total_paid', 0),  # Total amount paid
                'Grand Total': order.get('grand_total', 0),  # Grand total of the order
                'Total Refunded': order.get('total_refunded', 0),  # Total amount refunded
                'Bundle QTY Shipped': item_shipped_bundle_qty,  # Quantities of bundle items shipped
                'Simple QTY Shipped': item_shipped_simple_qty,  # Quantities of simple items shipped
                'PO Number': order.get('payment', {}).get('po_number', ''),  # Purchase order number
                'Simple Items': ', '.join(item_descriptions_simple),  # Comma-separated list of simple item descriptions
                'Bundle Items': ', '.join(item_descriptions_bundle),  # Comma-separated list of bundle item descriptions
                'Status': order.get('status'),  # Order status
                'Number of Items Bundle': len(non_blank_barcodes_descriptions_bundle),  # Number of unique bundle items
                'Number of Items Simple': len(non_blank_barcodes_descriptions_simple),  # Number of unique simple items
                'Barcode Bundle': ', '.join(non_blank_barcodes_descriptions_bundle),  # Comma-separated list of bundle item barcodes
                'Barcode Simple': ', '.join(non_blank_barcodes_descriptions_simple),  # Comma-separated list of simple item barcodes
                'QTY': ', '.join(str(qty) for qty in item_qty)  # Comma-separated list of quantities for simple items
                # 'Tracking': ', '.join(tracking_numbers),  # (Commented out) Comma-separated list of tracking numbers
                # 'Shipped Date': ', '.join(shipment_date)  # (Commented out) Comma-separated list of shipment dates
            }

            # Add the processed order entry to the orders list
            orders.append(order_entry)
        
        return orders  # Return the list of processed orders

    def has_more_pages(self, page, total_count, page_size):
        """
        Determine if there are more pages of data to fetch.

        Args:
            page (int): The current page number.
            total_count (int): The total number of items available.
            page_size (int): The number of items per page.

        Returns:
            bool: True if there are more pages, False otherwise.
        """
        total_pages = math.ceil(total_count / page_size)  # Calculate total pages based on total count and page size
        return page < total_pages  # Return True if the current page is less than the total number of pages

    def export_orders_to_excel(self, orders, file_save_location, file_name, date):
        """
        Export the processed orders data to an Excel file.

        Args:
            orders (list): A list of dictionaries representing the processed orders.
            file_save_location (str): The directory where the Excel file will be saved.
            file_name (str): The name of the Excel file.
            date (list or None): Optional date used to format the file name.

        Returns:
            str: The file path of the exported Excel file.
        """
        df = pd.DataFrame(orders)  # Convert the orders list to a DataFrame
        df['Order Date'] = pd.to_datetime(df['Order Date'])  # Convert 'Order Date' column to datetime format

        df.sort_values(by='Order Date', ascending=True, inplace=True)  # Sort DataFrame by 'Order Date'

        df = self.add_totals(df)  # Add totals to the DataFrame using a custom method

        # Define the formatting options for financial columns
        format_mapping_totals = {
            'SubTotal': '${:,.2f}',
            'Shipping and Handling': '${:,.2f}',
            'Total Due': '${:,.2f}',
            'Total Paid': '${:,.2f}',
            'Grand Total': '${:,.2f}',
            'Total Refunded': '${:,.2f}'
        }

        # Apply formatting to the specified columns
        for column, fmt in format_mapping_totals.items():
            df[column] = df[column].apply(fmt.format)

        # Determine the file name and path based on the provided date or current date
        if date is not None:
            update_date = datetime.strptime(date[0], '%Y-%m-%d')  # Parse the provided date
            formatted_date = update_date.strftime('%m.%Y')  # Format the date as 'MM.YYYY'
        else:
            current_date = datetime.now()  # Get the current date
            formatted_date = current_date.strftime('%m.%Y')  # Format the current date as 'MM.YYYY'

        excel_file_path = f"{file_save_location}/{formatted_date} {file_name}.xlsx"  # Generate the full file path

        self.save_to_excel(df, excel_file_path)  # Save the DataFrame to an Excel file

        print(f'Orders exported to {excel_file_path} successfully.')  # Print a confirmation message
        return excel_file_path  # Return the file path of the exported Excel file
                
    
    # def Reeds_API_orders_Search(self, store, access_token, file_save_location, file_name, date=None,shipment_date=None, customer_email=None, customer_last=None,total_paid=None,total_due=None,status=None, page=1, page_size=500, excel_file_path=None, skus:list=None):

    #     print(f"Fetching page {page}...")
    #     headers = {"Authorization": f"Bearer {self.login_controller.token}"}

    #     response = self.fetch_orders_page(self.login_controller.api_endpoint, headers, params)
    #     print(response.text)
    #     if response.status_code == 200 and not response.json()['total_count'] == 0:
    #         # Check if there are more pages and fetch them
    #         # Calculate total pages based on total count and page size
    #         total_count = response.json().get('total_count', 0)
    #         total_pages = math.ceil(total_count / page_size)
    #         if page < total_pages:
    #             self.Reeds_API_orders_Search(store, access_token, file_save_location, file_name, page=page+1, page_size=page_size, date=date, shipment_date=shipment_date, customer_email=customer_email, customer_last=customer_last, total_due=total_due, status=status)
            
    #         print(page, total_count)
    #         if not page < total_pages:
    #             df = pd.DataFrame(orders)

    #             df = self.add_totals(df)
    #             # Define the formatting options
    #             format_mapping_totals = {
    #                 'SubTotal': '${:,.2f}',
    #                 'Shipping and Handling': '${:,.2f}',
    #                 'Total Due': '${:,.2f}',
    #                 'Total Paid': '${:,.2f}',
    #                 'Grand Total': '${:,.2f}',
    #                 'Total Refunded': '${:,.2f}'
    #             }

    #             # Apply formatting to the columns
    #             for column, fmt in format_mapping_totals.items():
    #                 df[column] = df[column].apply(fmt.format)
    #             if excel_file_path is None:
    #                 # Export DataFrame to Excel file
    #                 if date is not None:
    #                     update_date = datetime.strptime(date[0], '%Y-%m-%d')
    #                     formatted_date = update_date.strftime('%m.%Y')
    #                 else:
    #                     # Get the current date and format it to display month and year
    #                     current_date = datetime.now()
    #                     formatted_date = current_date.strftime('%m.%Y')

    #             excel_file_path = f"{file_save_location}/{formatted_date} {file_name}.xlsx"

    #             self.save_to_excel(df, excel_file_path)

    #             print(f'Orders exported to {excel_file_path} successfully.')
    #             print(total_pages)
    #             return order_list

    #         elif response.json()['total_count'] == 0:
    #             print("No items found matching the search criteria.")
    #         else:
    #             error_response = response.json()
    #             print(f"Error occurred while retrieving order data: {error_response['message']}")
    #             return ValueError

    def get_order_details(self, increment_id):
        if not self.login_controller.is_logged_in():
            raise InvalidCredentialsError("Please log in to access product details.")

        endpoint = f"{self.login_controller.api_endpoint}/orders?searchCriteria[filterGroups][0][filters][0][field]=increment_id&searchCriteria[filterGroups][0][filters][0][value]={increment_id}"  # Endpoint to filter by increment_id
        headers = {"Authorization": f"Bearer {self.login_controller.token}"}

        response = requests.get(endpoint, headers=headers)
        if response.status_code == requests.codes.ok and response.json().get('total_count') > 0:
            order_data = response.json()
            # Process and return product data or perform other operations
            return order_data
        else:
            raise InvalidInputError(f"Order Number is invalid: {increment_id}")
    def save_details(self, data):
        return super().save_details(data,2)
    def add_comment_to_order(self, order_number, comment):
        """
        Add a comment/note to the specified order.

        Args:
            order_number (str): Order number to add the comment to
            comment (str): Comment/note to be added to the order

        Returns:
            dict or None: Response data from the comment addition request or None in case of failure
        """
        if not self.login_controller.is_logged_in():
            raise InvalidCredentialsError("Please log in to access product details.")
        try:
            try:
                order_data = self.get_order_details(order_number)
                status = order_data.get('items')[0].get('status')
                order_id = order_data.get('items', [])[0].get("entity_id")

            except (TypeError, AttributeError, IndexError):
                order_id = None

            # Prepare data for adding comment to the order
            comment_data = {
                "statusHistory": {
                    "comment": f"{comment}",
                    "is_customer_notified": 0,
                    "is_visible_on_front": 0,
                    "status": status
                }
            }

            if not order_id is None:

                # API endpoint to add comment to order
                update_endpoint = f"{self.login_controller.api_endpoint}/orders/{order_id}/comments"
                
                # Make a POST request to add the comment
                header = {"Authorization": f"Bearer {self.login_controller.token}"}

                response = requests.post(update_endpoint, headers=header, json=comment_data)
                response.raise_for_status()  # Raise HTTPError for bad responses (4xx and 5xx)
                
                return f"Order Comment Has Been Added To {order_number}"
        except requests.exceptions.RequestException as e:
            raise requests.exceptions.RequestException(f"An error occurred while adding a comment to order {order_number}: {str(e)}")
    def create_order(self,entity: Entity) -> dict:
        endpoint = f"{self.login_controller.api_endpoint}/orders"
        return self.make_api_request(endpoint,request_type='post',data=entity.to_json())
    def get_comments(self, order_id: str) -> dict:
        endpoint = f"{self.login_controller.api_endpoint}/orders/{order_id}/comments"
        return self.make_api_request(endpoint,request_type='get')
    def order_search(self, increment_ids: list = None, order_numbers: list = None):
        """
        Retrieves order information by order numbers or increment IDs.

        Parameters:
        - increment_ids (list): List of increment IDs for orders to search.
        - order_numbers (list): List of order numbers (entity_id) to search.

        Returns:
        - order_info (list): A list of dictionaries containing order details like:
            - 'Order#'
            - 'Order_ID'
            - 'Customer Name'
            - 'PO'
            - 'Status'
            - 'Ship To Name'
            - 'City'
            - 'State'
            - 'Street'
            - 'Total Shipping'
        """
        orders_endpoint = f"{self.login_controller.api_endpoint}/orders"
        order_info = []
        all_responses = []
        chunk_size = 500

        # Retry parameters
        max_retries = 3
        retry_delay = 30  # seconds between retries

        def fetch_orders(search_field, ids):
            """
            Helper function to make API calls in chunks with retries.

            Parameters:
            - search_field (str): The search field (e.g., 'entity_id' or 'increment_id').
            - ids (list): List of IDs (order numbers or increment IDs) to search.

            Returns:
            - None. Results are added to `all_responses`.
            """
            for i in range(0, len(ids), chunk_size):
                retries = 0
                while retries < max_retries:
                    chunk = ids[i:i + chunk_size]
                    filtergroup = FilterGroup()
                    filtergroup.add_filter('in', 'entity_id', ','.join(map(str, chunk)))
                    params = Params(filterGroups=filtergroup)

                    try:
                        response = self.make_api_request(orders_endpoint, params=params)
                        print(f"Request {i} success - Status: {response.status_code}")
                        all_responses.extend(response.get('items', []))

                    except requests.RequestException as e:
                        print(f"Request {i} failed: {e}, Retrying...")

                    retries += 1
                    time.sleep(retry_delay)

                else:
                    # If max retries are reached without success, log and exit
                    print(f"Max retries reached for chunk {i}. Could not retrieve order data.")
                    exit()

        # Fetch by order numbers if provided
        if order_numbers:
            fetch_orders('entity_id', order_numbers)

        # Fetch by increment IDs if provided
        if increment_ids:
            fetch_orders('increment_id', increment_ids)

        # Process the responses and construct the order information
        for order in all_responses:
            try:
                shipping = order.get('extension_attributes', {}).get('shipping_assignments', [])[0].get('shipping', {}).get('address', {})
                payment_po = order.get('payment', {}).get('po_number', None)

                order_entry = {
                    'Order#': order.get('increment_id'),
                    'Order_ID': order.get('entity_id'),
                    'Customer Name': f"{order.get('customer_firstname', '')} {order.get('customer_lastname', '')}",
                    'PO': payment_po,
                    'Status': order.get('status'),
                    'Ship To Name': f"{shipping.get('firstname', '')} {shipping.get('lastname', '')}",
                    'City': shipping.get('city', ''),
                    'State': shipping.get('region_code', ''),
                    'Street': shipping.get('street', [''])[0],
                    'Total Shipping': shipping.get('total', {}).get('base_shipping_amount', 0)
                }

                order_info.append(order_entry)

            except (IndexError, AttributeError) as e:
                print(f"Error processing order {order.get('increment_id')}: {e}")

        return order_info

class ShipmentHandler(Magento):
    def __init__(self, login_controller):
        self.error_event = threading.Event()  # Event to signal an error
        self.shipment_lock = threading.Lock()    # Lock for synchronizing access to orders list
        super().__init__()
        self.type = 3
        self.login_controller = login_controller
        if not self.login_controller.is_logged_in():
            self.login_controller.login()
    # def search_shipments(self,store, file_save_location, file_name,order_list=None, date=None, excel_file_path=None, page=1, page_size=500):
    #     all_shipments = []   # Use a global variable to accumulate shipments across all pages
    #     all_responses = []
    #     order_ids_list = []
    #     sku_attribute_mapping = {}
    #     bundle_sku_mapping = {}
    #     shipments_endpoint = f"{self.login_controller.api_endpoint}/shipments"
    #     magento = OrderHandler(self.login_controller)
    #     chunk_size = 500
    #     if order_list:
    #         # Call the order_search function to obtain order information
    #         order_info_entries = magento.get_order_detailsorder_search(order_list)

    #         order_info_mapping = {order_info_entry.get('Order_ID'): order_info_entry for order_info_entry in order_info_entries}
            
    #         # Convert dictionary keys to a list and iterate in chunks
    #         order_keys = list(order_info_mapping.keys())
    #         for i in range(0, len(order_keys), chunk_size):
    #             order_chunk = order_keys[i:i + chunk_size]  
    #             params = {
    #                 'searchCriteria[filterGroups][0][filters][0][field]': 'order_id',
    #                 'searchCriteria[filterGroups][0][filters][0][value]': ','.join(map(str, order_chunk)),
    #                 'searchCriteria[filterGroups][0][filters][0][conditionType]': 'in',
    #                 'searchCriteria[pageSize]': page_size,
    #                 'searchCriteria[currentPage]': page  # Set the page number
    #             }
    #             print("Order Chunk:", order_chunk, "\n\n\n\n")
    #             response = requests.get(shipments_endpoint, headers=headers, params=params)
    #             all_responses.extend(response.json().get('items', []))
    #             print(all_responses)
    #     else:
    #         # Initialize params outside the loop
    #         params = {
    #             'searchCriteria[filterGroups][0][filters][0][field]': 'store_id',
    #             'searchCriteria[filterGroups][0][filters][0][value]': store,
    #             'searchCriteria[filterGroups][0][filters][0][conditionType]': 'eq',
    #             'searchCriteria[pageSize]': page_size,
    #             'searchCriteria[currentPage]': page  # Set the page number
    #         }

    #         if date:
    #             params['searchCriteria[filterGroups][1][filters][0][field]'] = 'created_at'
    #             params['searchCriteria[filterGroups][1][filters][0][value]'] = datetime.strptime(date[0], '%Y-%m-%d').isoformat()
    #             params['searchCriteria[filterGroups][1][filters][0][conditionType]'] = 'from'
    #             params['searchCriteria[filterGroups][2][filters][0][field]'] = 'created_at'
    #             params['searchCriteria[filterGroups][2][filters][0][value]'] = datetime.strptime(date[1], '%Y-%m-%d').isoformat()
    #             params['searchCriteria[filterGroups][2][filters][0][conditionType]'] = 'to'

    #         # Initial request outside the loop
    #         response = requests.get(shipments_endpoint, headers=headers, params=params)
    #         total_count = response.json().get('total_count', 0)
    #         total_pages = math.ceil(total_count / page_size)
    #         print(total_pages)
    #         all_responses.extend(response.json().get('items', []))
    #         print(f"Fetching page {page}...")

    #         page += 1

    #         while page <= total_pages:  # Change condition to include the last page
    #             # Update only the necessary parameters inside the loop
    #             params['searchCriteria[currentPage]'] = page
    #             response = requests.get(shipments_endpoint, headers=headers, params=params)
    #             all_responses.extend(response.json().get('items', []))
    #             print(f"Fetching page {page}...")
    #             page += 1

        
    #     if not all_responses is None:
    #     # Extract order IDs from the shipment data
    #         if order_list is None:
                
    #             for response in all_responses:
    #                 order_ids_list.append(response.get('order_id'))
    #             # Call the order_search function to obtain order information
    #             order_info_entries = magento.order_search(order_numbers=order_ids_list)
    #             order_info_mapping = {order_info_entry.get('Order_ID'): order_info_entry for order_info_entry in order_info_entries}
    #         # Process the retrieved shipments data and filter by firstname
    #         shipments = []
    #         count = 0
    #         type(all_responses)
    #         sku_list = []
    #         for shipment in all_responses:
    #             for item in shipment.get('items'):
    #                 if item.get('sku') in nwtf_sku_mapping:
    #                     sku = nwtf_sku_mapping.get(item.get('sku')) 
    #                 else:
    #                     sku = item.get('sku')

    #                 sku_list.append(sku)

    #         all_product_responses = magento.product_search(sku_list)
    #         brand = ""
    #         style = ""
    #         attribute_data = magento.product_get_attributes(['brand'])

    #         for product in all_product_responses:
    #             # Extract SKU and attribute_set_id from the returned data

    #             sku = product.get('sku', '')
    #             for custom_attribute in product.get('custom_attributes'):
    #                 if custom_attribute.get('attribute_code') == 'style':
    #                     style = custom_attribute.get('value')
    #                 if custom_attribute.get('attribute_code') == 'brand':
    #                     options = attribute_data.get('brand').get('options')
    #                     for attribute in options:
    #                         if attribute.get('value') == custom_attribute.get('value'):
    #                             print(attribute.get('label'))
    #                             brand = attribute.get('label','')
    #             attribute_set_id = product.get('attribute_set_id', '')
    #             if product.get('type_id') == 'bundle':
    #                 bundle_sku_mapping[sku] = True

    #             # SKU already exists, update the mapping based on the attribute_set_id
    #             if attribute_set_id == 24:
    #                 sku_attribute_mapping[sku] = {
    #                     'Type': "Firearm",
    #                     'brand': brand,
    #                     'style': style                          
    #                                             }
    #                 print("Set Firearm")
    #             else:
    #                 sku_attribute_mapping[sku] = {
    #                     'Type': "Merch",
    #                     'brand': brand,
    #                     'style': style                          
    #                                             }
    #         auth_token = fedex_lookup.auth()
    #         tracking_number_list = []
    #         for shipment in all_responses:
    #             tracking_number_list.extend(track['track_number'] for track in shipment.get('tracks', '') if track.get('track_number'))
    #         tracking_number_info = fedex_lookup.track_package(auth_token,tracking_number_list)

    #         for shipment in all_responses:
    #             order_id = shipment.get('order_id')
    #             order_info_entry = order_info_mapping.get(order_id)
    #             tracking_numbers = [track['track_number'] for track in shipment.get('tracks', [])]

    #             for item in shipment.get('items'):
    #                 if order_info_entry and not item.get('sku') in drop_sku_list:
    #                     if item.get('sku') in nwtf_sku_mapping:
    #                         sku = nwtf_sku_mapping.get(item.get('sku')) 

    #                     else:
    #                         sku = item.get('sku')
    #                     weight = tracking_number_info.get(tracking_numbers[0], {}).get('weight','')
    #                     length = tracking_number_info.get(tracking_numbers[0], {}).get('length','')
    #                     width = tracking_number_info.get(tracking_numbers[0], {}).get('width','')
    #                     height = tracking_number_info.get(tracking_numbers[0], {}).get('height','')
    #                     shipment_type = tracking_number_info.get(tracking_numbers[0], {}).get('Shipment Type','')

    #                     associated_shipments = tracking_number_info.get(tracking_numbers[0], {}).get('associated_shipments','')
    #                     master_tracking_number =  tracking_number_info.get(tracking_numbers[0], {}).get('master_tracking_number','')
    #                     print(sku_attribute_mapping)
    #                     item_entry = {
    #                         'Order#': order_info_entry.get('Order#'),
    #                         'Shipment Date': shipment.get('created_at'),
    #                         'Customer Name': order_info_entry.get('Customer Name'),
    #                         'PO': order_info_entry.get('PO'),
    #                         'Status': order_info_entry.get('Status'),
    #                         'Master Pack': packing_array.get(item.get('sku', ''), ''),
    #                         'Product Name': item.get('name', ''),
    #                         'SKU': sku,
    #                         'Brand': sku_attribute_mapping.get(sku,{}).get('brand',''),
    #                         'Style': sku_attribute_mapping.get(sku,{}).get('style',''),
    #                         'Quantity': item.get('qty', 0),
    #                         'Price': item.get('price', 0),
    #                         # Add other relevant fields as needed
    #                         'Tracking Numbers': ', '.join(tracking_numbers) if tracking_numbers else '' ', '.join(tracking_numbers) if tracking_numbers else '',
    #                         'Master Tracking Number': master_tracking_number,
    #                         'Shipment Type': shipment_type,
    #                         'Shipping Cost': '',
    #                         'FedEx Shipping Vendor Tier Delivery Cost': '',
    #                         'Total Shipping Per Order': '',
    #                         'Weight (LB)': weight,
    #                         'Length (IN)': length,
    #                         'Width (IN)': width,
    #                         'Height (IN)': height,
    #                         'Ship To Name': order_info_entry.get('Ship To Name'),
    #                         'City': order_info_entry.get('City'),
    #                         'State':order_info_entry.get('State'),
    #                         'Street': order_info_entry.get('Street'),
    #                         'Total Shipping': order_info_entry.get('Total Shipping'),
    #                         'Type': sku_attribute_mapping.get(sku,{}).get('Type',''),
    #                         'Associated Shipments': associated_shipments,
    #                         'Picking Fee: Item 1': '',
    #                         'Additional Picks for Same Order':'',
    #                         'Carton Cost of Order':'',
    #                         'Total Fulfillment Costs, except for Shipping Cost': ''
                            
    #                     }
    #                     count += 1
    #                     all_shipments.append(item_entry)
    #                 else:
    #                     print(f"Order information not found for order ID: {order_id}")


    #     # Export DataFrame to Excel file after processing the last page
    #     df = pd.DataFrame(all_shipments)
    #     # Replace 'Master Pack' with the actual column header you are searching for
    #     column_to_search = 'Master Pack'

    #     # Check if the column exists in the DataFrame
    #     if column_to_search in df.columns:
    #         # Check if all values in the column are empty strings
    #         if (df[column_to_search] == '').all():
    #             # Remove the column
    #             df = df.drop(columns=[column_to_search])
    #     # Define the formatting options
    #     format_mapping_totals = {
    #         'SubTotal': '${:,.2f}',
    #         'Shipping and Handling': '${:,.2f}',
    #         'Total Due': '${:,.2f}',
    #         'Total Paid': '${:,.2f}',
    #         'Grand Total': '${:,.2f}',
    #         'Total Refunded': '${:,.2f}'
    #     }

    #     if excel_file_path is None:
    #         # Export DataFrame to Excel file
    #         if date is not None:
    #             update_date = datetime.strptime(date[0], '%Y-%m-%d')
    #             formatted_date = update_date.strftime('%m.%Y')
    #         else:
    #             # Get the current date and format it to display month and year
    #             current_date = datetime.now()
    #             formatted_date = current_date.strftime('%m.%Y')

    #         excel_file_path = f"{file_save_location}/{formatted_date} {file_name}.xlsx"

    #     magento.save_to_excel(df, excel_file_path)

    #     print(f'Shipments exported to {excel_file_path} successfully.')

    #     return all_shipments
    
if __name__ == "__main__":
    logincontroller = LoginController()
    headers = {
        "Authorization": f"Bearer {logincontroller.token}",
        "Accept": "application/json",
        "Content-Type" : "application/json"
        }
    orderhandler = OrderHandler(logincontroller)
    order_number = "65000000212"
    filtergroup = FilterGroup()
    filtergroup.add_filter('eq', 'entity_id', order_number)
    params = Params(filterGroups=filtergroup)
    print(params.to_dict())
    orderhandler.fetch_order_data(params)
    orderhandler.get_comments()
    
