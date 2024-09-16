import openpyxl


class AhmExcel:
    '''
    一系列处理Excel的方法
    '''

    @classmethod
    def json_to_excel(cls, json_data: list, excel_path: str, title: str = '数据') -> None:
        '''
        将json数据写入excel文件
        :param json_data: json数据
        :param excel_path: excel文件路径
        :return: None
        '''

        # 验证json数据
        cls._validate_json_data(json_data)

        # 将key设置为表头
        workbook = openpyxl.Workbook()
        sheet = workbook.active

        # 写入表头
        keys = list(json_data[0].keys())
        sheet.append(keys)

        # 写入数据
        for data in json_data:
            sheet.append([data[key] for key in keys])

        # 设置标题
        # 插入一行
        sheet.insert_rows(1)
        # 合并单元格
        sheet.merge_cells(start_row=1, start_column=1,
                          end_row=1, end_column=len(keys))
        # 设置数据
        sheet['A1'] = title
        # 设置样式
        sheet['A1'].alignment = openpyxl.styles.Alignment(
            horizontal='center', vertical='center')
        sheet['A1'].font = openpyxl.styles.Font(
            size=16, bold=True)

        # 边框加粗
        for row in sheet.iter_rows(min_row=1, max_row=len(json_data) + 2, max_col=len(keys)):
            for cell in row:
                cell.border = openpyxl.styles.Border(
                    left=openpyxl.styles.Side(style='thin'),
                    right=openpyxl.styles.Side(
                        style='thin'),
                    top=openpyxl.styles.Side(
                        style='thin'),
                    bottom=openpyxl.styles.Side(style='thin')
                )

        # 保存文件
        workbook.save(excel_path)
        print(f"excel文件已保存至 {excel_path}")

    @classmethod
    def _validate_json_data(cls, json_data: list):
        '''
        验证json数据是否为扁平结构
        '''
        if not isinstance(json_data, list):
            raise TypeError("json_data must be a list")

        # 检查子元素是否为字典
        for data in json_data:
            if not isinstance(data, dict):
                raise TypeError("child element of json_data must be a dict")

        # 检查子元素是否有相同的key
        keys = list(json_data[0].keys())
        for data in json_data:
            if keys != list(data.keys()):
                raise ValueError(
                    "child element of json_data must have the same keys")


if __name__ == '__main__':
    data_dict = [
        {'name': '张三', 'age': 18, },
        {'name': '李四', 'ages': 20, },
    ]
    AhmExcel.json_to_excel(data_dict, '测试.xlsx')
