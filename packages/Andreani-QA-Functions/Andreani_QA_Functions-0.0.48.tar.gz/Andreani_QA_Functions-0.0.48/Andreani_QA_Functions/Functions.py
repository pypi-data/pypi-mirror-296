# -*- coding: utf-8 -*-
import json
import pprint
import datetime
import platform
import unittest
import cx_Oracle
import pymsteams
import pyodbc
import time
import openpyxl
from openpyxl import Workbook
from openpyxl.utils import get_column_letter
import os
import smtplib
import base64
from email.mime.multipart import MIMEMultipart
from email.mime.text import MIMEText
from email.mime.base import MIMEBase
from email import encoders
from shareplum import Site, Office365
from shareplum.site import Version
import allure
import random
import pymongo
from pymongo.errors import ServerSelectionTimeoutError
from pymongo.errors import ConnectionFailure
from Andreani_QA_parameters.Parameters import Parameters
from cryptography.fernet import Fernet
import xml.etree.ElementTree as Et
import bz2
from typing import Any
# Loggin System
import logging
import urllib3

PATH_ORIGIN = os.path.join(Parameters.current_path, './environment_access.txt')
PATH_ORIGIN_XML = os.path.join(Parameters.current_path, './environment_access.xml')
PATH_TARGET = os.path.join(Parameters.current_path, './environment_access_e.txt')
ENVIRONMENT_VAR = 'PYBOT_KEY'
RED = '\033[31m'
BLUE = '\033[34m'
YELLOW = '\033[33m'
GREEN = '\033[32m'
DEFAULT = '\033[39m'


class Functions(Parameters):
    global_date = time.strftime(Parameters.date_format)  # formato aaaa/mm/dd
    global_time = time.strftime(Parameters.time_format)  # formato 24 houras
    project_name = None
    class_name = None
    case_name = None
    test_case_name = None
    file_name = None
    teams = None
    data_cache = {}
    data_resource = None
    path_downloads = None
    path_evidences = None
    path_files = None
    path_images = None
    path_outputs = None
    path_steps = None
    path_json = None
    path_resources = None
    path_map = None
    path_jmeter_executor = None
    path_config = None
    resource_remoto = False
    sharepoint_data_jmeter = False
    usuario_pybot_email = None
    email_pybot = None
    password_sharepoint = None
    sharepoint_url = None
    sharepoint_site = None
    sharepoint_doc = None

    def set_proyect(self, project_name=None):

        """
            Description:
                Setea variables de ambiente y rutas del proyecto.
            Args:
                project_name: Nombre del Proyecto
            Returns:
                Imprime por consola la siguiente configuración:
                    -Ambiente.
                    -Ruta de Resource.
                    -Ruta de Evidencias.
                    -Ruta de los Json.
                    -Ruta de las Imagenes de los json (reconocimiento por imagenes).
                    -Ruta de los Bass.
                Si hubo un error en la configuración, imprime por consola
                "No se pudieron detectar los datos de la ejecución".
        """
        # print(f"Plataforma {platform.system()} detectada.")
        self.LoggingFeature(f"Plataforma {platform.system()} detectada.").send_log()
        if platform.system() == "Windows":
            Functions.set_parameters_environment("Windows")
        elif platform.system() == "Linux":
            Functions.set_parameters_environment("Linux")
        if project_name is None and os.getenv('PROYECT') is None:
            if os.path.abspath(str(self)).split(' ')[0].split('\\')[-4] == 'src':
                Functions.project_name = os.path.abspath(str(self)).split(' ')[0].split('\\')[-5]
            else:
                Functions.project_name = os.path.abspath(str(self)).split(' ')[0].split('\\')[-4]
        elif os.getenv('PROYECT') is not None:
            Functions.project_name = os.getenv('PROYECT')
        else:
            Functions.project_name = project_name
        Functions.test_case_name = self.id().split('.')[-1]
        Functions.class_name = self.id().split('.')[-2]
        Functions.file_name = self.id().split('.')[-3]
        Functions.automatic_restore_row(self)
        if not Parameters.manual_increment:
            Functions.automatic_increment_row(self)
        if Parameters.environment == "Windows":
            base_path = os.path.join(Parameters.current_path)
            Functions.path_downloads = os.path.join(base_path, 'src', 'downloads')
            Functions.path_files = os.path.join(base_path, 'src', 'files')
            Functions.path_images = os.path.join(base_path, 'src', 'images')
            Functions.path_outputs = os.path.join(base_path, 'src', 'outputs')
            Functions.path_json = os.path.join(base_path, 'src', 'pages')
            Functions.path_resources = os.path.join(base_path, 'src', 'resources')
            Functions.path_steps = os.path.join(base_path, 'src', 'steps')
            Functions.path_config = os.path.join(base_path, 'config.yml')
            Functions.path_jmeter_executor = Parameters.path_jmeter
        if Parameters.environment == "Linux":
            base_path = os.path.join(Parameters.current_path)
            Functions.path_downloads = os.path.join(base_path, 'src', 'downloads')
            Functions.path_files = os.path.join(base_path, 'src', 'files')
            Functions.path_images = os.path.join(base_path, 'src', 'images')
            Functions.path_outputs = os.path.join(base_path, 'src', 'outputs')
            Functions.path_json = os.path.join(base_path, 'src', 'pages')
            Functions.path_resources = os.path.join(base_path, 'src', 'resources')
            Functions.path_steps = os.path.join(base_path, 'src', 'steps')
            Functions.path_config = os.path.join(base_path, 'config.yml')

        Functions.path_map = {"resources": Functions.path_resources,
                              "files": Functions.path_files,
                              "pages": Functions.path_json,
                              "images": Functions.path_images,
                              "downloads": Functions.path_downloads,
                              "outputs": Functions.path_outputs,
                              "steps": Functions.path_steps}

        Functions.create_folders_framework(Functions.path_map)

        if Functions.resource_remoto:
            Functions.download_file(self)

        if os.environ.get("rowXLSX"):
            Functions.set_manual_increment(True)
            Functions.set_excel_row(Functions.get_row_excel())
            Parameters.row = os.environ.get("rowXLSX")

        if os.environ.get("env"):
            if str(os.environ['env']).lower() == "qa":
                Parameters.env = "DataQa"
            if str(os.environ['env']).lower() == "prod":
                Parameters.env = "DataProd"
            if str(os.environ['env']).lower() == "test":
                Parameters.env = "DataTest"
            if str(os.environ['env']).lower() == "alt":
                Parameters.env = "DataAlt"
        else:
            Parameters.env = "DataTest"
        Functions.set_resources(self, Functions.class_name, Parameters.env)
        Functions.create_grid_by_sources(Functions.data_resource, "Datos del resource")
        return Functions.path_map

    def get_data_resource(self, key: str):
        """
        Recibe un diccionario del excel "resources" y una clave, devuelve el valor correspondiente.

        Args:
            resource (dict): Diccionario de resources.
            key (str): Clave para buscar en el diccionario de resources.

        Returns:
            str: Valor correspondiente a la clave dada. Devuelve un string vacío si la clave no está presente.
        """
        value = Functions.data_resource.get(key, None)
        if value is None:
            self.LoggingFeature(
                f"ADVERTENCIA: La clave {key} consultada no existe en el diccionario de resources.").send_log()
        return value

    def set_data_cache(self, key: str, value: Any):
        """
        Guarda un dato en el diccionario de cache.

        Args:
            key (str): Clave para guardar en el diccionario de cache.
            value (str): Valor correspondiente a la clave.

        Returns:
            None
        """
        Functions.data_cache[key] = value
        self.LoggingFeature(f"GUARDANDO DATO: clave '{key}' con valor: {value}").send_log()

    def get_data_cache(self, key):
        """
        Recibe un diccionario del cache y una clave, devuelve el valor correspondiente.

        Args:
            resource (dict): Diccionario de cache.
            key (str): Clave para buscar en el diccionario de cache.

        Returns:
            str: Valor correspondiente a la clave dada. Devuelve un string vacío si la clave no está presente.
        """
        value = Functions.data_cache.get(key, None)
        if value is None:
            print(f"ADVERTENCIA: La clave {key} consultada no existe en el diccionario de cache.")
        return value

    @staticmethod
    def create_grid_by_sources(resource: dict, message):

        body = """
            <!DOCTYPE html>
            <html>
            <head>
              <meta charset="utf-8">
              <title>Mi página web</title>
              <style>
                h1{
                  color: #D71920;
                  padding: 1%;
                  font-family: Arial, Helvetica, sans-serif;
                }

                ul {
                  list-style-type: disc; /* Tipo de viñeta, en este caso un círculo lleno */
                  margin: 0;
                  padding: 0;
                }

                li {
                  color:#D71920;
                  margin: 0 0 0 1em; /* Margen izquierdo para que se vea la viñeta */
                  font-family: Arial, Helvetica, sans-serif;
                  font-size: 15px;
                }
                span{
                  color: #757575;
                  font-size: 15px;
                  font-family: Arial, Helvetica, sans-serif;

                }
                .container{
                  background-color: #FFFFFF;
                  margin: 1%;
                  padding: 1%;
                  border-radius: 10px;
                  box-shadow: 0px 3px 10px #00000029;
                }
            </style>
            </head>
            <body>
                {list}
            </body>
            </html>
            """
        if len(resource) != 0:
            list_resources = ""
            for item in resource.items():
                resources_html = \
                    f"""<div class="container">
                        <ul>
                            <li><b>{item[0]}: </b><span>{item[1]}</span></li>
                        </ul>
                    </div>"""

                list_resources += resources_html
            body = body.replace("{list}", list_resources)
            try:
                allure.attach(body, message, attachment_type=allure.attachment_type.HTML)
            except Exception as e:
                Functions.exception_logger(e)

    @staticmethod
    def create_folders_framework(mapping: dict):
        for key in mapping.keys():
            if not os.path.exists(mapping[key]):
                os.makedirs(mapping[key])

            archivo_init = os.path.join(mapping[key], "__init__.py")
            open(archivo_init, 'a').close()

    @staticmethod
    def get_env():

        """

            Returns: Devuelve el valor de la variable de entorno 'env' en minúsculas'

        """

        return str(os.environ['env']).lower()

    @staticmethod
    def set_env(env):

        """
            Descripcion:
                Configura una variable para la lectura de resources.

            Args:
                env: QA, TEST, PROD, ALT

            Returns:
                Funcion que configura la variable de ambiente para la lectura del resources

        """
        os.environ['env'] = env

    @staticmethod
    def set_parameters_environment(system):

        """
            Description:
                Configura las opciones del framework en funcion del SO.
            Args:
                system: Sistema operativo anfitrion
        """

        Parameters.environment = system

    def set_retry(self, numbers_retries):

        """
            Description:
                Configura la cantidad de reintentos que se realizan al buscar algun objeto o realizar alguna espera.
            Args:
                numbers_retries: Cantidad de veces que se quiere reintentar.
        """

        Parameters.number_retries = numbers_retries
        print(f"La cantidad de reintentos configurada es {Parameters.number_retries}.")

    @staticmethod
    def set_remote_resource(value):

        """
            Description:
                Activar el modo para leer resources remotamente desde SharePoint de Pybot Team & qa.
            Args:
                value (bool): activar o desactivar el modo de resource remoto.
        """

        Functions.resource_remoto = value

    def auth(self):

        """
            Description:
                Autenticarse con un usuario de SharePoint y obtener una instancia para interactuar con el content
                del mismo.
            Returns:
                site (Site): Instancia del SharePoint.
        """

        Functions.password_sharepoint = Functions.Encryptor("CLAVES", "id", "Sharepoint Andreani", "PASS").main()
        Functions.sharepoint_url = Functions.Encryptor('CLAVES', 'id', 'Sharepoint Andreani', 'URL').main()
        Functions.email_pybot = Functions.Encryptor("CLAVES", "id", "Email de Pybot", "USER").main()
        Functions.sharepoint_site = Functions.Encryptor('CLAVES', 'id', 'Sharepoint Andreani', 'SITE').main()
        authcookie = Office365(Functions.sharepoint_url,
                               username=Functions.email_pybot,
                               password=Functions.password_sharepoint).GetCookies()
        site = Site(Functions.sharepoint_site, version=Version.v365, authcookie=authcookie)

        return site

    def connect_folder(self, folder_name):

        """
            Description:
                Obtener la instancia de una carpeta del SharePoint para interactuar con su content.
            Args:
                folder_name (str): nombre de la carpeta en SharePoint Ej. Ebuyplace/src/outputs.
            Returns:
                folder (Folder): instancia de la carpeta para interactuar con su content.
        """

        auth_site = self.auth()
        Functions.sharepoint_doc = Functions.Encryptor('CLAVES', 'id', 'Sharepoint Andreani', 'LOCATION').main()

        sharepoint_dir = '\\'.join([Functions.sharepoint_doc, folder_name])
        folder = auth_site.Folder(sharepoint_dir)

        return folder

    def upload_file(self, file, file_name, folder_name):

        """
            Description:
                Subir un archivo local en SharePoint de Pybot Team & QA.
            Args:
                file (str): Dirección local del archivo.
                file_name (str): Nombre del archivo Ej. factura.pdf
                folder_name (strt): Dirección donde se subirá el archivo en SharePoint Ej. Ebuyplace/src/outputs
        """

        _folder = self.connect_folder(folder_name)

        with open(file, mode='rb') as file_obj:
            file_content = file_obj.read()

        _folder.upload_file(file_content, file_name)

    def download_file(self):

        """
            Description:
                Lee el content de el resource correspondiente para el caso de prueba que va a ejecutarse y lo
                descarga. El nombre del resource debe ser el mismo que el de caso de prueba, y se debe respetar
                la estructura de carpetas, por ejemplo:
                Para un conjunto de pruebas llamado A.py debe existir una ruta en SharPoint
                /Ebuyplace/src/resources/A.xlxs
                que es de donde se leerá los datos a utilizarse en las pruebas.
        """

        # set the folder name
        folder_name = Functions.path_resources.split("\\")[-3] + "/" + Functions.path_resources.split("\\")[-2] + "/" + \
                      Functions.path_resources.split("\\")[-1] + "/"

        # set file name
        file_name = f"{Functions.file_name}.xlsx"

        # set dowload path
        download_path = f"{Functions.path_resources}\\{Functions.file_name}.xlsx"

        _folder = Functions.connect_folder(folder_name)
        file = _folder.get_file(file_name)

        # save file
        with open(download_path, 'wb') as f:
            f.write(file)
            f.close()

    @staticmethod
    def attach_data_test_in_allure_step(msg, data):

        """
            Description:
                Crea una tabla HTML con los datos (diccionario) pasados como parámetro y
                lo adjunta a un step allure.
            Args:
                msg (str): Mensaje que se quiere dejar en el step allure.
                data: Diccionario que se quiere mostrar en forma de tabla html.
        """

        table_html = "<!DOCTYPE html><html><head><style>table " \
                     "{font-family: arial, sans-serif;border-collapse: collapse;width: 100%;}" \
                     "td, th {border: 1px solid #dddddd;text-align: left;padding: 8px;}tr:nth-child(1) " \
                     "{background-color: #009efb;}</style></head><body><table><tr>"
        if len(data.keys()) > 0:
            for key in data.keys():
                table_html += f"<th>{key}</th>"
            table_html += f"</tr><tr>"
            for value in data.values():
                table_html += f"<td>{value}</td>"
            table_html += "</tr></table></body></html>"
            allure.attach(table_html, msg, allure.attachment_type.HTML)

    def automatic_restore_row(self):

        """
            Description:
                Restaura la variable Parameters.row a 2 cuando se pasa de un archivo a.py a b.py.
        """

        file_name = str(self)[str(self).find("(") + 1:str(self).find(")")]
        if Functions.get_file_name_stored() is None:
            Functions.set_file_name_stored(file_name)
        else:
            if file_name != Functions.get_file_name_stored():
                Functions.set_restore_excel_row()
                Functions.set_file_name_stored(file_name)

    def automatic_increment_row(self):

        """
            Description:
                Incrementa Parameters.row de a cuerdo al test_id que se está ejecutando.
        """
        # dudas aca 1
        if not Parameters.manual_increment:
            id_case = str(Functions.test_case_name.split('_')[1])
            if id_case.isdigit():
                Parameters.row = int(id_case) + 2
            else:
                unittest.TestCase.skipTest(self, f"No es posible realizar la ejecución del caso"
                                                 f" '{Functions.test_case_name}' ya que el nombre del mismo no respeta"
                                                 f" la siguiente convención de nombres 'test_xxx_descripcion'"
                                                 f" siendo x un numero entero.")

    @staticmethod
    def get_file_name_stored():

        """
            Description:
                Obtiene el parámetro file_name_stored de la configuracion.
            Returns:
                Devuelve el parámetro el valor del parámetro file_named_stored.
        """

        return Parameters.file_name_stored

    @staticmethod
    def set_file_name_stored(file):

        """
            Description:
                Setea el valor del parámetro file_name_stored de la configuración.
            Args:
                file: El nuevo nombre del archivo.
        """

        Parameters.file_name_stored = file

    @staticmethod
    def get_path_system():

        """
            Description:
                Obtiene el directorio base del sistema
            Returns:
                Devuelve el directorio base del proyecto
       """

        return Parameters.current_path

    @staticmethod
    def get_row_excel():

        """
            Description:
                Obtiene la row actual del excel.
            Returns:
                Imprime por consola "El numero del registro consultado es: "+ str(row)"
                y retorna la row.
        """

        print(f"El numero del registro consultado es: {Parameters.row} .")
        return Parameters.row

    @staticmethod
    def set_excel_row(value: int):
        Parameters.row = value

    @staticmethod
    def set_manual_increment(value: bool):
        Parameters.manual_increment = value

    @staticmethod
    def set_increment_row():

        """
            Description:
                Incrementa en 1 el número de registro que será consultado en el resource.
        """

        Parameters.row += 1
        Parameters.manual_increment = True
        print(f"El numero del registro fue aumentado a: {Parameters.row}")

    @staticmethod
    def set_restore_excel_row():

        """
            Description:
                Restaura al value original en "2" el número de registro que será consultado en el resource.
        """

        Parameters.row = 2
        print(f"El numero del registro fue restaruado a: {Parameters.row}")

    def read_cell(self, celda, file_name=None, specific_sheet=None):

        """
            Description:
                Lee la cell de un resource.
            Args:
                celda (obj): Celda del resource.
                file_name (str): Nombre del caso.
                specific_sheet (str): Hoja del resource.
            Returns:
                Retorna el value de la cell del resource.
        """

        if file_name is None:
            print("El nombre del caso es : " + Functions.file_name)
            file_name = self.file_name
        resource = f"{Functions.path_resources}\\{file_name}.xlsx"
        if not os.path.isfile(resource):
            resource = f"{Functions.path_resources}{file_name}.xlsx"
            if not os.path.isfile(resource):
                raise Exception('El resource no existe')
        wb = openpyxl.load_workbook(resource, data_only=True)
        if specific_sheet is None:
            sheet = wb["DataTest"]
        else:
            sheet = wb[specific_sheet]
        if sheet[celda].value is None:
            value = ""
        else:
            value = str(sheet[celda].value)
        print(f"El libro de excel utilizado es de es: {resource}")
        print(f"El value de la cell es: {value}")
        return value

    def write_cell(self, cell, value, name, folder, sheet=None):

        """
            Description:
                Permite escribir en una celda indicada de una hoja especifica para un
                libro de excel en directorio ./inputs/.
            Args:
                cell (obj): Celda de la hoja, se espera COLUMNA+FILA.
                value (str): Valor a ingresar en la celda.
                name (str): Nombre del libro de excel, en el directorio ./inputs/.
                sheet (str): Hoja especifica del libro excel.
                folder (str): Nombre de la carpeta que contiene el libro excel. Es 'files' por default o puede ser
                'downloads'.
            Returns:
                Imprime por consola la celda, hoja y valor escrito, y devuelve TRUE
                en caso contrario imprime por consola "VERIFICAR: No se pudo escribir el archivo."
                y devuelve FALSE.
        """
        resource = ''
        try:
            if folder == 'files':
                resource = f"{Functions.path_files}\\{name}.xlsx"
            elif folder == 'downloads':
                resource = f"{Functions.path_downloads}\\{name}.xlsx"
            print(resource)
            wb = openpyxl.load_workbook(resource)
            if sheet is None:
                hoja = wb["DataTest"]
            else:
                hoja = wb[sheet]
            hoja[cell] = value
            print(value)
            print(sheet)
            print(cell)
            wb.save(filename=resource)
            wb.close()
            flag = True
            print(f"El libro de excel utilizado es: {resource}")
            if not (sheet is None):
                print(f"Se escribio en la celda {str(cell)} de la hoja {str(sheet)} el valor: {str(value)}")
            else:
                print(f"Se escribio en la celda {str(cell)} el valor: {str(value)}")
            print(flag)
            return flag
        except Exception as e:
            flag = False
            Functions.exception_logger(e)
            print("VERIFICAR: No se pudo escribir el archivo.")
            return flag

    @staticmethod
    def wait(time_load, logger=Parameters.loggin_time, reason=None):

        """
            Description:
                Espera un elemento, el tiempo es dado en segundos.
            Args:
                time_load (int): Tiempo en segundos.
                logger: Indica si se requieren logear los mensajes de espera.
                reason: Razón por la que se quiere esperar un elemento.
            Returns:
                Cuando termina el tiempo de espera imprime "Esperar: Carga Finalizada ... ".
        """
        actual_hour = datetime.datetime.now().time().strftime(Parameters.time_format)
        if logger:
            print(f"AGUARDANDO: Inicia espera de '{str(time_load)}' segundo/s a las {actual_hour}.")
        if reason is not None:
            print(reason)
        try:
            total_wait = 0
            while total_wait < time_load:
                time.sleep(1)
                total_wait = total_wait + 1
        finally:
            actual_hour = datetime.datetime.now().time().strftime(Parameters.time_format)
            if logger:
                print(f"AGUARDANDO: Finaliza espera a las {actual_hour}")

    # FUNCIONES BASE DE DATOS ##########################################################################################
    def set_timeout_base_sql_server(self, time_seconds):

        """
            Description:
               Configura el value de timeout (segundos) configurado para las conexiones a bases sqlServer.
            Args:
                time_seconds: Valor (int) que representa una cantidad en segundos.
        """

        Parameters.timeout_base_sql_server = time_seconds
        time_timeout = Parameters.timeout_base_sql_server
        print(f"El nuevo value de timeout para la conexion de la base sql es de {time_timeout} segundos.")

    def get_timeout_base_sql_server(self):

        """
            Description:
                Devuelve el value de timeout configurado para la conexión a bases sqlServer.
            Returns:
                Devuelve el value de timeout (segundos) configurado para la conexion a bases sqlServer.
        """

        time_timeout = Parameters.timeout_base_sql_server
        print(f"El value de timeout para la conexion de la base sql es de {time_timeout} segundos.")
        return time_timeout

    def establish_connection_sqlserver(self, db_name):

        """
            Description:
                Realiza conexión a una base de datos sqlServer.
            Args:
                server: Servidor ip.
                base: Nombre de la base.
                user: Usuario.
                password: Contraseña.
            Returns:
                Devuelve una variable con la conexion a la base de datos sqlServer.
        """
        driver = None
        conection = None
        ################################# GET DATA FROM XML FILE FOR DATABASE CONNECTION ###############################
        server = Functions.Encryptor("CLAVES", "id", db_name, 'IP', False).main()
        db_port = Functions.Encryptor("CLAVES", "id", db_name, 'PORT', False).main()
        base = Functions.Encryptor("CLAVES", "id", db_name, 'BASE', False).main()
        user = Functions.Encryptor("CLAVES", "id", db_name, 'USER', False).main()
        password = Functions.Encryptor("CLAVES", "id", db_name, 'PASS', False).main()
        ################################# GET DATA FROM XML FILE FOR DATABASE CONNECTION ###############################
        if Parameters.environment == "Linux":
            driver = "/usr/lib/libtdsodbc.so"
        if Parameters.environment == "Windows":
            driver = "{SQL Server}"
        try:
            conection = pyodbc.connect(f"Driver={driver};Server={server};PORT={db_port};Database={base};UID={user};"
                                       f"PWD={password}")
        except pyodbc.OperationalError:
            unittest.TestCase().fail("El servidor no existe o el acceso al mismo fué denegado.")
        finally:
            if conection is not None:
                conection.timeout = Parameters.timeout_base_sql_server
                return conection

    def check_base_sqlserver(self, db_name, query):
        """
            Description:
                Realiza conexión y consulta a base de datos con la libreria pyodbc. El método incluye la
                desconexión.
            Args:
                server: Servidor ip.
                base: Nombre de la base.
                user: usuario.
                password: Contraseña.
                query: consulta Query.
            Returns:
                <class 'pyodbc.Row'>: Retorna un class 'pyodbc.Row' si la consulta y la conexión es exitosa. De lo
                 contrario imprime por consola "Se produjo un error en la base de datos."
        """

        cursor = None
        recordset = []
        conn = Functions.establish_connection_sqlserver(self, db_name)
        try:
            cursor = conn.cursor()
            cursor.execute(query)
            for row in cursor:
                recordset = row

        except Exception as e:
            Functions.exception_logger(e)
            print(f"Se produjo un error en la base de datos.")

        finally:
            cursor.close()
            conn.close()
            return recordset

    def execute_sp_base_sqlserver(self, db_name, query, parameters: tuple):

        """
            Description:
                Realiza conexión y consulta a base de datos con la libreria pyodbc. El método incluye la
                desconexión.
            Args:
                server (str): Servidor ip.
                base (str): Nombre de la base.
                user (str): usuario.
                password (str): Contraseña.
                query (str): consulta Query.
                parameters (tuple): tupla con parametros para el sp.
            Returns:
                <class 'pyodbc.Row'>: Retorna un class 'pyodbc.Row' si la consulta y la conexión es exitosa. De lo
                 contrario imprime por consola "Se produjo un error en la base de datos."
        """

        recordset = []
        cursor = None
        connection = Functions.establish_connection_sqlserver(self, db_name)

        try:
            cursor = connection.cursor()
            cursor.execute(query, parameters)
            for row in cursor:
                recordset.append(row)

        except Exception as e:
            Functions.exception_logger(e)
            print("Se produjo un error en la base de datos.")

        finally:
            cursor.close()
            connection.close()
            return recordset

    def get_list_base_sqlserver(self, db_name, query):

        """
            Description:
                Realiza conexión y consulta a base de datos con la libreria pyodbc. El método incluye la
                desconexión.
            Args:
                server (str): Servidor ip.
                base (str): Nombre de la base.
                user (str): usuario.
                password (str): Contraseña.
                query (str): consulta Query.
            Returns:
                results: Lista con los resultados.
        """

        recordset = []
        cursor = None
        connection = Functions.establish_connection_sqlserver(self, db_name)

        try:
            cursor = connection.cursor()
            cursor.execute(query)
            for row in cursor:
                recordset.append(row)

        except Exception as e:
            Functions.exception_logger(e)
            print("A ocurrido un error en la base de datos.")

        finally:
            cursor.close()
            connection.close()

        return recordset

    def get_recordset_sqlserver(self, db_name, query):

        """
            Description:
                Realiza conexión y consulta a base de datos con la libreria pyodbc. El método incluye la
                desconexión.
            Args:
                server (str): Servidor ip.
                base (str): Nombre de la base.
                user (str): usuario.
                password (str): Contraseña.
                query (str): consulta Query.
            Returns:
                results: Lista con diccionarios que referencian las valores con sus correspondientes columnas.
        """

        recordset = []
        cursor = None
        connection = Functions.establish_connection_sqlserver(self, db_name)

        try:
            cursor = connection.cursor()
            cursor.execute(query)
            records = cursor.fetchall()
            column_names = [column[0] for column in cursor.description]
            for record in records:
                recordset.append(dict(zip(column_names, record)))

        except Exception as e:
            Functions.exception_logger(e)
            print(f"A ocurrido un error en la base de datos. {e}")

        finally:
            cursor.close()
            connection.close()

        return recordset

    def delete_reg_base_sqlserver(self, db_name, query):

        """
            Description:
                Elimina un registro de la base de datos. El método incluye la desconexión.
            Args:
                server (str): Servidor ip.
                base (str): Nombre de la base.
                user (str): usuario.
                password (str): Contraseña.
                query (str): consulta Query.
        """

        cursor = None
        conn = Functions.establish_connection_sqlserver(self, db_name)
        try:
            cursor = conn.cursor()
            cursor.execute(query)
            conn.commit()
            print("Borrado de registro en base de datos realizado exitosamente.")
        except Exception as e:
            Functions.exception_logger(e)
            print("Ocurrio un error en la base al intentar eliminar un registro.")
        finally:
            cursor.close()
            conn.close()

    def insert_row_base_sqlserver(self, db_name, query):

        """
            Description:
                Inserta un nuevo registro en la base de datos. El método incluye la desconexión.
            Args:
                server (str): Servidor ip.
                base (str): Nombre de la base.
                user (str): usuario.
                password (str): Contraseña.
                query (str): consulta Query.
        """

        cursor = None
        conn = Functions.establish_connection_sqlserver(self, db_name)
        try:
            cursor = conn.cursor()
            cursor.execute(query)
            conn.commit()
            print("Insertado de registro en base de datos realizado exitosamente.")
        except Exception as e:
            Functions.exception_logger(e)
            print(f"Ocurrio un error en la base al intentar un registro.")
        finally:
            cursor.close()
            conn.close()

    def update_row_base_sqlserver(self, db_name, query):

        """
            Description:
                Actualiza un registro en la base de datos. El método incluye la desconexión.
            Args:
                server (str): Servidor ip.
                base (str): Nombre de la base.
                user (str): usuario.
                password (str): Contraseña.
                query (str): consulta Query.
        """

        cursor = None
        conn = Functions.establish_connection_sqlserver(self, db_name)
        try:
            cursor = conn.cursor()
            cursor.execute(query)
            conn.commit()
            print("Actualización de registro en base de datos realizado exitosamente.")
        except Exception as e:
            Functions.exception_logger(e)
            print(f"Ocurrio un error en la base al intentar actualizar un registro.")
        finally:
            cursor.close()
            conn.close()

    ##     ORACLE         ###
    ##     ORACLE         ###
    def establish_connection_oracle_db(self, db_name):
        """
            Description:
                Realiza conexión a una base de datos Oracle.
            Args:
                server: nombre desde archivo encriptado.
                base: Nombre de la base. IP:PUERTO/base
                user: Usuario.
                password: Contraseña.
            Returns:
                Devuelve una variable con la conexion a la base de datos Oracle.
        """

        ################################# GET DATA FROM XML FILE FOR DATABASE CONNECTION ###############################
        server = Functions.Encryptor("CLAVES", "id", db_name, 'IP', False).main()
        db_port = Functions.Encryptor("CLAVES", "id", db_name, 'PORT', False).main()
        base = Functions.Encryptor("CLAVES", "id", db_name, 'BASE', False).main()
        user = Functions.Encryptor("CLAVES", "id", db_name, 'USER', False).main()
        password = Functions.Encryptor("CLAVES", "id", db_name, 'PASS', False).main()
        ################################# GET DATA FROM XML FILE FOR DATABASE CONNECTION ###############################
        connection = cx_Oracle.connect(user, password, base, encoding="UTF-8")
        return connection

    def get_recordset_oracle_db(self, db_name, query):

        """
            Description:
                Realiza conexión y consulta a base de datos con la libreria cx_Oracle. El método incluye la
                desconexión.
            Args:
                server (str): Servidor ip.
                base (str): Nombre de la base.
                user (str): usuario.
                password (str): Contraseña.
                query (str): consulta Query.
            Returns:
                results: Lista con diccionarios que referencian las valores con sus correspondientes columnas.
        """

        recordset = []
        cursor = None
        connection = Functions.establish_connection_oracle_db(self, db_name)

        try:
            cursor = connection.cursor()
            cursor.execute(query)
            records = cursor.fetchall()
            column_names = [column[0] for column in cursor.description]
            for record in records:
                recordset.append(dict(zip(column_names, record)))

        except Exception as e:
            Functions.exception_logger(e)
            print(f"La intereacción con la DB de Oracle arrojó el siguiente error: {e}")

        finally:
            cursor.close()
            connection.close()

        return recordset

    def check_base_oracle_db(self, db_name, query):

        """
            Description:
                Realiza conexión y consulta a base de datos con la libreria cx_Oracle. El método incluye la
                desconexión.
            Args:
                server: Servidor ip.
                base: Nombre de la base.
                user: usuario.
                password: Contraseña.
                query: consulta Query.
            Returns:
                <class 'pyodbc.Row'>: Retorna un class 'pyodbc.Row' si la consulta y la conexión es exitosa. De lo
                 contrario imprime por consola "Se produjo un error en la base de datos."
        """

        cursor = None
        recordset = []
        conn = Functions.establish_connection_oracle_db(self, db_name)
        try:
            cursor = conn.cursor()
            cursor.execute(query)

            for row in cursor:
                recordset = row

        except Exception as e:
            Functions.exception_logger(e)
            print(f"La intereacción con la DB de Oracle arrojó el siguiente error: {e}")

        finally:
            cursor.close()
            conn.close()
            return recordset

    def get_oracle_db_headers(self, db_name, query):

        """
            Description:
                Realiza conexión y consulta a base de datos con la libreria cx_Oracle. El método incluye la
                desconexión.
            Args:
                server (str): Servidor ip.
                base (str): Nombre de la base.
                user (str): usuario.
                password (str): Contraseña.
                query (str): consulta Query.
            Returns:
                results: Lista con los nombres de la cabecera correspondiente a la consulta.
        """
        cursor = None
        connection = Functions.establish_connection_oracle_db(self, db_name)
        column_names = None

        try:
            cursor = connection.cursor()
            cursor.execute(query)
            column_names = [column[0] for column in cursor.description]

        except Exception as e:
            Functions.exception_logger(e)
            print(f"La intereacción con la DB de Oracle arrojó el siguiente error: {e}")

        finally:
            cursor.close()
            connection.close()

        return column_names

    def create_file_validations(self, data_validations, name_template):

        """
            Description:
                Crea un archivo html con las validaciones realizadas.
            Args:
                data_validations (dict): Diccionario con información de las validaciones realizadas.
                name_template (str): Nombre del template a utilizar.
            Return:
                Retorna la ruta del archivo html creado.
        """

        from datetime import datetime
        replacement = ""
        path = Functions.path_outputs

        date_time = datetime.now().strftime("%d-%m-%y_%H-%M-%S-%f")
        file_name = f"{date_time}.html"

        path_file = os.path.join(path, file_name)

        with open(path_file, 'a', encoding="utf-8") as f:
            template = open(path + f'\\{name_template}.html', 'r', encoding="utf-8")
            data_template = template.read()
            f.write(data_template)
            template.close()
            f.close()

        with open(path_file, 'r', encoding="utf-8") as f:
            data_file = f.readlines()
            data_string = json.dumps(data_validations)
            data_string = f"data = {data_string};"
            data_without_spaces = [i.strip() for i in data_file]

            try:
                index = data_without_spaces.index("window.onload = (event) => {")
            except ValueError:
                print("Hubo un error al generar el template de evidencias")

            data_file[index + 1] = data_string
            for line in data_file:
                line = line.strip()
                changes = line.replace('\n', "")
                changes = changes.replace('\\"', '"')
                replacement = replacement + changes
            f.close()

        with open(path_file, 'w', encoding="utf-8") as f:
            f.writelines(replacement)
            f.close()

        return path_file

    # FUNCIONES INFORMES ###############################################################################################

    @staticmethod
    def create_message_html(message_text: str, special_strings=None):

        """
            Description:
                Crea un párrafo en formato html.
            Args:
                message_text: mensaje en formato string.
                special_strings: Lista de palabras que deben ser resaltadas en negrita dentro del mensaje.
            Returns:
                Devuelve el párrafo en formato html.
        """

        message_html = f'<p>{message_text}</p>'

        if special_strings is not None:
            for string in special_strings:
                message_html = message_html.replace(string, f"<strong>{string}</strong>")

        return message_html

    @staticmethod
    def create_message_teams(message_text: str, special_strings=None):

        """
            Description:
                Crea un párrafo en formato de notificacion teams.
            Args:
                message_text: mensaje en formato string.
                special_strings: Lista de palabras que deben ser resaltadas en negrita dentro del mensaje.
            Returns:
                Devuelve el párrafo en formato html.
        """
        style = 'style=font-size: 14px;font: "Helvetica";'
        message = f'<p style="{style}">{message_text}</p>'
        if special_strings is not None:
            for string in special_strings:
                message = message.replace(string, f"<strong>{string}</strong>")
        return message

    @staticmethod
    def send_message_teams(channel, title, message=None, sections=None):

        """
        Description:
            Realiza el envio de notificaciones via microsoft teams.
        Args:
            channel: Canal objetivo de la notificacion. Debe ser configurado el webhook en teams.
            title: Titulo de la notificacion.
            message: Mensaje de la notificacion.
            sections: lista de secciones con contenido para la generación de la notificación.
        """

        message_teams = pymsteams.connectorcard(channel)
        message_teams.color(Parameters.teams_notifications_colors)
        message_teams.title(title)
        if message is None:
            message_teams.text("<p></p>")
        else:
            message_teams.text(message)
        if sections is not None:
            if type(sections) is list:
                for section in sections:
                    message_teams.addSection(section)
            else:
                message_teams.addSection(sections)

        message_teams.addSection(Functions.create_section_teams(Functions.footer_signature_teams()))
        message_teams.send()

    @staticmethod
    def create_section_teams(message, title=None, content=None):

        """
            Description:
                Crea secciones para notificaciones de teams.
            Args:
                message: Mensaje de la seccion.
                title: Titulo de la seccion.
                content: Contenido de la seccion.
            Returns:
                Seccion de microsoft teams.
        """

        my_section = pymsteams.cardsection()
        if title is not None:
            my_section.title(f"<h3>{title}</h3>")
        my_section.text(message)
        if content is not None:
            my_section.text(content)
        return my_section

    @staticmethod
    def add_button_teams():
        pass

    @staticmethod
    def footer_signature_teams():

        """
        Descripcion:
            Agrega una firma a las notificaciones generadas en teams.
        Returns:
            string con la firma en teams
        """

        signature = f'''<footer>
                          <p class="member"><strong class="team">Equipo QA Cross Automation</strong></p>
                          <p class="member"><strong>Joel Pino</strong> || jpino@andreani.com</p>
                          <p class="member"><strong>Lucas Cariboni</strong> || lcariboni@andreani.com</p>
                        </footer>'''
        return signature

    @staticmethod
    def create_title(title_text: str, format_table="HTML"):

        """
            Description:
                Crea un título en formato html.
            Args:
                title_text: título en formato value_text.
                format_table: el formato de la tabla que se desea crear. formatos admitidos HTML y MSTEAMS.
            Returns:
                Devuelve título en formato html.
        """

        return f'<h5{title_text}</h5>'

    @staticmethod
    def create_table(list_data_head: list, list_data_content: list, format_table="HTML"):

        """
            Description:
                Crea una tabla html.
            Args:
                list_data_head: lista con los encabezados de la tabla.
                list_data_content: Matriz (lista con lista) con los datos de la tabla.
                format_table: el formato de la tabla que se desea crear. formatos admitidos HTML y MSTEAMS.
            Returns:
                Devuelve una tabla en formato html.
        """
        table_head_html = ""
        table_content_html = ""
        table_colums_html = ""
        table_html = ""
        if format_table == "HTML":
            for row in list_data_head:
                table_head_html = f'{table_head_html}<th>{row}</th>'
            table_head_html = f"<tr>{table_head_html}</tr>"
            for rows in list_data_content:
                for col in rows:
                    table_colums_html = f"{table_colums_html}<td>{col}</td>"
                table_content_html = f"{table_content_html}<tr>{table_colums_html}</tr>"
                table_colums_html = ""
            table_html = f"{table_head_html}{table_content_html}"

        elif format_table == "MSTEAMS":
            for row in list_data_head:
                table_head_html = f'{table_head_html}<th style="border: 1px solid gray; padding: 1rem 2rem 1rem 2rem; background-color: {Parameters.teams_notifications_colors}; text-align: center; color: white;">{row}</th>'
            table_head_html = f'<tr>{table_head_html}</tr>'
            for rows in list_data_content:
                for col in rows:
                    table_colums_html = f'{table_colums_html}<td style="border: 1px solid gray; padding: 1rem 2rem 1rem 2rem; background-color: white; text-align: center; color: {Parameters.teams_notifications_colors};">{col}</td>'
                table_content_html = f'{table_content_html}<tr>{table_colums_html}</tr>'
                table_colums_html = ""
            table_html = f"{table_head_html}{table_content_html}"

        if format_table == "HTML":
            return f"<table>{table_html}</table>"
        elif format_table == "MSTEAMS":
            return f'<pre style="margin: 0px; padding: 0px;"><table style="border-spacing: 0px; text-align:center; margin: 0px; width:90%: heigh: 100%;">{table_html}</table></pre>'

    @staticmethod
    def create_style_html():

        """
            Description:
                Devuelve el código css con los estilos que deben aplicarse a un bloque HTML.
            Returns:
                Devuelve el estilo para aplicar al código html.
        """

        style = '''<style>
                      * {
                        font-family: "Calibri", "Helvetica", "Arial", "Trebuchet MS", "sans-serif";
                        padding: none;
                        margin: none;
                        outline: none;
                        font-size: 14px;
                        margin-bottom: 2rem;
                      }
                      h5 {
                        font-size: 20px;
                      }
                      p {
                        padding-left: 1rem;
                        font-size: 14px;
                      }
                      strong {
                        font-size: 14px;
                        font-style: inherit;
                        color: #616161;
                      }
                      td,
                      th {
                        text-align: center;
                        border: 1px solid gray !important;
                        padding: 1rem 2rem 1rem 2rem;
                        margin: 1rem;
                      }
                      tr:nth-child(even) {
                        background-color: #f2f2f2;
                        padding-bottom: 1em;
                      }
                      tr:hover {
                        background-color: #d9534f;
                      }
                      th {
                        padding: 1rem 2rem 1rem 2rem;
                        margin: 1rem;
                        text-align: center;
                        background-color: #d9534f;
                        color: white;
                      }
                      table {
                        padding-left: 1rem;
                      }
                      img {
                        width: 10rem;
                      }
                      .team {
                        font-size: 16px;
                        font-style: inherit;
                        color: #ff644b;
                      }
                      .member {
                        margin: 0px !important;
                        padding: 0px !important;
                      }
                </style>'''

        return style

    @staticmethod
    def footer_signature_html():
        signature = f'''<footer>
                  <p class="member"><strong class="team">Equipo QA Cross Automation</strong></p>
                  <p class="member"><strong>Joel Pino</strong> || jpino@andreani.com</p>
                  <p class="member"><strong>Lucas Cariboni</strong> || lcariboni@andreani.com</p>
                </footer>'''
        return signature

    @staticmethod
    def create_image(image):

        """
            Description:
                Crea una imágen en formato html.
            Args:
                image: imágen a adjuntar.
            Returns:
                Devuelve imágen en formato html.
        """

        data = open(image, 'rb').read()  # read bytes from file
        data_base64 = base64.b64encode(data)  # encode to base64 (bytes)
        data_base64 = data_base64.decode()  # convert bytes to value_text
        return f'<img src="data:image/jpeg;base64,{data_base64}">'

    @staticmethod
    def apply_style_css_to_block(block_html: str):

        """
            Description:
                Aplica estilos css a un bloque html.
            Args:
                block_html: bloque html que recibirá los estilos css.
            Returns:
                Devuelve un bloque html con estilos aplicados
        """

        block_html_css = f"{block_html}{Functions.create_style_html()}"
        return block_html_css

    # FUNCIONES MONGODB ################################################################################################

    @staticmethod
    def insert_collection_into_mongodb(connection: list, coleccion_datos: list):

        """
            Description:
                Dada una conexión a base y una colección de "documentos" se insertan los mismos en mongoDB.
            Args:
                connection: Lista con la conexión a la base de mongoBD, tener en cuenta que tiene que tener:
                        -servidor/puerto/timeout para la conexión
                        -nombre de la base de datos
                        -nombre de la colección de documentos (para extrapolar a otros tipos de bases, "esquemas")
                coleccion_datos: Lista que contiene los documentos a insertar en base, los documentos puden ser
                diccionarios de datos o datos más complejos.
            Returns:
                Devuelve un texto con el resultado de la inserción.
        """

        client = None
        try:
            uri_connection = f"mongodb://{connection['MONGODB_HOST']}:{connection['MONGODB_PORT']}/"
            client = pymongo.MongoClient(uri_connection, serverSelectionTimeoutMS=connection['MONGODB_TIMEOUT'])
            collection = client[connection['DB']][connection['COLLECTION']]
            collection.insert_many(coleccion_datos)
            print('Se inserto la coleccion de datos en la base')
        except ServerSelectionTimeoutError as e:
            Functions.exception_logger(e)
            print('No se ha podido establecer una conexion con la base mongoDB.')
        except ConnectionFailure as e:
            Functions.exception_logger(e)
            print('Fallo la conexion con la base mongoDB.')
        finally:
            client.close()  # --> cierra coneccion NUNCA OLVIDAR

    # FUNCIONES DE ESCRITURA DE ARCHIVOS ###############################################################################
    def write_file(self, collection_data, format_file, delimiter=',', head=True, file_name=None, specific_sheet=None):

        """
            Description:
                Función de escritura de archivo, toma una lista de diccionarios con los datos a escribir y los
                guarda en el archivo especificado (crea el archivo).
                El archivo generado se guardará en la carpeta Output del mismo nombre del proyecto a testear.
            Args:
                collection_data: Lista de diccionario de datos con el content a escribir (obligatorio).
                format_file: Extension del archivo a escribir (obligatorio).
                delimiter: Delimitador (',').
                head: si el archivo contiene cabecera (por defecto esta en true).
                file_name: el nombre del archivo que se escribira.
                specific_sheet: si el archivo es un excel se le puede identificar el nombre de la hoja a utilizar.
        """

        f = None
        if file_name is None:
            print(f"El nombre del archivo a escribir es: {Functions.file_name}")  # tomo el nombre del
            file_name = Functions.file_name  # archivo de los datos de "inicializacion"
            resource = Functions.path_outputs + file_name + "." + format_file
        else:
            print("El nombre del archivo a escribir es : " + file_name)
            resource = Functions.path_outputs + file_name + "." + format_file

        if format_file == "xlsx":
            Functions.convert_data_to_excel_format(collection_data, resource, head, specific_sheet)
            print("Se escribio el archivo correctamete.")
        else:
            contents = Functions.convert_data_to_csv_format(collection_data, delimiter, head)
            # el archivo csv para poder escribirlo
            try:
                f = open(resource, "w", encoding='utf8')
                f.write(contents)
                print("Se escribio el archivo correctamete")
            except Exception as e:
                Functions.exception_logger(e)
                print("No se pudo escribir el archivo")
            finally:
                f.close()

    @staticmethod
    def convert_data_to_csv_format(collection_data, delimiter, head):

        """
            Description:
                Función que arma un value_text con formato separado por un delimitador para guardar en un archivo.
            Args:
                collection_data: Lista de diccionario de datos con el content a escribir (obligatorio).
                delimiter:
                head: Si el archivo contiene cabecera (por defecto esta en true).
            Returns:
                Devuelve el content unificado en un value_text para guardar en el archivo.
        """

        head_line = None
        line = None
        index = None

        if head:  # si tiene cabecera en verdadero
            for key in collection_data[0].keys():  # obtengo las key del diccionario de un item de la lista
                if head_line is not None:
                    head_line = f'{head_line}{delimiter}{key}'
                    # si el value_text de la linea cabecera no esta vacio sumo la key a lo que ya contiene mas
                    # el delimitador especificado
                else:
                    head_line = key
            head_line = head_line + "\n"  # al final de la linea hago un salto de linea

        # Ahora genero el content propiamente dicho, recorro la lista de de diccionarios y por cada diccionario obtengo
        # los valores de cada key
        for index in range(len(collection_data)):
            for key in collection_data[index].keys():
                value = collection_data[index][key]
                if line is not None:
                    if line.endswith(
                            "\n"):  # si lo ultimo que tiene el value_text es un salto de linea agrego solo el nuevo
                        # value a la linea
                        line = line + value
                    else:
                        line = f'{line}{delimiter}{value}'  # si no agrego el delimitador mas el value nuevo
                else:
                    line = value
            line = line + "\n"  # Al final de recorrer cada diccionario agrego un salto de linea al content
        index += 1

        # Para finalizar agrego la linea cabecera al resto de content para retornar un unico gran value_text con
        if head:
            contents = head_line + line
        else:
            contents = line
        return contents

    def read_excel(self, row_start: int, row_end: int, cols: int, file_name=None, specific_sheet=None):

        """
            Description:
                Lee un archivo excel con un pool de datos para utilizar en pruebas automatizadas y lo convierte
                en una lista de diccionarios de datos.
            Args:
                row_start: indico desde que row debe comenzar a leer (obligarotio)
                row_end: indico hasta que row debe leer (obligatorio)
                cols: indico cuantas columnas se leeran por cada row (obligatorio)
                file_name: el nombre del archivo excel que se leera (opcional)
                specific_sheet: nombre de la hoja que se leera (opcional)
            Returns:
                Devuelve una lista con los datos obtenidos de una fila del excel.
        """

        if file_name is None:
            print(f"Se leera el archivo con nombre: {Functions.file_name}")
            file_name = Functions.file_name
            resource = f"{Functions.path_resources}\\{file_name}.xlsx"
            book = openpyxl.load_workbook(resource, data_only=True)
        else:
            resource = f"{Functions.path_resources}\\{file_name}.xlsx"
            book = openpyxl.load_workbook(resource, data_only=True)

        if specific_sheet is None:
            sheet = book["Pool Data"]
        else:
            sheet = book[specific_sheet]
        key = []
        value = []
        collection_data = []
        dict_data = {}
        for row in sheet.iter_rows(min_row=row_start, max_col=cols, max_row=row_start):
            for cel in row:
                key.append(sheet[cel.coordinate].value)  # saco los valores que van a ser las Key del diccionario
        for row in sheet.iter_rows(min_row=row_start + 1, max_col=cols, max_row=row_end):
            for cel in row:
                value.append(sheet[cel.coordinate].value)  # obtengo los valores de las celdas
            for index in range(len(key)):
                dict_data.update({key[index]: value[index]})  # formo un diccionario
            value = []  # limpio la lista de valores de cell para la proxima iteracion
            collection_data.append(dict_data)  # guardo el diccionario en la lista
            dict_data = {}  # limpio el diccionario para la proxima iteracion

        return collection_data

    @staticmethod
    def convert_data_to_excel_format(collection_data, resource, head=True, specific_sheet=None):
        print(f"que pasa acá_ {collection_data}")
        """
            Description:
                Convierte una colección de datos en un formato excel.
            Args:
                collection_data: Lista de diccionario de datos con el content a escribir (obligatorio).
                resource: el nombre del archivo completo armado.
                head: si el archivo contiene cabecera (por defecto esta en true).
                specific_sheet: nombre de la hoja a utilizar.
        """

        # Inicializacion de variables
        keys = []
        values = []
        sheet = None

        # intento abrir el archivo a escribir, si no existe, lo creo y borro la hoja por defecto que trae
        try:
            book = openpyxl.load_workbook(resource, data_only=True)
        except FileNotFoundError:
            book = Workbook()
            book.save(resource)
            del book['Sheet']

        # reviso si me trajo un nombre de hoja especifico, borro la hoja existente y creo una nuevo para escribir en
        # una hoja limpia
        if specific_sheet is None:
            print(len(book.sheetnames))
            if len(book.sheetnames) == 0:
                sheet = book.create_sheet(title="OutputData")
            if "OutputData" in book.sheetnames:
                del book["OutputData"]
                sheet = book.create_sheet(title="OutputData")
        else:
            if len(book.sheetnames) == 0:
                sheet = book.create_sheet(title=specific_sheet)
            if specific_sheet in book.sheetnames:
                del book[specific_sheet]
                sheet = book.create_sheet(title=specific_sheet)

        if head:  # si tiene cabecera en verdadero
            for key in collection_data[0].keys():  # obtengo las key del diccionario de un item de la lista
                keys.append(key)  # lo guardo en una lista
            sheet.append(keys)  # lo guardo en la lista de la hoja
            # ahora recorro la lista de diccionarios para guardar en la hoja todos los datos
            for index in range(len(collection_data)):
                for row in collection_data[index].values():
                    values.append(row)  # guardo los valores de cada key en una lista
                sheet.append(values)  # guardo la lista en la lista "hoja"
                values = []  # limpio la lista de valores para la proxima
                # iteracion porque si no se acumulan los diccionarios
            book.save(filename=resource)  # guardo el libro de excel
        else:  # si no tiene cabecera hago lo mismo que arriba pero sin la iteracion de cabecera
            for index in range(len(collection_data)):
                for row in collection_data[index].values():
                    values.append(row)
                sheet.append(values)
                values = []
            book.save(filename=resource)
        book.close()

    # FUNCIONES ENCRIPTS################################################################################################
    @staticmethod
    def get_enviroment_key_from_file():

        """
            Description:
                Obtiene la key (bytes) de la variable de entorno "PYBOT_KEY".
            Returns:
                Devuelve la key en bytes.
        """

        key = None
        enviroment_key = os.getenv(ENVIRONMENT_VAR)
        if enviroment_key is not None:
            try:
                with open(enviroment_key, 'rb') as file:
                    key = file.read()
            except FileNotFoundError:
                print(f"No existe el archivo '{enviroment_key}'")
        else:
            print(f"No se encuentra cargada correctamente la variable de entorno f{ENVIRONMENT_VAR}")
        return key

    def get_password_from_file_encrypted(self, enviroment, user):

        """
            Description:
                Busca una password en un archivo encriptado.
            Args:
                enviroment: Nombre del ambiente asociado al usuario del cual se pretende recuperar la password.
                user: Nombre del usuario del que se pretende recuperar la password.
            Returns:
                Devuelve la password del usuario.
        """

        password = None
        key = Functions.get_enviroment_key_from_file()
        fe = Fernet(key)

        with open(PATH_ORIGIN, 'rb') as file:
            encrypte_data = file.read()

        decrypted_data = fe.decrypt(encrypte_data)
        pass_list = decrypted_data.decode('utf-8').split('\r')
        if 'AMBIENTE;IP;BASE;USUARIO;PASS' in pass_list:
            for row in pass_list:
                list_data_row = row.split(';')
                if enviroment and user in list_data_row:
                    password = list_data_row[-1]
                    break

            if password is None:
                unittest.TestCase().skipTest(f"--PasswordNotFound-- No se encontro la password de acceso en el archivo"
                                             f" para {enviroment};{user}")
            else:
                return password

    def get_data_from_xml_encrypted(self, father_attribute, attribute_to_search, attribute_name, inner_search):
        """
            Description:
                Busca y retorna la información requerida por el usuario desde el xml encriptado
            Args:
                father_attribute: Nombre del Tag padre.
                attribute_to_search: Tipo de atributo que desea buscar.
                attribute_name: Nombre del atributo que desea buscar.
                inner_search: Nombre del tag interno del que se desea obtener el texto.
            Returns:
                Retorna el texto interno del dato requerido.
        """
        key = Functions.get_enviroment_key_from_file()
        fe = Fernet(key)
        return_data = None
        try:
            with open(PATH_ORIGIN_XML, 'rb') as file:
                data = file.read()
            deencrypted_data = fe.decrypt(data)
            decompressed_data = bz2.decompress(deencrypted_data)
            file.close()
            read_xml_file = Et.fromstring(decompressed_data)
            # el siguiente for revisa utilizando un formato XPATH los datos requeridos por el usuario
            # y lo retorna si este existe
            for element in read_xml_file.findall(f"./{father_attribute}[@{attribute_to_search}='{attribute_name}']/"):
                if element.tag == inner_search and (element.text is not None or element.text != ""
                                                    or element.text != " "):
                    return_data = element.text
        except:
            raise "Ha Ocurrido un Error en el Tiempo de Ejecución -> ERROR CODE 1523 (Functions)"

        return return_data

    @staticmethod
    def use_xml_connect_to_db(ip_db, db_user_name):
        """
            Description:
                Busca y retorna la contraseña de la db requerida desde el xml encriptado
            Args:
                ip_db: IP servidor a conectar.
                db_user_name: Nombre de usuario de la DB.
            Returns:
                Retorna la contraseña de la db.
        """
        key = Functions.get_enviroment_key_from_file()
        fe = Fernet(key)
        return_db_password = None
        try:
            with open(PATH_ORIGIN_XML, 'rb') as file:
                data = file.read()
            deencrypted_data = fe.decrypt(data)
            decompressed_data = bz2.decompress(deencrypted_data)
            file.close()
            read_xml_file = Et.fromstring(decompressed_data)
            # el siguiente for revisa utilizando un formato XPATH los datos requeridos por el usuario
            # y lo retorna si este existe
            elements_search = read_xml_file.findall(f"./CLAVES/IP[.='{ip_db}']/../USER[.='{db_user_name}']/../PASS")
            return_db_password = elements_search[0].text
        except:
            unittest.TestCase.skipTest(Functions, "Error al intentar obtener información del archivo XML")
        return return_db_password

    @staticmethod
    def obtain_port_from_xml(ip_db, db_user_name):
        """
            Description:
                Busca y retorna la contraseña de la db requerida desde el xml encriptado
            Args:
                ip_db: IP servidor a conectar.
                db_user_name: Nombre de usuario de la DB.
            Returns:
                Retorna la contraseña de la db.
        """
        key = Functions.get_enviroment_key_from_file()
        fe = Fernet(key)
        return_db_port = None
        try:
            with open(PATH_ORIGIN_XML, 'rb') as file:
                data = file.read()
            deencrypted_data = fe.decrypt(data)
            decompressed_data = bz2.decompress(deencrypted_data)
            file.close()
            read_xml_file = Et.fromstring(decompressed_data)
            # el siguiente for revisa utilizando un formato XPATH los datos requeridos por el usuario
            # y lo retorna si este existe
            elements_search = read_xml_file.findall(f"./CLAVES/IP[.='{ip_db}']/../USER[.='{db_user_name}']/../PORT")
            return_db_port = elements_search[0].text
        except:
            unittest.TestCase.skipTest(Functions, "Error al intentar obtener información del archivo XML")
        return return_db_port

    # FUNCIONES NOTIFICACIONES #########################################################################################

    def send_mail(self, receiver_email: list, title, content, file_attach=None):

        """
            Description:
                Envia un informe vía email.
            Args:
                receiver_email (str): Lista de destinatarios de correos.
                title (str): Asunto del correo.
                content (str): Cuerpo del correo
                file_attach (file): Archivos adjuntos del correo.
            Returns:
                True: Si el correo fue enviado con éxito.
                False: Si el correo no pudo ser enviado.
        """
        content = f'{content}{Functions.footer_signature_html()}'
        content = Functions.apply_style_css_to_block(content)
        port = Functions.get_data_from_xml_encrypted(self, "CLAVES", "id", "Email Sender Info", "PORT")
        smtp_server = Functions.get_data_from_xml_encrypted(self, "CLAVES", "id", "Email Sender Info", "IP")
        password = Functions.get_data_from_xml_encrypted(self, "CLAVES", "id", "Email de Pybot", "PASS")
        sender_email = Functions.get_data_from_xml_encrypted(self, "CLAVES", "id", "Email de Pybot", "USER")
        message = MIMEMultipart("alternative")
        message['To'] = ",".join(receiver_email)
        message['Subject'] = 'No-responder: ' + title
        message.attach(MIMEText(content, 'html'))
        if file_attach is not None:
            attachment = open(file_attach, "rb")
            p = MIMEBase('application', 'octet-stream')
            p.set_payload(attachment.read())
            encoders.encode_base64(p)
            file_name = file_attach.split('\\')
            p.add_header('Content-Disposition', "attachment; filename= %s" % file_name[-1])
            message.attach(p)
            # img_data = open(file_attach, 'rb').read()
            # image = MIMEImage(img_data, name=os.path.basename(file_attach))
            # message.attach(image)
        try:
            with smtplib.SMTP(smtp_server, port) as server:
                server.ehlo()  # Can be omitted
                if Parameters.environment == "Windows":
                    server.starttls()
                    server.login(sender_email, password)
                text = message.as_string()
                server.sendmail(sender_email, receiver_email, text)
                return True
        except Exception as e:
            print(f'El mail no pudo ser enviado. // exception: {e}')
            Functions.exception_logger(e)
            return False
        finally:
            server.close()

    def set_resources(self, file_name=None, specific_sheet=None, test_cases_name_list=None):
        """
        Description:
            Genera un pool de datos realizando una lectura completa del archivo excel,
            identificado en el proyecto. Esta etapa se realiza en el SetUp,
            y entrega los datos correspondientes al caso ejecutado.
        Args:
            file_name = Nombre del archivo excel.
            specific_sheet = Hoja específica de trabajo.
            test_cases_name_list = Nombre del caso de prueba.
        """
        try:
            if file_name is None:
                print(f"Se leerá el resource con nombre: '{file_name}.xlsx'")
                file_name = Functions.file_name
                resource = os.path.join(Functions.path_resources, f"{file_name}.xlsx")

                # Verificar si el archivo existe
                if not os.path.isfile(resource):
                    print("El archivo Excel no existe. Se creará uno nuevo.")
                    try:
                        # Crear un nuevo excel
                        book = openpyxl.Workbook()
                        # Crear una nueva hoja llamada "DataTest"
                        sheet = book.create_sheet(title="DataTest", index=0)
                        # Insertar encabezado "DATO" en la primera columna
                        sheet.cell(row=1, column=1, value="SIN_DATOS")
                        # Guardar el excel con el nombre del test
                        book.save(resource)
                        print(f"Archivo Excel '{file_name}.xlsx' creado exitosamente con hoja 'DataTest'.")
                    except Exception as e:
                        print(f"Error al crear el archivo Excel: {e}")
                        raise
                else:
                    # El archivo ya existe, cargarlo normalmente
                    book = openpyxl.load_workbook(resource, data_only=True)
            else:
                resource = os.path.join(Functions.path_resources, f"{file_name}.xlsx")
                book = openpyxl.load_workbook(resource, data_only=True)

            if test_cases_name_list is None:
                test_cases_name_list = unittest.getTestCaseNames(self.__class__, "test")
            if specific_sheet is None:
                print("Utilizando hoja default 'DataTest'")
                sheet = book["DataTest"]
            else:
                print(f"Utilizando la hoja: '{specific_sheet}'")
                sheet = book[specific_sheet]

            records = list(sheet.values)  # pool de data completo
            headers = list(records.pop(0))  # solo datos del header
            long_header = len(headers)
            for header_value in range(0, long_header):
                if headers[header_value] is None:
                    none_cell_letter = get_column_letter(header_value + 1)
                    none_cell_data = sheet[f'{none_cell_letter}{1}'].value
                    # Se valida si el header está completo, de lo contrario avisa al usuario
                    print(
                        f"Error: Existe un campo vacío en el header. {none_cell_letter}1 y con value: {none_cell_data} "
                        f"Acción: SkipTest del proyecto  ->{Functions.project_name} y Archivo -> "
                        f"{Functions.file_name}.xlsx")
                    unittest.TestCase().skipTest(f"--DataResourceNullHeaderError-- Existe un Header vacio en la Pos -> "
                                                 f"{none_cell_letter}1 y con value: {none_cell_data} en el proyecto->"
                                                 f"{Functions.project_name} y Archivo "
                                                 f"-> {Functions.file_name}.xlsx")

            # validando headers duplicados
            for pos_header in range(0, long_header - 1):
                header_list_checker = headers.copy()
                long_list_copy = len(header_list_checker)
                # Utilizando la lista de paso, se extrae el value para ser comparado
                check_header = header_list_checker.pop(pos_header)
                for pos_in_list in range(0, long_list_copy - 1):
                    # Se utiliza el largo de la lista de paso ya que ahora posee 1 campo menos
                    # Se normaliza el tipo de texto para evitar problemas al comparar
                    if check_header.upper() == header_list_checker[pos_in_list].upper():
                        # Se obtiene la letra de la columna duplicada utilizando un método de la libreria
                        # Y pasando como parámetros el espacio de la lista original.
                        cell_letter = get_column_letter(headers.index(header_list_checker[pos_in_list]) + 1)
                        cell_data = sheet[f'{cell_letter}{1}'].value
                        print(
                            f"Error: Existe un campo duplicado en el header. Pos {cell_letter}1 y con value: {cell_data} "
                            f"Acción: SkipTest del proyecto ->{Functions.project_name} y Archivo -> "
                            f"{Functions.file_name}.xlsx")
                        unittest.TestCase().skipTest(f"--DataResourceDuplicateKeyError-- "
                                                     f"Existe un campo duplicado en el Header -> "
                                                     f"Pos {cell_letter}1 y con value: {cell_data} en el proyecto->"
                                                     f"{Functions.project_name} y Archivo -> "
                                                     f"{Functions.file_name}.xlsx")
            for upper_function in range(long_header):
                headers[upper_function] = headers[upper_function].upper()
            # unificación de headers y data recolectado
            data_global_resource = [dict(zip(headers, row)) for row in records]

            try:
                # unificación de nombres test cases y data_global_resource
                if not Parameters.manual_increment:
                    Functions.data_resource = dict(zip(test_cases_name_list, data_global_resource))
                else:
                    Functions.data_resource = data_global_resource[int(Parameters.row) - 1]
                    print(f"El resource correra con el caso: {Functions.test_case_name}")
                    pprint.pprint(Functions.data_resource)
                    print("============================================")
                    return None
            except KeyError as key_error:
                print(f"Error: A cada row del archivo {Functions.file_name}.xlsx le corresponde un caso")
                # No detenemos el caso de prueba, solo mostramos una advertencia
                print(f"ADVERTENCIA: El archivo de recursos no tiene la estructura esperada.")

            # se pasa como parámetro el nombre del test case ejecutado
            try:
                Functions.data_resource = Functions.data_resource[Functions.test_case_name]
            except KeyError as key_error:
                print(f"ERROR: A cada row del archivo {Functions.file_name}.xlsx le corresponde a un caso")
                # No detenemos el caso de prueba, solo mostramos una advertencia
                print(f"ADVERTENCIA: No tiene datos cargados en el resource.")
        except Exception as ex:
            print(f"ERROR: Ocurrió un error durante la configuración de los recursos. {ex}")
            # No detenemos el caso de prueba, solo mostramos una advertencia
        print(f"El resource correra con el caso: '{Functions.test_case_name}'")
        print("==================Inicio Resource==================")
        pprint.pprint(Functions.data_resource)
        print("===================Fin Resource====================")

    def get_random(self, min_range, max_range):

        """
            Description:
                Obtiene un número aleatorio del rango especificado.
            Args:
                min_range (int): Rango mínimo.
                max_range (int): Rango máximo.
            Returns:
                Retorna un número aleatorio.
        """

        random_number = random.randint(min_range, max_range)
        return random_number

    def create_teams_notifications(self, teams_channel=None, table_color=None, msg_tittle=None,
                                   section_text=None, btn_name=None, btn_link=None):

        """
            Description:
                Realiza la notificación a teams.
            Args:
                teams_channel: Canal de teams donde se producirá la notificación.
                table_color: Color de la tabla.
                msg_tittle: Titulo del mensaje.
                section_text: Contenido del mensaje.
                btn_name: Nombre de un boton.
                btn_link: Link asociado a un boton.
        """

        if table_color is None:
            table_color = "F03A2E"
        if msg_tittle is None:
            msg_tittle = "Notificación creada automaticamente"
        if section_text is None:
            section_text = "Texto Sección 1"
        if btn_name is None:
            btn_name = "Hazme Click!"

        Functions.teams = pymsteams.connectorcard(teams_channel)
        Functions.teams.color(table_color)  # Color de la tarjeta.
        Functions.teams.title(msg_tittle)  # Titulo del mensaje.
        Functions.teams.text(" ")  # Texto al mensaje.
        my_message_section = pymsteams.cardsection()  # Agrega una sección a la tarjeta.
        # my_message_section.title(section_tittle)  # Titulo de la nueva seccion.
        my_message_section.text(section_text)  # Texto de la sección.
        my_message_section.linkButton(btn_name, btn_link)  # Se agrega un botón a la sección.
        Functions.teams.addSection(my_message_section)  # Agregar sección a la tarjeta.
        # self.teams.printme()  # Payload del mensaje.

        if teams_channel is None or btn_link is None:
            print(f"\nNo puede enviarse la notificación ya que alguno de los datos requeridos es inválido")
            print(f"\nChannel: {teams_channel}")
            print(f"\nButton Link: {btn_link}")
        else:
            Functions.teams.send()  # Envia el mensaje.

    @staticmethod
    def exception_logger(exception):

        """
            Description:
                Realiza la la impresión de una excepción por conosola.
            Args:
                exception: Excepción producida durante el tiempo de ejecución.
        """

        if Parameters.loggin_exceptions:
            print(exception)

    @staticmethod
    def set_exception_loggin(value: bool):

        """

        Description:
            Configura el logeo de las excepciones

        Args:
            value: true o false

        """

        Parameters.loggin_exceptions = value

    @staticmethod
    def color_message(color, message):
        """
            Description: Colorea el string del color indicado de entre una lista de colores.

            Args:
                color: puede ser de color red, blue, yellow o green.
                message: string a colorear.

            Returns:
                string coloreado o string por default

        """
        if Parameters.environment_configuration is None:
            if color.upper() == "RED":
                return f"{RED}{message}{DEFAULT}"
            elif color.upper() == "BLUE":
                return f"{BLUE}{message}{DEFAULT}"
            elif color.upper() == "YELLOW":
                return f"{YELLOW}{message}{DEFAULT}"
            elif color.upper() == "GREEN":
                return f"{GREEN}{message}{DEFAULT}"
            else:
                return f"{DEFAULT}{message}{DEFAULT}"
        else:
            if Parameters.environment_configuration.lower() == 'server':
                if color.upper() == "RED":
                    return f"{RED}{message}{DEFAULT}"
                elif color.upper() == "BLUE":
                    return f"{BLUE}{message}{DEFAULT}"
                elif color.upper() == "YELLOW":
                    return f"{YELLOW}{message}{DEFAULT}"
                elif color.upper() == "GREEN":
                    return f"{GREEN}{message}{DEFAULT}"
                else:
                    return f"{DEFAULT}{message}{DEFAULT}"
            else:
                return message

    ####################################################################################################################
    ############################################## Encryptor ###########################################################
    ####################################################################################################################

    class Encryptor:
        ENVIRONMENT_PROJ_FILE_PATH = os.path.join(os.path.abspath(os.path.join(os.getcwd(), "../..")),
                                                  "environment_access.xml")
        ENVIRONMENT_BACKUP_FILE_NAME = os.path.join(os.path.abspath(os.path.join(os.getcwd(), "../../../..")),
                                                    "environment_access.xml")

        def __init__(self, father_attribute, attribute_to_search, data_find, data_inner_key, xml_en_consola=False):
            self.father_attribute = father_attribute
            self.atribute_to_search = attribute_to_search
            self.data_find = data_find
            self.data_inner_key = data_inner_key
            self.key = self.get_enviroment_key_from_file()
            self.encriptor = Fernet(self.key)
            self.xml_en_consola = xml_en_consola

        def main(self):
            data_returned = None
            # Proceso:
            # 1 Si el archivo del proyecto existe
            if self.project_file_exists() is True:
                # print("Archivo del proyecto encontrado!")
                data_returned = self.get_data_from_proj_file(self.father_attribute,
                                                             self.atribute_to_search,
                                                             self.data_find,
                                                             self.data_inner_key)
                # Si la información obtenida del archivo del proyecto no existe
                if data_returned is None:
                    print(f"Parace que el archivo no contiene la información buscada: {self.data_inner_key}")
                    print(f"Buscando {self.data_inner_key} ahora en el archivo BackUp")
                    # reviso que el archivo backup exista
                    if self.backup_file_exists() is True:
                        # print("Archivo BackUp encontrado!")
                        # obtengo el dato si existe, desde el archivo backup
                        data_returned = self.get_data_from_backup_file(self.father_attribute,
                                                                       self.atribute_to_search,
                                                                       self.data_find,
                                                                       self.data_inner_key)
                        # Si el dato existe
                        if data_returned is not None:
                            # print("Se a encontrado el dato buscado, pero este se encuentra en el archivo BackUp")
                            print(f"Copiando [{self.father_attribute};{self.atribute_to_search}] al archivo del "
                                  f"proyecto actual")
                            # Agrego la información faltante desde el archivo backup, al archivo del proyecto
                            self.add_data_to_proj_file(self.father_attribute, self.atribute_to_search, self.data_find)
                            print("La información fué agregada al archivo del proyecto actual correctamente!")
                        # Si el dato NO existe
                        else:
                            unittest.case.TestCase.skipTest(self,
                                                            "Parace que ninguno de los archivos contiene la información buscada")
                    else:
                        # Si el archivo backup no existe, crea uno en base a un template vacio
                        unittest.case.TestCase.skipTest(self,
                                                        "Archivo BackUp no encontrado!, Creando archivo BackUp con template base para futuras ejecuciones")
                        self.create_backup_file()
                        print("Archivo BackUp creado con éxito!")
            else:
                print("Archivo del proyecto no encontrado!")
                # 2 - Si el archivo del proyecto NO existe, reviso que el archivo backup exista
                if self.backup_file_exists() is True:
                    # print("Archivo BackUp encontrado!")
                    # Si el archivo backup existe, crea un template en la ubicación del proyecto
                    # print("Creando archivo de proyecto actual con template base para futuras ejecuciones")
                    self.create_project_file()
                    # print("Archivo del proyecto creado con éxito!")
                    self.add_data_to_proj_file(self.father_attribute, self.atribute_to_search, self.data_find)
                    print("Data transferida al archivo del proyecto actual")
                    # Corro el proceso otra vez
                    data_returned = Functions.Encryptor(self.father_attribute, self.atribute_to_search, self.data_find,
                                                        self.data_inner_key).main()
                # Si no existe el archivo del proyecto y tampoco existe el archivo backup
                else:
                    unittest.case.TestCase.skipTest(self,
                                                    "No se puede continuar sin la presencia de al menos uno de los archivos environment_access.xml")

            return data_returned

        def project_file_exists(self):
            project_exists = False
            if os.path.exists(self.ENVIRONMENT_PROJ_FILE_PATH) is True:
                project_exists = True
            return project_exists

        def backup_file_exists(self):
            ignored_exists = False
            # Si existe el archivo ignorado ubicado en testing-Automation devuelve True, sino existe False
            if os.path.exists(self.ENVIRONMENT_BACKUP_FILE_NAME):
                ignored_exists = True
            return ignored_exists

        ####################################################################################################################
        #                                           ENCRYPTOR FUNCTIONS                                                    #
        ####################################################################################################################

        def get_enviroment_key_from_file(self):

            """

                Description:
                    Obtiene la key (bytes) de la variable de entorno "PYBOT_KEY".

                Returns:
                    Devuelve la key en bytes.

            """

            key = ""
            enviroment_key = os.getenv(ENVIRONMENT_VAR)
            if enviroment_key is not None:
                try:
                    with open(enviroment_key, 'rb') as file:
                        key = file.read()
                except FileNotFoundError:
                    print(f"No existe el archivo '{enviroment_key}'")
            else:
                print(f"No se encuentra cargada correctamente la variable de entorno f{ENVIRONMENT_VAR}")
            return key

        # Decript
        def decompress_and_deencrypt_xml(self, read_file):
            if self.is_file_encrypted(read_file) is True:
                try:
                    with open(read_file, 'rb') as file:
                        data = file.read()
                    deencrypted_data = self.encriptor.decrypt(data)
                    decompressed_data = bz2.decompress(deencrypted_data)
                    file.close()
                    os.remove(read_file)
                    with open(read_file, 'wb') as output_file:
                        output_file.write(decompressed_data)
                except FileNotFoundError:
                    print("El archivo buscado no existe en el directorio especificado.")
            else:
                print(f"No se puede desencryptar el archivo, ya que este es su estado actual -> {read_file}")

        # Encript
        def compress_and_encrypt_xml(self, read_file):
            if self.is_file_encrypted(read_file) is not True:
                try:
                    with open(read_file, 'rb') as file:
                        data = file.read()
                    compressed_data = bz2.compress(data)
                    encrypted_data = self.encriptor.encrypt(compressed_data)
                    file.close()
                    os.remove(read_file)
                    with open(read_file, 'wb') as output_file:
                        output_file.write(encrypted_data)
                except FileNotFoundError:
                    print("El archivo buscado no existe en el directorio especificado.")
            else:
                print(f"No se puede encryptar el archivo, ya que este es su estado actual -> {read_file}")

        # obtener data requerida desde el archivo
        def get_data_from_proj_file(self, father_attribute, atribute_to_search, dato_a_buscar, inner_search):
            return_data = None
            try:
                read_xml_file, project_tree = self.get_xml_root(self.ENVIRONMENT_PROJ_FILE_PATH)
                for element in read_xml_file.findall(f"./{father_attribute}[@{atribute_to_search}='{dato_a_buscar}']/"):
                    if element.tag == inner_search and (element.text is not None or element.text != ""
                                                        or element.text != " "):
                        return_data = element.text
            except Exception:
                print("Ha Ocurrido un Error en el Tiempo de Ejecución -> ERROR CODE 204 (Encriptor)")

            return return_data

        def get_data_from_backup_file(self, father_attribute, atribute_to_search, dato_a_buscar, inner_search):
            return_data = None
            try:
                read_xml_file, backup_tree = self.get_xml_root(self.ENVIRONMENT_BACKUP_FILE_NAME)
                for element in read_xml_file.findall(f"./{father_attribute}[@{atribute_to_search}='{dato_a_buscar}']/"):
                    if element.tag == inner_search and (element.text is not None or element.text != ""
                                                        or element.text != " "):
                        return_data = element.text
            except Exception:
                raise "Ha Ocurrido un Error en el Tiempo de Ejecución -> ERROR CODE 204 (Encriptor)"

            return return_data

        # Actualizar data
        def add_data_to_proj_file(self, father_attribute, atribute_to_search, dato_a_buscar):
            # Necesito:
            #   *   Ahora que sé que el dato existe en el archivo backup, obtenerlo con su bloque de información.
            #   *   VALIDAR MEDIANTE TAG SI ESTE ELEMENTO EXISTE EN EL ARCHIVO
            backup_xml, backup_tree = self.get_xml_root(self.ENVIRONMENT_BACKUP_FILE_NAME)
            project_xml, project_tree = self.get_xml_root(self.ENVIRONMENT_PROJ_FILE_PATH)
            self.validate(project_xml)
            # Create a new CLAVES element
            new_claves = Et.Element('CLAVES')
            new_claves.set(atribute_to_search, dato_a_buscar)
            # Create sub-elements for the new CLAVES
            port = Et.SubElement(new_claves, 'PORT')
            ip = Et.SubElement(new_claves, 'IP')
            environment = Et.SubElement(new_claves, 'ENVIRONMENT')
            base = Et.SubElement(new_claves, 'BASE')
            user = Et.SubElement(new_claves, 'USER')
            password = Et.SubElement(new_claves, 'PASS')
            # Append the new CLAVES element to the root
            project_xml.append(new_claves)
            for element in backup_xml.findall(f"./{father_attribute}[@{atribute_to_search}='{dato_a_buscar}']/"):
                if element.tag == "PORT":
                    port.text = element.text
                if element.tag == "IP":
                    ip.text = element.text
                if element.tag == "ENVIRONMENT":
                    environment.text = element.text
                if element.tag == "BASE":
                    base.text = element.text
                if element.tag == "USER":
                    user.text = element.text
                if element.tag == "PASS":
                    password.text = element.text
            Et.indent(project_xml, "   ")
            project_tree.write(self.ENVIRONMENT_PROJ_FILE_PATH)
            self.compress_and_encrypt_xml(self.ENVIRONMENT_PROJ_FILE_PATH)
            if self.xml_en_consola is True:
                Et.dump(project_xml)
                print("---------------------")
                Et.dump(backup_xml)

        def create_backup_file(self):
            xml_body = b'<?xml version="1.0" encoding="UTF-8" ?> ' \
                       b'<root>' \
                       b'<!--<CLAVES id="Test_Id">-->' \
                       b'<!--<PORT>9999</PORT>-->' \
                       b'<!--<IP>255.255.255.255</IP>-->' \
                       b'<!--<ENVIRONMENT>Test</ENVIRONMENT>-->' \
                       b'<!--<BASE>SQL_BASE</BASE>-->' \
                       b'<!--<USER>Test_User</USER>-->' \
                       b'<!--<PASS>Test_Password</PASS>-->' \
                       b'<!--</CLAVES>-->' \
                       b'</root>'
            # Formateo del archivo output
            compressed_data = bz2.compress(xml_body)
            encripted_xml = self.encriptor.encrypt(compressed_data)
            with open(self.ENVIRONMENT_BACKUP_FILE_NAME, 'wb') as output_file:
                output_file.write(encripted_xml)
            output_file.close()

        def create_project_file(self):
            with open(rf"{self.ENVIRONMENT_PROJ_FILE_PATH}", "wb") as env_file:
                env_file.write(b'<root>\n')
                env_file.write(b'</root>\n')
            tree = Et.parse(self.ENVIRONMENT_PROJ_FILE_PATH)
            root = tree.getroot()
            Et.indent(root, "   ")
            tree.write(self.ENVIRONMENT_PROJ_FILE_PATH)
            self.compress_and_encrypt_xml(self.ENVIRONMENT_PROJ_FILE_PATH)

        def get_xml_root(self, file_loc):
            self.decompress_and_deencrypt_xml(file_loc)
            tree = Et.parse(file_loc)
            current_root = tree.getroot()
            self.compress_and_encrypt_xml(file_loc)
            return current_root, tree

        def is_file_encrypted(self, file_path):
            try:
                # Intentar analizar el archivo XML
                Et.parse(file_path)
                return False
            except Et.ParseError:
                # Si se produce un error al analizar, se considera que el archivo está encriptado
                return True

        def validate(self, root_tree):
            elemento_prueba = root_tree.find(f'.//*[@{self.atribute_to_search}="{self.data_find}"]')
            if elemento_prueba is not None:
                root_tree.remove(elemento_prueba)

    class LoggingFeature():
        def __init__(self, mensaje, enable_logging=False):
            self.enable_logging = enable_logging
            self.mensaje = mensaje
            # Inicialización y configuración del logger
            self.logger = logging.getLogger(__name__)
            logging.basicConfig(level="WARNING", encoding='utf-8')
            # Se desactivan los warnings sobre reintentos
            urllib3.disable_warnings(urllib3.exceptions.InsecureRequestWarning)
            logging.getLogger("urllib3").setLevel(logging.ERROR)

        def send_log(self):
            if self.enable_logging is True:
                self.logger.warning(self.mensaje)
            else:
                print(self.mensaje)
